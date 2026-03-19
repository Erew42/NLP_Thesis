from __future__ import annotations

from collections import Counter
import datetime as dt
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable

import polars as pl
from xlsxwriter.exceptions import FileCreateError

from thesis_pkg.core.ccm.transforms import common_stock_pass_expr
from thesis_pkg.io.excel import (
    write_refinitiv_bridge_workbook,
    write_refinitiv_null_ric_diagnostics_workbook,
    write_refinitiv_ownership_universe_workbook,
    write_refinitiv_ownership_validation_workbook,
    write_refinitiv_ownership_smoke_testing_workbook,
    write_refinitiv_resolution_diagnostic_workbook,
    write_refinitiv_ric_lookup_extended_workbook,
    write_refinitiv_ric_lookup_workbook,
)


REQUIRED_DAILY_COLUMNS: tuple[str, ...] = (
    "KYPERMNO",
    "CALDT",
    "KYGVKEY_final",
    "LIID",
    "CIK_final",
    "CUSIP",
    "ISIN",
    "TICKER",
)

OPTIONAL_DAILY_COLUMNS: tuple[str, ...] = (
    "LINKTYPE",
    "LINKPRIM",
    "link_quality_flag",
    "HEXCNTRY",
    "n_filings",
)

BRIDGE_SOURCE_COLUMNS: tuple[str, ...] = (
    "bridge_row_id",
    "KYPERMNO",
    "KYGVKEY_final",
    "LIID",
    "CIK_final",
    "CUSIP",
    "ISIN",
    "TICKER",
    "LINKTYPE",
    "LINKPRIM",
    "link_quality_flag",
    "HEXCNTRY",
    "company_name",
    "first_seen_caldt",
    "last_seen_caldt",
    "n_daily_rows",
    "n_filing_days",
    "n_filings_total",
)

BRIDGE_VENDOR_COLUMNS: tuple[str, ...] = (
    "vendor_match_status",
    "vendor_primary_id_type",
    "vendor_primary_id",
    "vendor_primary_ric",
    "vendor_ds_mnemonic",
    "vendor_permid",
    "vendor_returned_name",
    "vendor_returned_cusip",
    "vendor_returned_isin",
    "vendor_candidate_ids_raw",
    "vendor_notes",
)

BRIDGE_OUTPUT_COLUMNS: tuple[str, ...] = BRIDGE_SOURCE_COLUMNS + BRIDGE_VENDOR_COLUMNS

WORKBOOK_TEXT_COLUMNS: tuple[str, ...] = (
    "bridge_row_id",
    "KYPERMNO",
    "KYGVKEY_final",
    "LIID",
    "CIK_final",
    "CUSIP",
    "ISIN",
    "TICKER",
    "LINKTYPE",
    "LINKPRIM",
    "link_quality_flag",
    "HEXCNTRY",
    "company_name",
    "first_seen_caldt",
    "last_seen_caldt",
) + BRIDGE_VENDOR_COLUMNS

RIC_LOOKUP_COLUMNS: tuple[str, ...] = (
    "bridge_row_id",
    "KYPERMNO",
    "CUSIP",
    "ISIN",
    "TICKER",
    "first_seen_caldt",
    "last_seen_caldt",
    "preferred_lookup_id",
    "preferred_lookup_type",
    "vendor_primary_ric",
    "vendor_returned_name",
    "vendor_returned_cusip",
    "vendor_returned_isin",
    "vendor_match_status",
    "vendor_notes",
)

RIC_LOOKUP_TEXT_COLUMNS: tuple[str, ...] = RIC_LOOKUP_COLUMNS


@dataclass(frozen=True)
class RicLookupFilterProfile:
    name: str
    predicates: tuple[str, ...]
    require_common_stock: bool = False
    require_gvkey: bool = False


RIC_LOOKUP_FILTER_PROFILES: tuple[RicLookupFilterProfile, ...] = (
    RicLookupFilterProfile(
        name="common_stock",
        predicates=(
            "KYPERMNO is not null",
            "common stock (SHRCD in {10, 11})",
        ),
        require_common_stock=True,
    ),
    RicLookupFilterProfile(
        name="common_stock_with_gvkey",
        predicates=(
            "KYPERMNO is not null",
            "common stock (SHRCD in {10, 11})",
            "KYGVKEY_final is not null",
        ),
        require_common_stock=True,
        require_gvkey=True,
    ),
)

LOOKUP_IDENTIFIER_TYPES: tuple[str, ...] = ("ISIN", "CUSIP", "TICKER")
LOOKUP_IDENTIFIER_PAIRS: tuple[tuple[str, str], ...] = (
    ("ISIN", "CUSIP"),
    ("ISIN", "TICKER"),
    ("CUSIP", "TICKER"),
)

RIC_LOOKUP_EXTENDED_BASE_COLUMNS: tuple[str, ...] = (
    "bridge_row_id",
    "KYPERMNO",
    "CUSIP",
    "ISIN",
    "TICKER",
    "first_seen_caldt",
    "last_seen_caldt",
    "vendor_primary_ric",
    "vendor_returned_name",
    "vendor_returned_cusip",
    "vendor_returned_isin",
    "vendor_match_status",
    "vendor_notes",
)


def _extended_lookup_columns() -> tuple[str, ...]:
    columns = list(RIC_LOOKUP_EXTENDED_BASE_COLUMNS)
    for identifier_type in LOOKUP_IDENTIFIER_TYPES:
        columns.extend(
            (
                f"{identifier_type}_lookup_input",
                f"{identifier_type}_attempted",
                f"{identifier_type}_returned_ric",
                f"{identifier_type}_returned_name",
                f"{identifier_type}_returned_isin",
                f"{identifier_type}_returned_cusip",
                f"{identifier_type}_success",
            )
        )
    for left_type, right_type in LOOKUP_IDENTIFIER_PAIRS:
        columns.extend(
            (
                f"{left_type}_vs_{right_type}_same_ric",
                f"{left_type}_vs_{right_type}_same_isin",
                f"{left_type}_vs_{right_type}_same_cusip",
            )
        )
    columns.append("all_successful_attempts_consistent")
    return tuple(columns)


RIC_LOOKUP_EXTENDED_COLUMNS: tuple[str, ...] = _extended_lookup_columns()
RIC_LOOKUP_EXTENDED_TEXT_COLUMNS: tuple[str, ...] = tuple(
    name
    for name in RIC_LOOKUP_EXTENDED_COLUMNS
    if name not in {"all_successful_attempts_consistent"}
    and not name.endswith("_attempted")
    and not name.endswith("_success")
    and "_same_" not in name
)

RIC_LOOKUP_EXTENDED_SUMMARY_COLUMNS: tuple[str, ...] = (
    "summary_category",
    "summary_key",
    "value",
)
RIC_LOOKUP_EXTENDED_SUMMARY_TEXT_COLUMNS: tuple[str, ...] = (
    "summary_category",
    "summary_key",
)

LOOKUP_FAILURE_MARKERS: frozenset[str] = frozenset({"NULL", "NO UNIVERSE DEFINED."})
RIC_LOOKUP_EXTENDED_DATE_COLUMNS: tuple[str, ...] = ("first_seen_caldt", "last_seen_caldt")
RIC_LOOKUP_EXTENDED_BOOLEAN_COLUMNS: tuple[str, ...] = tuple(
    name
    for name in RIC_LOOKUP_EXTENDED_COLUMNS
    if name.endswith("_attempted")
    or name.endswith("_success")
    or "_same_" in name
    or name == "all_successful_attempts_consistent"
)
RIC_LOOKUP_RESOLUTION_COLUMNS: tuple[str, ...] = (
    "accepted_identity_returned_isin",
    "accepted_identity_returned_cusip",
    "accepted_ric",
    "accepted_ric_source",
    "accepted_resolution_status",
    "conventional_identity_conflict",
    "ticker_candidate_ric",
    "ticker_candidate_available",
    "ticker_candidate_conflicts_with_conventional",
    "extended_ric",
    "extended_from_bridge_row_id",
    "extended_from_span_start",
    "extended_from_span_end",
    "extension_direction",
    "extension_status",
    "effective_collection_ric",
    "effective_collection_ric_source",
    "effective_resolution_status",
)
RIC_LOOKUP_RESOLUTION_OUTPUT_COLUMNS: tuple[str, ...] = (
    *RIC_LOOKUP_EXTENDED_COLUMNS,
    *RIC_LOOKUP_RESOLUTION_COLUMNS,
)

RESOLUTION_DIAGNOSTIC_TARGET_COLUMNS: tuple[str, ...] = (
    *RIC_LOOKUP_RESOLUTION_OUTPUT_COLUMNS,
    "LIID_inferred",
    "logical_bridge_group_id",
    "diagnostic_case_id",
    "case_target_bridge_row_id",
    "target_class",
    "group_sequence_index",
    "group_row_count",
    "case_previous_row_available",
    "case_next_row_available",
    "case_any_adjacent_row_available",
    "case_previous_bridge_row_id",
    "case_next_bridge_row_id",
    "case_previous_effective_available",
    "case_next_effective_available",
    "case_any_adjacent_effective_available",
    "case_both_adjacent_effective_available",
    "case_previous_effective_collection_ric",
    "case_next_effective_collection_ric",
    "case_previous_effective_collection_ric_source",
    "case_next_effective_collection_ric_source",
    "case_previous_accepted_identity_returned_isin",
    "case_next_accepted_identity_returned_isin",
    "case_previous_accepted_identity_returned_cusip",
    "case_next_accepted_identity_returned_cusip",
    "isin_candidate_matches_case_previous_effective_ric",
    "isin_candidate_matches_case_next_effective_ric",
    "cusip_candidate_matches_case_previous_effective_ric",
    "cusip_candidate_matches_case_next_effective_ric",
    "ticker_candidate_matches_case_previous_effective_ric",
    "ticker_candidate_matches_case_next_effective_ric",
    "raw_isin_matches_case_previous_accepted_identity",
    "raw_isin_matches_case_next_accepted_identity",
    "raw_cusip_matches_case_previous_accepted_identity",
    "raw_cusip_matches_case_next_accepted_identity",
    "ticker_returned_isin_matches_case_previous_accepted_identity",
    "ticker_returned_isin_matches_case_next_accepted_identity",
    "ticker_returned_cusip_matches_case_previous_accepted_identity",
    "ticker_returned_cusip_matches_case_next_accepted_identity",
)

RESOLUTION_DIAGNOSTIC_CONTEXT_COLUMNS: tuple[str, ...] = (
    *RIC_LOOKUP_RESOLUTION_OUTPUT_COLUMNS,
    "LIID_inferred",
    "logical_bridge_group_id",
    "diagnostic_case_id",
    "case_target_bridge_row_id",
    "target_class",
    "diagnostic_role",
    "context_offset",
    "group_sequence_index",
    "group_row_count",
    "case_previous_row_available",
    "case_next_row_available",
    "case_any_adjacent_row_available",
    "case_previous_bridge_row_id",
    "case_next_bridge_row_id",
    "case_previous_effective_available",
    "case_next_effective_available",
    "case_any_adjacent_effective_available",
    "case_both_adjacent_effective_available",
    "case_previous_effective_collection_ric",
    "case_next_effective_collection_ric",
    "case_previous_effective_collection_ric_source",
    "case_next_effective_collection_ric_source",
    "case_previous_accepted_identity_returned_isin",
    "case_next_accepted_identity_returned_isin",
    "case_previous_accepted_identity_returned_cusip",
    "case_next_accepted_identity_returned_cusip",
    "isin_candidate_matches_case_previous_effective_ric",
    "isin_candidate_matches_case_next_effective_ric",
    "cusip_candidate_matches_case_previous_effective_ric",
    "cusip_candidate_matches_case_next_effective_ric",
    "ticker_candidate_matches_case_previous_effective_ric",
    "ticker_candidate_matches_case_next_effective_ric",
    "raw_isin_matches_case_previous_accepted_identity",
    "raw_isin_matches_case_next_accepted_identity",
    "raw_cusip_matches_case_previous_accepted_identity",
    "raw_cusip_matches_case_next_accepted_identity",
    "ticker_returned_isin_matches_case_previous_accepted_identity",
    "ticker_returned_isin_matches_case_next_accepted_identity",
    "ticker_returned_cusip_matches_case_previous_accepted_identity",
    "ticker_returned_cusip_matches_case_next_accepted_identity",
)

RESOLUTION_DIAGNOSTIC_HANDOFF_COLUMNS: tuple[str, ...] = (
    "diagnostic_case_id",
    "case_target_bridge_row_id",
    "target_class",
    "retrieval_sequence_index",
    "retrieval_role",
    "diagnostic_role",
    "bridge_row_id",
    "KYPERMNO",
    "CUSIP",
    "ISIN",
    "TICKER",
    "lookup_input",
    "lookup_input_source",
    "request_start_date",
    "request_end_date",
    "effective_collection_ric",
    "effective_collection_ric_source",
    "accepted_ric",
    "accepted_ric_source",
    "ISIN_returned_ric",
    "CUSIP_returned_ric",
    "ticker_candidate_ric",
    "case_previous_effective_collection_ric",
    "case_next_effective_collection_ric",
)

OWNERSHIP_VALIDATION_RETRIEVAL_ROLE_ORDER: tuple[str, ...] = (
    "PREVIOUS_EFFECTIVE",
    "TARGET_ISIN_CANDIDATE",
    "TARGET_CUSIP_CANDIDATE",
    "TARGET_TICKER_CANDIDATE",
    "NEXT_EFFECTIVE",
)

OWNERSHIP_VALIDATION_VISIBLE_INPUT_FIELDS: tuple[str, ...] = (
    "diagnostic_case_id",
    "target_class",
    "retrieval_role",
    "diagnostic_role",
    "bridge_row_id",
    "KYPERMNO",
    "lookup_input",
    "lookup_input_source",
    "request_start_date",
    "request_end_date",
    "effective_collection_ric",
    "effective_collection_ric_source",
    "accepted_ric",
    "accepted_ric_source",
)

OWNERSHIP_VALIDATION_CASES_PER_SHEET = 100
OWNERSHIP_VALIDATION_CASE_BAND_HEIGHT = 4000
OWNERSHIP_VALIDATION_BLANK_RESULT_STREAK = 50

OWNERSHIP_UNIVERSE_RETRIEVAL_ROLE_ORDER: tuple[str, ...] = (
    "UNIVERSE_EFFECTIVE",
    "UNIVERSE_TARGET_ISIN_CANDIDATE",
    "UNIVERSE_TARGET_CUSIP_CANDIDATE",
    "UNIVERSE_TARGET_TICKER_CANDIDATE",
)

OWNERSHIP_UNIVERSE_RETRIEVAL_SHEET_NAME = "ownership_retrieval"
OWNERSHIP_UNIVERSE_BLOCK_HEADERS: tuple[str, ...] = (
    "input_data",
    "returned_ric",
    "returned_date",
    "returned_value",
    "returned_category",
)

OWNERSHIP_UNIVERSE_VISIBLE_INPUT_FIELDS: tuple[str, ...] = (
    "bridge_row_id",
    "candidate_ric",
    "request_start_date",
    "request_end_date",
    "candidate_slot",
    "lookup_input_source",
    "effective_collection_ric",
    "effective_collection_ric_source",
    "accepted_ric",
    "accepted_ric_source",
)

OWNERSHIP_VALIDATION_HANDOFF_COLUMNS: tuple[str, ...] = (
    *RESOLUTION_DIAGNOSTIC_HANDOFF_COLUMNS,
    "sheet_name",
    "sheet_case_index",
    "case_band_row_start",
    "case_band_row_end",
    "block_slot_index",
    "block_slot_role",
)

OWNERSHIP_VALIDATION_HANDOFF_INT_COLUMNS: frozenset[str] = frozenset(
    {
        "retrieval_sequence_index",
        "sheet_case_index",
        "case_band_row_start",
        "case_band_row_end",
        "block_slot_index",
    }
)

OWNERSHIP_VALIDATION_HANDOFF_TEXT_COLUMNS: tuple[str, ...] = tuple(
    name for name in OWNERSHIP_VALIDATION_HANDOFF_COLUMNS if name not in OWNERSHIP_VALIDATION_HANDOFF_INT_COLUMNS
)

OWNERSHIP_VALIDATION_RESULTS_COLUMNS: tuple[str, ...] = (
    *OWNERSHIP_VALIDATION_HANDOFF_COLUMNS,
    "returned_ric",
    "returned_date",
    "returned_category",
    "returned_value",
)

OWNERSHIP_UNIVERSE_HANDOFF_COLUMNS: tuple[str, ...] = (
    *RIC_LOOKUP_RESOLUTION_OUTPUT_COLUMNS,
    "diagnostic_case_id",
    "candidate_slot",
    "candidate_ric",
    "ownership_lookup_row_id",
    "ownership_lookup_role",
    "lookup_input",
    "lookup_input_source",
    "request_start_date",
    "request_end_date",
    "retrieval_eligible",
    "retrieval_exclusion_reason",
)

OWNERSHIP_UNIVERSE_HANDOFF_INT_COLUMNS: frozenset[str] = frozenset()

OWNERSHIP_UNIVERSE_HANDOFF_TEXT_COLUMNS: tuple[str, ...] = tuple(
    name for name in OWNERSHIP_UNIVERSE_HANDOFF_COLUMNS if name not in OWNERSHIP_UNIVERSE_HANDOFF_INT_COLUMNS
)

OWNERSHIP_UNIVERSE_RESULTS_COLUMNS: tuple[str, ...] = (
    *OWNERSHIP_UNIVERSE_HANDOFF_COLUMNS,
    "returned_ric",
    "returned_date",
    "returned_category",
    "returned_value",
)

OWNERSHIP_UNIVERSE_RESULTS_TEXT_COLUMNS: tuple[str, ...] = tuple(
    name for name in OWNERSHIP_UNIVERSE_RESULTS_COLUMNS if name not in {"returned_date", "returned_value"}
)

OWNERSHIP_UNIVERSE_ROW_SUMMARY_COLUMNS: tuple[str, ...] = (
    *OWNERSHIP_UNIVERSE_HANDOFF_COLUMNS,
    "retrieval_row_present",
    "ownership_rows_returned",
    "ownership_first_date",
    "ownership_last_date",
    "ownership_distinct_categories",
    "ownership_nonnull_value_count",
    "ownership_single_returned_ric",
    "ownership_returned_ric_nunique",
)

OWNERSHIP_VALIDATION_RESULTS_TEXT_COLUMNS: tuple[str, ...] = tuple(
    name for name in OWNERSHIP_VALIDATION_RESULTS_COLUMNS if name not in {"returned_date", "returned_value"}
)

OWNERSHIP_VALIDATION_RETRIEVAL_SUMMARY_COLUMNS: tuple[str, ...] = (
    *OWNERSHIP_VALIDATION_HANDOFF_COLUMNS,
    "ownership_rows_returned",
    "ownership_first_date",
    "ownership_last_date",
    "ownership_distinct_categories",
    "ownership_nonnull_value_count",
    "ownership_single_returned_ric",
    "ownership_returned_ric_nunique",
)

OWNERSHIP_VALIDATION_RETRIEVAL_SUMMARY_TEXT_COLUMNS: tuple[str, ...] = tuple(
    name
    for name in OWNERSHIP_VALIDATION_RETRIEVAL_SUMMARY_COLUMNS
    if name
    not in {
        "retrieval_sequence_index",
        "sheet_case_index",
        "case_band_row_start",
        "case_band_row_end",
        "block_slot_index",
        "ownership_rows_returned",
        "ownership_distinct_categories",
        "ownership_nonnull_value_count",
        "ownership_returned_ric_nunique",
        "ownership_single_returned_ric",
        "ownership_first_date",
        "ownership_last_date",
    }
)

OWNERSHIP_VALIDATION_PAIRWISE_COMPARISON_COLUMNS: tuple[str, ...] = (
    "diagnostic_case_id",
    "target_class",
    "case_target_bridge_row_id",
    "target_retrieval_role",
    "target_lookup_input",
    "target_lookup_input_source",
    "target_bridge_row_id",
    "adjacent_retrieval_role",
    "adjacent_lookup_input",
    "adjacent_lookup_input_source",
    "adjacent_bridge_row_id",
    "adjacent_direction",
    "overlap_date_count",
    "overlap_category_count",
    "overlap_date_category_pair_count",
    "same_returned_ric",
    "same_category_set_on_overlap",
    "matched_value_pair_count",
    "mean_abs_value_diff",
    "median_abs_value_diff",
    "max_abs_value_diff",
    "pair_has_useful_overlap",
    "pair_supports_corrobation",
    "pair_supports_same_identity_ric_variant",
    "pair_conflicts",
)

OWNERSHIP_VALIDATION_PAIRWISE_TEXT_COLUMNS: tuple[str, ...] = tuple(
    name
    for name in OWNERSHIP_VALIDATION_PAIRWISE_COMPARISON_COLUMNS
    if name
    not in {
        "overlap_date_count",
        "overlap_category_count",
        "overlap_date_category_pair_count",
        "same_returned_ric",
        "same_category_set_on_overlap",
        "matched_value_pair_count",
        "mean_abs_value_diff",
        "median_abs_value_diff",
        "max_abs_value_diff",
        "pair_has_useful_overlap",
        "pair_supports_corrobation",
        "pair_supports_same_identity_ric_variant",
        "pair_conflicts",
    }
)

OWNERSHIP_VALIDATION_CASE_SUMMARY_COLUMNS: tuple[str, ...] = (
    "diagnostic_case_id",
    "target_class",
    "case_target_bridge_row_id",
    "candidate_has_ownership_data",
    "any_adjacent_effective_has_ownership_data",
    "candidate_matches_previous_effective_ownership",
    "candidate_matches_next_effective_ownership",
    "candidate_matches_any_adjacent_effective_ownership",
    "candidate_retrieval_rows_with_data",
    "adjacent_effective_retrieval_rows_with_data",
    "pair_supports_corrobation_count",
    "pair_supports_same_identity_ric_variant_count",
    "pair_conflicts_count",
    "ownership_validation_bucket",
)

OWNERSHIP_VALIDATION_CASE_SUMMARY_TEXT_COLUMNS: tuple[str, ...] = tuple(
    name
    for name in OWNERSHIP_VALIDATION_CASE_SUMMARY_COLUMNS
    if name
    not in {
        "candidate_has_ownership_data",
        "any_adjacent_effective_has_ownership_data",
        "candidate_matches_previous_effective_ownership",
        "candidate_matches_next_effective_ownership",
        "candidate_matches_any_adjacent_effective_ownership",
        "candidate_retrieval_rows_with_data",
        "adjacent_effective_retrieval_rows_with_data",
        "pair_supports_corrobation_count",
        "pair_supports_same_identity_ric_variant_count",
        "pair_conflicts_count",
    }
)

OWNERSHIP_VALIDATION_USEFUL_OVERLAP_MIN_MATCHED_VALUE_PAIRS = 12
OWNERSHIP_VALIDATION_USEFUL_OVERLAP_MIN_DATES = 3
OWNERSHIP_VALIDATION_USEFUL_OVERLAP_MIN_CATEGORIES = 3
OWNERSHIP_VALIDATION_SUPPORT_MEDIAN_ABS_DIFF_MAX = 0.5
OWNERSHIP_VALIDATION_SUPPORT_MEAN_ABS_DIFF_MAX = 1.0
OWNERSHIP_VALIDATION_SUPPORT_MAX_ABS_DIFF_MAX = 5.0
OWNERSHIP_VALIDATION_CONFLICT_MEAN_ABS_DIFF_MIN = 3.0
OWNERSHIP_VALIDATION_CONFLICT_MAX_ABS_DIFF_MIN = 10.0

RESOLUTION_DIAGNOSTIC_DATE_COLUMNS: tuple[str, ...] = (
    "first_seen_caldt",
    "last_seen_caldt",
    "extended_from_span_start",
    "extended_from_span_end",
)

RESOLUTION_DIAGNOSTIC_BOOLEAN_COLUMNS: frozenset[str] = frozenset(
    {
        *RIC_LOOKUP_EXTENDED_BOOLEAN_COLUMNS,
        "conventional_identity_conflict",
        "ticker_candidate_available",
        "ticker_candidate_conflicts_with_conventional",
        "case_previous_row_available",
        "case_next_row_available",
        "case_any_adjacent_row_available",
        "case_previous_effective_available",
        "case_next_effective_available",
        "case_any_adjacent_effective_available",
        "case_both_adjacent_effective_available",
        "isin_candidate_matches_case_previous_effective_ric",
        "isin_candidate_matches_case_next_effective_ric",
        "cusip_candidate_matches_case_previous_effective_ric",
        "cusip_candidate_matches_case_next_effective_ric",
        "ticker_candidate_matches_case_previous_effective_ric",
        "ticker_candidate_matches_case_next_effective_ric",
        "raw_isin_matches_case_previous_accepted_identity",
        "raw_isin_matches_case_next_accepted_identity",
        "raw_cusip_matches_case_previous_accepted_identity",
        "raw_cusip_matches_case_next_accepted_identity",
        "ticker_returned_isin_matches_case_previous_accepted_identity",
        "ticker_returned_isin_matches_case_next_accepted_identity",
        "ticker_returned_cusip_matches_case_previous_accepted_identity",
        "ticker_returned_cusip_matches_case_next_accepted_identity",
    }
)

RESOLUTION_DIAGNOSTIC_INT_COLUMNS: frozenset[str] = frozenset(
    {
        "group_sequence_index",
        "group_row_count",
        "context_offset",
        "retrieval_sequence_index",
    }
)

RESOLUTION_DIAGNOSTIC_TARGET_TEXT_COLUMNS: tuple[str, ...] = tuple(
    name
    for name in RESOLUTION_DIAGNOSTIC_TARGET_COLUMNS
    if name not in RESOLUTION_DIAGNOSTIC_BOOLEAN_COLUMNS
    and name not in RESOLUTION_DIAGNOSTIC_INT_COLUMNS
)

RESOLUTION_DIAGNOSTIC_CONTEXT_TEXT_COLUMNS: tuple[str, ...] = tuple(
    name
    for name in RESOLUTION_DIAGNOSTIC_CONTEXT_COLUMNS
    if name not in RESOLUTION_DIAGNOSTIC_BOOLEAN_COLUMNS
    and name not in RESOLUTION_DIAGNOSTIC_INT_COLUMNS
)


@dataclass(frozen=True)
class LookupIdentityCandidate:
    source: str
    returned_ric: str
    returned_isin: str | None
    returned_cusip: str | None


NULL_RIC_DIAGNOSTIC_COLUMNS: tuple[str, ...] = (
    *RIC_LOOKUP_COLUMNS,
    "failed_lookup_flag",
    "failed_lookup_reason",
    "invalid_identifier_signal",
    "successful_row_exists_for_kypermno",
    "successful_row_exists_before_span",
    "successful_row_exists_after_span",
    "successful_row_overlap_exists",
    "nearest_successful_gap_days_before",
    "nearest_successful_gap_days_after",
    "unique_successful_identifier_pair_count",
    "unique_successful_ric_count",
    "failed_identifier_pair_matches_success",
    "alternative_identifier_available",
    "alternative_identifier",
    "alternative_identifier_type",
    "candidate_successful_ric_available",
    "candidate_successful_ric",
    "candidate_successful_cusip",
    "candidate_successful_isin",
)

NULL_RIC_DIAGNOSTIC_TEXT_COLUMNS: tuple[str, ...] = (
    "bridge_row_id",
    "KYPERMNO",
    "CUSIP",
    "ISIN",
    "TICKER",
    "first_seen_caldt",
    "last_seen_caldt",
    "preferred_lookup_id",
    "preferred_lookup_type",
    "vendor_primary_ric",
    "vendor_returned_name",
    "vendor_returned_cusip",
    "vendor_returned_isin",
    "vendor_match_status",
    "vendor_notes",
    "failed_lookup_reason",
    "alternative_identifier_available",
    "candidate_successful_ric",
    "candidate_successful_cusip",
    "candidate_successful_isin",
    "alternative_identifier",
    "alternative_identifier_type",
    "candidate_successful_ric_available",
)

NULL_RIC_REVIEW_COLUMNS: tuple[str, ...] = (
    "bridge_row_id",
    "KYPERMNO",
    "TICKER",
    "first_seen_caldt",
    "last_seen_caldt",
    "preferred_lookup_id",
    "preferred_lookup_type",
    "vendor_primary_ric",
    "alternative_identifier",
    "alternative_identifier_type",
    "candidate_successful_ric",
    "successful_row_exists_for_kypermno",
    "successful_row_exists_before_span",
    "successful_row_exists_after_span",
    "successful_row_overlap_exists",
    "nearest_successful_gap_days_before",
    "nearest_successful_gap_days_after",
    "unique_successful_identifier_pair_count",
    "unique_successful_ric_count",
    "alternative_identifier_available",
    "candidate_successful_ric_available",
    "test_category",
    "test_method",
    "test_result",
    "test_notes",
)

NULL_RIC_REVIEW_TEXT_COLUMNS: tuple[str, ...] = (
    "bridge_row_id",
    "KYPERMNO",
    "TICKER",
    "first_seen_caldt",
    "last_seen_caldt",
    "preferred_lookup_id",
    "preferred_lookup_type",
    "vendor_primary_ric",
    "alternative_identifier",
    "alternative_identifier_type",
    "candidate_successful_ric",
    "test_category",
    "test_method",
    "test_result",
    "test_notes",
)

OWNERSHIP_SMOKE_SAMPLE_COLUMNS: tuple[str, ...] = (
    "sample_category",
    "bridge_row_id",
    "KYPERMNO",
    "TICKER",
    "lookup_input",
    "lookup_input_source",
    "request_start_date",
    "request_end_date",
    "preferred_lookup_id",
    "preferred_lookup_type",
    "alternative_identifier",
    "alternative_identifier_type",
    "candidate_successful_ric",
    "successful_row_exists_before_span",
    "successful_row_exists_after_span",
    "successful_row_overlap_exists",
    "alternative_identifier_available",
    "candidate_successful_ric_available",
    "unique_successful_identifier_pair_count",
    "unique_successful_ric_count",
)

OWNERSHIP_SMOKE_SAMPLE_SCHEMA: dict[str, pl.DataType] = {
    "sample_category": pl.Utf8,
    "bridge_row_id": pl.Utf8,
    "KYPERMNO": pl.Utf8,
    "TICKER": pl.Utf8,
    "lookup_input": pl.Utf8,
    "lookup_input_source": pl.Utf8,
    "request_start_date": pl.Utf8,
    "request_end_date": pl.Utf8,
    "preferred_lookup_id": pl.Utf8,
    "preferred_lookup_type": pl.Utf8,
    "alternative_identifier": pl.Utf8,
    "alternative_identifier_type": pl.Utf8,
    "candidate_successful_ric": pl.Utf8,
    "successful_row_exists_before_span": pl.Boolean,
    "successful_row_exists_after_span": pl.Boolean,
    "successful_row_overlap_exists": pl.Boolean,
    "alternative_identifier_available": pl.Boolean,
    "candidate_successful_ric_available": pl.Boolean,
    "unique_successful_identifier_pair_count": pl.Int64,
    "unique_successful_ric_count": pl.Int64,
}

OWNERSHIP_SMOKE_CATEGORY_SPECS: tuple[tuple[str, Callable[[dict[str, Any]], bool]], ...] = (
    ("successful_row_exists_after_span", lambda row: bool(row.get("successful_row_exists_after_span"))),
    ("successful_row_overlap_exists", lambda row: bool(row.get("successful_row_overlap_exists"))),
    ("successful_row_exists_before_span", lambda row: bool(row.get("successful_row_exists_before_span"))),
    ("alternative_identifier_available", lambda row: bool(row.get("alternative_identifier_available"))),
    (
        "no_successful_row_for_kypermno",
        lambda row: not bool(row.get("successful_row_exists_for_kypermno")),
    ),
    (
        "multiple_successful_identifier_pairs_or_rics",
        lambda row: int(row.get("unique_successful_identifier_pair_count") or 0) > 1
        or int(row.get("unique_successful_ric_count") or 0) > 1,
    ),
)


def _require_columns(lf: pl.LazyFrame, required: tuple[str, ...], label: str) -> None:
    schema = lf.collect_schema()
    missing = [name for name in required if name not in schema]
    if missing:
        raise ValueError(f"{label} missing required columns: {missing}")


def _normalize_company_description_lf(
    company_description_lf: pl.LazyFrame | None,
    *,
    target_gvkey_dtype: pl.DataType,
) -> pl.LazyFrame | None:
    if company_description_lf is None:
        return None

    schema = company_description_lf.collect_schema()
    if "KYGVKEY" not in schema:
        raise ValueError("company_description missing required column: KYGVKEY")

    name_candidates = [name for name in ("CONM", "CONML") if name in schema]
    if not name_candidates:
        return None

    return (
        company_description_lf.select(
            pl.col("KYGVKEY").cast(target_gvkey_dtype, strict=False).alias("KYGVKEY_final"),
            pl.coalesce([pl.col(name).cast(pl.Utf8, strict=False) for name in name_candidates]).alias("company_name"),
        )
        .drop_nulls(subset=["KYGVKEY_final"])
        .filter(pl.col("company_name").is_not_null())
        .unique(subset=["KYGVKEY_final"], keep="first")
    )


def _bridge_row_id_expr() -> pl.Expr:
    return pl.concat_str(
        [
            pl.col("KYPERMNO").cast(pl.Utf8, strict=False),
            pl.coalesce([pl.col("LIID"), pl.lit("-")]),
            pl.coalesce([pl.col("CUSIP"), pl.lit("-")]),
            pl.coalesce([pl.col("ISIN"), pl.lit("-")]),
            pl.coalesce([pl.col("TICKER"), pl.lit("-")]),
        ],
        separator=":",
    )


def build_refinitiv_step1_bridge_universe(
    daily_lf: pl.LazyFrame,
    *,
    company_description_lf: pl.LazyFrame | None = None,
) -> pl.LazyFrame:
    """Build the compact security-identifier bridge universe for Refinitiv step 1."""
    _require_columns(daily_lf, REQUIRED_DAILY_COLUMNS, "daily panel")
    schema = daily_lf.collect_schema()

    projected_cols = list(REQUIRED_DAILY_COLUMNS)
    projected_cols.extend(name for name in OPTIONAL_DAILY_COLUMNS if name in schema)

    n_filings_expr: pl.Expr
    if "n_filings" in schema:
        n_filings_expr = pl.col("n_filings").cast(pl.Int64, strict=False)
    else:
        n_filings_expr = pl.lit(0, dtype=pl.Int64)

    base = (
        daily_lf.select(projected_cols)
        .drop_nulls(subset=["KYPERMNO", "CALDT"])
        .with_columns(
            pl.col("KYPERMNO").cast(pl.Int32, strict=False),
            pl.col("CALDT").cast(pl.Date, strict=False),
            pl.col("KYGVKEY_final").cast(pl.Utf8, strict=False),
            pl.col("LIID").cast(pl.Utf8, strict=False),
            pl.col("CIK_final").cast(pl.Utf8, strict=False),
            _clean_identifier_expr("CUSIP"),
            _clean_identifier_expr("ISIN"),
            _clean_identifier_expr("TICKER"),
            (
                pl.col("LINKTYPE").cast(pl.Utf8, strict=False)
                if "LINKTYPE" in schema
                else pl.lit(None, dtype=pl.Utf8)
            ).alias("LINKTYPE"),
            (
                pl.col("LINKPRIM").cast(pl.Utf8, strict=False)
                if "LINKPRIM" in schema
                else pl.lit(None, dtype=pl.Utf8)
            ).alias("LINKPRIM"),
            (
                pl.col("link_quality_flag").cast(pl.Utf8, strict=False)
                if "link_quality_flag" in schema
                else pl.lit(None, dtype=pl.Utf8)
            ).alias("link_quality_flag"),
            (
                pl.col("HEXCNTRY").cast(pl.Utf8, strict=False)
                if "HEXCNTRY" in schema
                else pl.lit(None, dtype=pl.Utf8)
            ).alias("HEXCNTRY"),
            n_filings_expr.alias("n_filings"),
        )
        .sort("KYPERMNO", "LIID", "CUSIP", "ISIN", "TICKER", "CALDT")
    )

    grouped = base.group_by(["KYPERMNO", "LIID", "CUSIP", "ISIN", "TICKER"], maintain_order=True).agg(
        pl.col("KYGVKEY_final").drop_nulls().last().alias("KYGVKEY_final"),
        pl.col("CIK_final").drop_nulls().last().alias("CIK_final"),
        pl.col("LINKTYPE").drop_nulls().last().alias("LINKTYPE"),
        pl.col("LINKPRIM").drop_nulls().last().alias("LINKPRIM"),
        pl.col("link_quality_flag").drop_nulls().last().alias("link_quality_flag"),
        pl.col("HEXCNTRY").drop_nulls().last().alias("HEXCNTRY"),
        pl.col("CALDT").min().alias("first_seen_caldt"),
        pl.col("CALDT").max().alias("last_seen_caldt"),
        pl.len().cast(pl.Int64).alias("n_daily_rows"),
        (pl.col("n_filings") > pl.lit(0)).cast(pl.Int64).sum().alias("n_filing_days"),
        pl.col("n_filings").sum().cast(pl.Int64).alias("n_filings_total"),
    )

    company_names = _normalize_company_description_lf(
        company_description_lf,
        target_gvkey_dtype=pl.Utf8,
    )
    if company_names is not None:
        grouped = grouped.join(company_names, on="KYGVKEY_final", how="left")
    else:
        grouped = grouped.with_columns(pl.lit(None, dtype=pl.Utf8).alias("company_name"))

    bridge = grouped.with_columns(_bridge_row_id_expr().alias("bridge_row_id"))

    for name in BRIDGE_VENDOR_COLUMNS:
        bridge = bridge.with_columns(pl.lit(None, dtype=pl.Utf8).alias(name))

    return bridge.select(BRIDGE_OUTPUT_COLUMNS).sort(
        "KYPERMNO",
        "LIID",
        "CUSIP",
        "ISIN",
        "TICKER",
        "first_seen_caldt",
    )


def _build_handoff_df(df: pl.DataFrame) -> pl.DataFrame:
    return _coerce_text_columns(df.select(BRIDGE_OUTPUT_COLUMNS), WORKBOOK_TEXT_COLUMNS)


def _coerce_text_columns(df: pl.DataFrame, text_columns: tuple[str, ...]) -> pl.DataFrame:
    exprs: list[pl.Expr] = []
    text_col_set = set(text_columns)
    for name in df.columns:
        if name in text_col_set:
            if name in ("first_seen_caldt", "last_seen_caldt"):
                exprs.append(
                    pl.when(pl.col(name).is_null())
                    .then(pl.lit(None, dtype=pl.Utf8))
                    .otherwise(pl.col(name).cast(pl.Date, strict=False).dt.strftime("%Y-%m-%d"))
                    .alias(name)
                )
            else:
                exprs.append(
                    pl.when(pl.col(name).is_null())
                    .then(pl.lit(None, dtype=pl.Utf8))
                    .otherwise(pl.col(name).cast(pl.Utf8, strict=False))
                    .alias(name)
                )
        else:
            exprs.append(pl.col(name))
    return df.select(exprs)


def _clean_text_value_expr(name: str) -> pl.Expr:
    cleaned = (
        pl.when(pl.col(name).is_null())
        .then(pl.lit(None, dtype=pl.Utf8))
        .otherwise(pl.col(name).cast(pl.Utf8, strict=False).str.strip_chars())
    )
    return pl.when(cleaned.is_null() | cleaned.eq("")).then(pl.lit(None, dtype=pl.Utf8)).otherwise(cleaned)


def _clean_text_expr(name: str) -> pl.Expr:
    return _clean_text_value_expr(name).alias(name)


def _clean_identifier_value_expr(name: str) -> pl.Expr:
    return _clean_text_value_expr(name)


def _clean_identifier_expr(name: str) -> pl.Expr:
    return (
        _clean_identifier_value_expr(name).alias(name)
    )


def _build_ric_lookup_base_frame(df: pl.DataFrame) -> pl.DataFrame:
    return df.select(
        pl.col("bridge_row_id"),
        pl.col("KYPERMNO"),
        _clean_identifier_expr("CUSIP"),
        _clean_identifier_expr("ISIN"),
        _clean_identifier_expr("TICKER"),
        pl.col("first_seen_caldt"),
        pl.col("last_seen_caldt"),
        _clean_text_expr("vendor_primary_ric"),
        _clean_text_expr("vendor_returned_name"),
        _clean_identifier_expr("vendor_returned_cusip"),
        _clean_identifier_expr("vendor_returned_isin"),
        _clean_text_expr("vendor_match_status"),
        _clean_text_expr("vendor_notes"),
    ).with_columns(
        pl.when(pl.col("ISIN").is_not_null() & pl.col("ISIN").ne(""))
        .then(pl.col("ISIN"))
        .when(pl.col("CUSIP").is_not_null() & pl.col("CUSIP").ne(""))
        .then(pl.col("CUSIP"))
        .when(pl.col("TICKER").is_not_null() & pl.col("TICKER").ne(""))
        .then(pl.col("TICKER"))
        .otherwise(pl.lit(None, dtype=pl.Utf8))
        .alias("preferred_lookup_id"),
        pl.when(pl.col("ISIN").is_not_null() & pl.col("ISIN").ne(""))
        .then(pl.lit("ISIN"))
        .when(pl.col("CUSIP").is_not_null() & pl.col("CUSIP").ne(""))
        .then(pl.lit("CUSIP"))
        .when(pl.col("TICKER").is_not_null() & pl.col("TICKER").ne(""))
        .then(pl.lit("TICKER"))
        .otherwise(pl.lit(None, dtype=pl.Utf8))
        .alias("preferred_lookup_type"),
    )


def _build_ric_lookup_handoff_frames(df: pl.DataFrame) -> tuple[pl.DataFrame, pl.DataFrame]:
    lookup_base = _build_ric_lookup_base_frame(df)

    ordered_lookup = lookup_base.select(RIC_LOOKUP_COLUMNS)
    lookup_df = ordered_lookup.filter(pl.col("preferred_lookup_id").is_not_null())
    manual_review_df = ordered_lookup.filter(pl.col("preferred_lookup_id").is_null())
    return (
        _coerce_text_columns(lookup_df, RIC_LOOKUP_TEXT_COLUMNS),
        _coerce_text_columns(manual_review_df, RIC_LOOKUP_TEXT_COLUMNS),
    )


def _combine_lookup_frames(lookup_df: pl.DataFrame, manual_review_df: pl.DataFrame) -> pl.DataFrame:
    frames = [lookup_df]
    if manual_review_df.height:
        frames.append(manual_review_df)
    return pl.concat(frames, how="vertical_relaxed").sort("bridge_row_id") if frames else pl.DataFrame()


def _profile_output_key(profile_name: str) -> str:
    return f"refinitiv_ric_lookup_handoff_{profile_name}_xlsx"


def _profile_output_path(output_dir: Path, profile_name: str) -> Path:
    return output_dir / f"refinitiv_ric_lookup_handoff_{profile_name}.xlsx"


def _extended_profile_output_key(profile_name: str) -> str:
    return f"refinitiv_ric_lookup_handoff_{profile_name}_extended_xlsx"


def _extended_profile_output_path(output_dir: Path, profile_name: str) -> Path:
    return output_dir / f"refinitiv_ric_lookup_handoff_{profile_name}_extended.xlsx"


def _build_lookup_profile_bridge_ids(
    daily_lf: pl.LazyFrame,
    profile: RicLookupFilterProfile,
) -> pl.DataFrame:
    _require_columns(
        daily_lf,
        REQUIRED_DAILY_COLUMNS + ("SHRCD",),
        f"daily panel for lookup profile {profile.name}",
    )

    eligible = (
        daily_lf.select("KYPERMNO", "CALDT", "LIID", "CUSIP", "ISIN", "TICKER", "SHRCD")
        .drop_nulls(subset=["KYPERMNO", "CALDT"])
        .with_columns(
            pl.col("KYPERMNO").cast(pl.Int32, strict=False),
            pl.col("CALDT").cast(pl.Date, strict=False),
            pl.col("LIID").cast(pl.Utf8, strict=False),
            _clean_identifier_expr("CUSIP"),
            _clean_identifier_expr("ISIN"),
            _clean_identifier_expr("TICKER"),
        )
        .with_columns(_bridge_row_id_expr().alias("bridge_row_id"))
    )
    if profile.require_common_stock:
        eligible = eligible.filter(common_stock_pass_expr("SHRCD"))
    return eligible.select("bridge_row_id").unique(subset=["bridge_row_id"]).collect()


def _build_filtered_ric_lookup_profile_artifact(
    bridge_df: pl.DataFrame,
    qualifying_bridge_ids: pl.DataFrame,
    profile: RicLookupFilterProfile,
) -> tuple[pl.DataFrame, pl.DataFrame, dict[str, Any]]:
    filtered_bridge = bridge_df.join(qualifying_bridge_ids, on="bridge_row_id", how="semi")
    if profile.require_gvkey:
        filtered_bridge = filtered_bridge.filter(pl.col("KYGVKEY_final").is_not_null())

    lookup_df, manual_review_df = _build_ric_lookup_handoff_frames(filtered_bridge)
    summary: dict[str, Any] = {
        "profile_name": profile.name,
        "predicates_applied": list(profile.predicates),
        "rows_before_filter": int(bridge_df.height),
        "rows_after_filter": int(filtered_bridge.height),
        "ric_lookup_rows": int(lookup_df.height),
        "ric_manual_review_rows": int(manual_review_df.height),
    }
    return lookup_df, manual_review_df, summary


def _lookup_input_expr(identifier_type: str) -> pl.Expr:
    return _clean_identifier_value_expr(identifier_type).alias(f"{identifier_type}_lookup_input")


def _pairwise_same_value_expr(left_type: str, right_type: str, suffix: str) -> pl.Expr:
    left_column = f"{left_type}_{suffix}"
    right_column = f"{right_type}_{suffix}"
    result_name = f"{left_type}_vs_{right_type}_same_{suffix.split('_', 1)[1]}"
    return (
        pl.when(
            pl.col(f"{left_type}_success")
            & pl.col(f"{right_type}_success")
            & pl.col(left_column).is_not_null()
            & pl.col(right_column).is_not_null()
        )
        .then(pl.col(left_column) == pl.col(right_column))
        .otherwise(pl.lit(None, dtype=pl.Boolean))
        .alias(result_name)
    )


def _successful_distinct_value_count_expr(suffix: str) -> pl.Expr:
    return (
        pl.concat_list(
            [
                pl.when(pl.col(f"{identifier_type}_success"))
                .then(pl.col(f"{identifier_type}_{suffix}"))
                .otherwise(pl.lit(None, dtype=pl.Utf8))
                for identifier_type in LOOKUP_IDENTIFIER_TYPES
            ]
        )
        .list.drop_nulls()
        .list.unique()
        .list.len()
    )


def _normalized_lookup_result_value_expr(name: str) -> pl.Expr:
    cleaned = _clean_text_value_expr(name)
    upper_cleaned = cleaned.str.to_uppercase()
    invalid_identifier_signal = (
        cleaned.str.to_lowercase().str.contains("invalid identifier").fill_null(False)
    )
    return (
        pl.when(
            cleaned.is_null()
            | upper_cleaned.is_in(list(LOOKUP_FAILURE_MARKERS))
            | invalid_identifier_signal
        )
        .then(pl.lit(None, dtype=pl.Utf8))
        .otherwise(cleaned)
        .alias(name)
    )


def _extended_lookup_schema() -> dict[str, pl.DataType]:
    schema: dict[str, pl.DataType] = {}
    for name in RIC_LOOKUP_EXTENDED_COLUMNS:
        if name in RIC_LOOKUP_EXTENDED_DATE_COLUMNS:
            schema[name] = pl.Date
        elif name in RIC_LOOKUP_EXTENDED_BOOLEAN_COLUMNS:
            schema[name] = pl.Boolean
        else:
            schema[name] = pl.Utf8
    return schema


def _resolution_output_schema() -> dict[str, pl.DataType]:
    schema = _extended_lookup_schema()
    schema.update(
        {
            "accepted_identity_returned_isin": pl.Utf8,
            "accepted_identity_returned_cusip": pl.Utf8,
            "accepted_ric": pl.Utf8,
            "accepted_ric_source": pl.Utf8,
            "accepted_resolution_status": pl.Utf8,
            "conventional_identity_conflict": pl.Boolean,
            "ticker_candidate_ric": pl.Utf8,
            "ticker_candidate_available": pl.Boolean,
            "ticker_candidate_conflicts_with_conventional": pl.Boolean,
            "extended_ric": pl.Utf8,
            "extended_from_bridge_row_id": pl.Utf8,
            "extended_from_span_start": pl.Date,
            "extended_from_span_end": pl.Date,
            "extension_direction": pl.Utf8,
            "extension_status": pl.Utf8,
            "effective_collection_ric": pl.Utf8,
            "effective_collection_ric_source": pl.Utf8,
            "effective_resolution_status": pl.Utf8,
        }
    )
    return schema


def _cast_df_to_schema(df: pl.DataFrame, schema: dict[str, pl.DataType]) -> pl.DataFrame:
    exprs: list[pl.Expr] = []
    for name, dtype in schema.items():
        if dtype == pl.Date:
            exprs.append(
                pl.when(pl.col(name).is_null())
                .then(pl.lit(None, dtype=pl.Date))
                .otherwise(pl.col(name).cast(pl.Utf8, strict=False).str.strptime(pl.Date, strict=False))
                .alias(name)
            )
        else:
            exprs.append(pl.col(name).cast(dtype, strict=False).alias(name))
    return df.select(exprs)


def _empty_resolution_df() -> pl.DataFrame:
    schema = _resolution_output_schema()
    return _cast_df_to_schema(
        pl.DataFrame(schema={name: dtype for name, dtype in schema.items()}),
        schema,
    )


def _normalize_resolution_input_df(lookup_df: pl.DataFrame) -> pl.DataFrame:
    normalized = lookup_df.select(
        _clean_text_expr("bridge_row_id"),
        _clean_text_expr("KYPERMNO"),
        _clean_identifier_expr("CUSIP"),
        _clean_identifier_expr("ISIN"),
        _clean_identifier_expr("TICKER"),
        pl.col("first_seen_caldt").cast(pl.Date, strict=False).alias("first_seen_caldt"),
        pl.col("last_seen_caldt").cast(pl.Date, strict=False).alias("last_seen_caldt"),
        _clean_text_expr("vendor_primary_ric"),
        _clean_text_expr("vendor_returned_name"),
        _clean_identifier_expr("vendor_returned_cusip"),
        _clean_identifier_expr("vendor_returned_isin"),
        _clean_text_expr("vendor_match_status"),
        _clean_text_expr("vendor_notes"),
        *[_clean_identifier_expr(f"{identifier_type}_lookup_input") for identifier_type in LOOKUP_IDENTIFIER_TYPES],
        *[
            _normalized_lookup_result_value_expr(f"{identifier_type}_returned_ric")
            for identifier_type in LOOKUP_IDENTIFIER_TYPES
        ],
        *[
            _normalized_lookup_result_value_expr(f"{identifier_type}_returned_name")
            for identifier_type in LOOKUP_IDENTIFIER_TYPES
        ],
        *[
            _normalized_lookup_result_value_expr(f"{identifier_type}_returned_isin")
            for identifier_type in LOOKUP_IDENTIFIER_TYPES
        ],
        *[
            _normalized_lookup_result_value_expr(f"{identifier_type}_returned_cusip")
            for identifier_type in LOOKUP_IDENTIFIER_TYPES
        ],
    )
    normalized = normalized.with_columns(
        *[
            pl.col(f"{identifier_type}_lookup_input").is_not_null().alias(f"{identifier_type}_attempted")
            for identifier_type in LOOKUP_IDENTIFIER_TYPES
        ],
        *[
            pl.col(f"{identifier_type}_returned_ric").is_not_null().alias(f"{identifier_type}_success")
            for identifier_type in LOOKUP_IDENTIFIER_TYPES
        ],
    )
    normalized = normalized.with_columns(
        *[
            _pairwise_same_value_expr(left_type, right_type, "returned_ric")
            for left_type, right_type in LOOKUP_IDENTIFIER_PAIRS
        ],
        *[
            _pairwise_same_value_expr(left_type, right_type, "returned_isin")
            for left_type, right_type in LOOKUP_IDENTIFIER_PAIRS
        ],
        *[
            _pairwise_same_value_expr(left_type, right_type, "returned_cusip")
            for left_type, right_type in LOOKUP_IDENTIFIER_PAIRS
        ],
    )
    success_count_expr = pl.sum_horizontal(
        [pl.col(f"{identifier_type}_success").cast(pl.Int64) for identifier_type in LOOKUP_IDENTIFIER_TYPES]
    )
    return normalized.with_columns(
        pl.when(success_count_expr >= 2)
        .then(
            (_successful_distinct_value_count_expr("returned_ric") <= 1)
            & (_successful_distinct_value_count_expr("returned_isin") <= 1)
            & (_successful_distinct_value_count_expr("returned_cusip") <= 1)
        )
        .otherwise(pl.lit(None, dtype=pl.Boolean))
        .alias("all_successful_attempts_consistent")
    ).select(RIC_LOOKUP_EXTENDED_COLUMNS)


def build_refinitiv_lookup_extended_diagnostic_artifact(
    lookup_df: pl.DataFrame,
    manual_review_df: pl.DataFrame | None = None,
) -> tuple[pl.DataFrame, pl.DataFrame, dict[str, Any]]:
    combined = _combine_lookup_frames(
        lookup_df,
        manual_review_df if manual_review_df is not None else pl.DataFrame(schema={name: pl.Utf8 for name in RIC_LOOKUP_COLUMNS}),
    )
    if combined.height == 0:
        empty_extended = pl.DataFrame(schema={name: pl.Utf8 for name in RIC_LOOKUP_EXTENDED_COLUMNS}).with_columns(
            *[
                pl.col(name).cast(pl.Boolean, strict=False)
                for name in RIC_LOOKUP_EXTENDED_COLUMNS
                if name.endswith("_attempted")
                or name.endswith("_success")
                or "_same_" in name
                or name == "all_successful_attempts_consistent"
            ]
        )
        empty_summary = pl.DataFrame(schema={
            "summary_category": pl.Utf8,
            "summary_key": pl.Utf8,
            "value": pl.Int64,
        })
        summary_payload = {
            "attempt_counts_by_identifier_type": {identifier_type: 0 for identifier_type in LOOKUP_IDENTIFIER_TYPES},
            "success_counts_by_identifier_type": {identifier_type: 0 for identifier_type in LOOKUP_IDENTIFIER_TYPES},
            "agreement_counts": {},
            "conflict_counts": {},
            "rows_where_only_one_identifier_type_succeeds": 0,
        }
        return empty_extended, empty_summary, summary_payload

    base = combined.select(
        _clean_text_expr("bridge_row_id"),
        _clean_text_expr("KYPERMNO"),
        _clean_identifier_expr("CUSIP"),
        _clean_identifier_expr("ISIN"),
        _clean_identifier_expr("TICKER"),
        pl.when(pl.col("first_seen_caldt").is_null())
        .then(pl.lit(None, dtype=pl.Utf8))
        .otherwise(pl.col("first_seen_caldt").cast(pl.Utf8, strict=False))
        .alias("first_seen_caldt"),
        pl.when(pl.col("last_seen_caldt").is_null())
        .then(pl.lit(None, dtype=pl.Utf8))
        .otherwise(pl.col("last_seen_caldt").cast(pl.Utf8, strict=False))
        .alias("last_seen_caldt"),
        _clean_text_expr("vendor_primary_ric"),
        _clean_text_expr("vendor_returned_name"),
        _clean_identifier_expr("vendor_returned_cusip"),
        _clean_identifier_expr("vendor_returned_isin"),
        _clean_text_expr("vendor_match_status"),
        _clean_text_expr("vendor_notes"),
    )

    per_identifier_exprs: list[pl.Expr] = []
    for identifier_type in LOOKUP_IDENTIFIER_TYPES:
        per_identifier_exprs.extend(
            (
                _lookup_input_expr(identifier_type),
                pl.lit(None, dtype=pl.Utf8).alias(f"{identifier_type}_returned_ric"),
                pl.lit(None, dtype=pl.Utf8).alias(f"{identifier_type}_returned_name"),
                pl.lit(None, dtype=pl.Utf8).alias(f"{identifier_type}_returned_isin"),
                pl.lit(None, dtype=pl.Utf8).alias(f"{identifier_type}_returned_cusip"),
            )
        )

    extended = base.with_columns(per_identifier_exprs)
    extended = extended.with_columns(
        *[
            pl.col(f"{identifier_type}_lookup_input").is_not_null().alias(f"{identifier_type}_attempted")
            for identifier_type in LOOKUP_IDENTIFIER_TYPES
        ],
        *[
            pl.col(f"{identifier_type}_returned_ric").is_not_null().alias(f"{identifier_type}_success")
            for identifier_type in LOOKUP_IDENTIFIER_TYPES
        ],
    )
    extended = extended.with_columns(
        *[
            pl.lit(None, dtype=pl.Boolean).alias(f"{left_type}_vs_{right_type}_same_ric")
            for left_type, right_type in LOOKUP_IDENTIFIER_PAIRS
        ],
        *[
            pl.lit(None, dtype=pl.Boolean).alias(f"{left_type}_vs_{right_type}_same_isin")
            for left_type, right_type in LOOKUP_IDENTIFIER_PAIRS
        ],
        *[
            pl.lit(None, dtype=pl.Boolean).alias(f"{left_type}_vs_{right_type}_same_cusip")
            for left_type, right_type in LOOKUP_IDENTIFIER_PAIRS
        ],
    )
    extended = extended.with_columns(
        pl.lit(None, dtype=pl.Boolean).alias("all_successful_attempts_consistent")
    ).select(RIC_LOOKUP_EXTENDED_COLUMNS)

    summary_rows: list[dict[str, Any]] = []
    attempt_counts_by_identifier_type: dict[str, int] = {}
    success_counts_by_identifier_type: dict[str, int] = {}
    agreement_counts: dict[str, int] = {}
    conflict_counts: dict[str, int] = {}

    for identifier_type in LOOKUP_IDENTIFIER_TYPES:
        attempted_count = int(
            extended.select(pl.col(f"{identifier_type}_attempted").cast(pl.Int64).sum()).item()
        )
        success_count = int(
            extended.select(pl.col(f"{identifier_type}_success").cast(pl.Int64).sum()).item()
        )
        attempt_counts_by_identifier_type[identifier_type] = attempted_count
        success_counts_by_identifier_type[identifier_type] = success_count
        summary_rows.extend(
            (
                {
                    "summary_category": "attempt_count_by_identifier_type",
                    "summary_key": identifier_type,
                    "value": attempted_count,
                },
                {
                    "summary_category": "success_count_by_identifier_type",
                    "summary_key": identifier_type,
                    "value": success_count,
                },
            )
        )

    for left_type, right_type in LOOKUP_IDENTIFIER_PAIRS:
        pair_name = f"{left_type}_vs_{right_type}"
        for field_name in ("ric", "isin", "cusip"):
            column_name = f"{pair_name}_same_{field_name}"
            agreement_count = int(
                extended.select(pl.col(column_name).fill_null(False).cast(pl.Int64).sum()).item()
            )
            conflict_count = int(
                extended.select(
                    pl.when(pl.col(column_name).is_null())
                    .then(pl.lit(0, dtype=pl.Int64))
                    .otherwise((~pl.col(column_name)).cast(pl.Int64))
                    .sum()
                ).item()
            )
            agreement_key = f"{pair_name}_same_{field_name}"
            conflict_key = f"{pair_name}_same_{field_name}"
            agreement_counts[agreement_key] = agreement_count
            conflict_counts[conflict_key] = conflict_count
            summary_rows.extend(
                (
                    {
                        "summary_category": f"agreement_count_same_{field_name}",
                        "summary_key": pair_name,
                        "value": agreement_count,
                    },
                    {
                        "summary_category": f"conflict_count_same_{field_name}",
                        "summary_key": pair_name,
                        "value": conflict_count,
                    },
                )
            )

    all_consistent_true = int(
        extended.select(pl.col("all_successful_attempts_consistent").fill_null(False).cast(pl.Int64).sum()).item()
    )
    all_consistent_false = int(
        extended.select(
            pl.when(pl.col("all_successful_attempts_consistent").is_null())
            .then(pl.lit(0, dtype=pl.Int64))
            .otherwise((~pl.col("all_successful_attempts_consistent")).cast(pl.Int64))
            .sum()
        ).item()
    )
    agreement_counts["all_successful_attempts_consistent"] = all_consistent_true
    conflict_counts["all_successful_attempts_consistent"] = all_consistent_false
    summary_rows.extend(
        (
            {
                "summary_category": "agreement_count_all_successful_attempts_consistent",
                "summary_key": "all_successful_attempts_consistent",
                "value": all_consistent_true,
            },
            {
                "summary_category": "conflict_count_all_successful_attempts_consistent",
                "summary_key": "all_successful_attempts_consistent",
                "value": all_consistent_false,
            },
        )
    )

    single_success_counts: dict[str, int] = {}
    for identifier_type in LOOKUP_IDENTIFIER_TYPES:
        only_this_success_count = int(
            extended.select(
                (
                    pl.col(f"{identifier_type}_success")
                    & pl.all_horizontal(
                        [
                            (~pl.col(f"{other_identifier_type}_success"))
                            for other_identifier_type in LOOKUP_IDENTIFIER_TYPES
                            if other_identifier_type != identifier_type
                        ]
                    )
                )
                .cast(pl.Int64)
                .sum()
            ).item()
        )
        single_success_counts[identifier_type] = only_this_success_count
        summary_rows.append(
            {
                "summary_category": "only_one_identifier_type_succeeds",
                "summary_key": identifier_type,
                "value": only_this_success_count,
            }
        )

    rows_where_only_one_identifier_type_succeeds = int(sum(single_success_counts.values()))
    summary_rows.append(
        {
            "summary_category": "only_one_identifier_type_succeeds",
            "summary_key": "total",
            "value": rows_where_only_one_identifier_type_succeeds,
        }
    )

    summary_df = pl.DataFrame(summary_rows).select(
        pl.col("summary_category").cast(pl.Utf8, strict=False),
        pl.col("summary_key").cast(pl.Utf8, strict=False),
        pl.col("value").cast(pl.Int64, strict=False),
    )
    summary_payload = {
        "attempt_counts_by_identifier_type": attempt_counts_by_identifier_type,
        "success_counts_by_identifier_type": success_counts_by_identifier_type,
        "agreement_counts": agreement_counts,
        "conflict_counts": conflict_counts,
        "rows_where_only_one_identifier_type_succeeds": rows_where_only_one_identifier_type_succeeds,
    }
    return extended, summary_df, summary_payload


def _write_json(out_path: Path, payload: dict[str, Any]) -> Path:
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(payload, indent=2, sort_keys=True), encoding="utf-8")
    return out_path


def _write_workbook_or_reuse_locked_output(
    writer: Callable[..., Path],
    df: pl.DataFrame,
    output_path: Path,
    /,
    **writer_kwargs: Any,
) -> Path:
    output_path = Path(output_path)
    try:
        return writer(df, output_path, **writer_kwargs)
    except FileCreateError as exc:
        if output_path.exists():
            print(
                {
                    "warning": "reusing existing locked workbook output",
                    "path": str(output_path),
                    "reason": str(exc),
                }
            )
            return output_path
        raise


def run_refinitiv_step1_bridge_pipeline(
    daily_lf: pl.LazyFrame,
    output_dir: Path,
    *,
    company_description_lf: pl.LazyFrame | None = None,
    source_daily_path: Path | None = None,
) -> dict[str, Path]:
    """Write the compact Refinitiv step-1 bridge artifacts."""
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    bridge_df = build_refinitiv_step1_bridge_universe(
        daily_lf,
        company_description_lf=company_description_lf,
    ).collect()

    parquet_path = output_dir / "refinitiv_bridge_universe.parquet"
    extended_profile_path = _extended_profile_output_path(output_dir, "common_stock")
    extended_snapshot_path = output_dir / "refinitiv_ric_lookup_handoff_common_stock_extended_snapshot.parquet"
    manifest_path = output_dir / "refinitiv_step1_manifest.json"

    bridge_df.write_parquet(parquet_path, compression="zstd")

    source_rows = int(bridge_df["n_daily_rows"].sum()) if bridge_df.height else 0
    bridge_rows = bridge_df.height
    distinct_permno = bridge_df["KYPERMNO"].n_unique() if bridge_df.height else 0
    with_vendor_id = (
        bridge_df.select(
            (
                pl.col("CUSIP").is_not_null()
                | pl.col("ISIN").is_not_null()
                | pl.col("TICKER").is_not_null()
            ).sum()
        ).item()
        if bridge_df.height
        else 0
    )
    generated_at_utc = dt.datetime.now(dt.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

    common_stock_profile = next(
        profile for profile in RIC_LOOKUP_FILTER_PROFILES if profile.name == "common_stock"
    )
    qualifying_bridge_ids = _build_lookup_profile_bridge_ids(daily_lf, common_stock_profile)
    filtered_lookup_df, filtered_manual_review_df, profile_summary = _build_filtered_ric_lookup_profile_artifact(
        bridge_df,
        qualifying_bridge_ids,
        common_stock_profile,
    )
    extended_df, extended_summary_df, extended_summary_payload = (
        build_refinitiv_lookup_extended_diagnostic_artifact(
            filtered_lookup_df,
            filtered_manual_review_df,
        )
    )

    manifest_payload: dict[str, Any] = {
        "pipeline_name": "refinitiv_step1_bridge",
        "artifact_version": "v1",
        "generated_at_utc": generated_at_utc,
        "source_daily_path": str(source_daily_path) if source_daily_path is not None else None,
        "bridge_rows": bridge_rows,
        "source_daily_rows": source_rows,
        "distinct_permno": int(distinct_permno),
        "rows_with_vendor_identifier": int(with_vendor_id),
        "rows_missing_vendor_identifier": int(bridge_rows - with_vendor_id),
        "ric_lookup_rows": int(filtered_lookup_df.height),
        "ric_manual_review_rows": int(filtered_manual_review_df.height),
        "authoritative_format": "parquet",
        "source_columns": list(BRIDGE_SOURCE_COLUMNS),
        "vendor_columns": list(BRIDGE_VENDOR_COLUMNS),
        "extended_lookup_columns": list(RIC_LOOKUP_EXTENDED_COLUMNS),
        "common_stock_lookup_profile": profile_summary,
        "extended_lookup_diagnostic_summary": extended_summary_payload,
        "artifacts": {
            "refinitiv_bridge_universe_parquet": str(parquet_path),
            "refinitiv_ric_lookup_handoff_common_stock_extended_xlsx": str(extended_profile_path),
            "refinitiv_step1_manifest": str(manifest_path),
        },
    }

    _write_workbook_or_reuse_locked_output(
        write_refinitiv_ric_lookup_extended_workbook,
        extended_df,
        extended_profile_path,
        readme_payload=manifest_payload,
        text_columns=RIC_LOOKUP_EXTENDED_TEXT_COLUMNS,
        summary_df=extended_summary_df,
        summary_text_columns=RIC_LOOKUP_EXTENDED_SUMMARY_TEXT_COLUMNS,
    )
    extended_df.write_parquet(extended_snapshot_path, compression="zstd")
    _write_json(manifest_path, manifest_payload)

    return {
        "refinitiv_bridge_universe_parquet": parquet_path,
        "refinitiv_ric_lookup_handoff_common_stock_extended_xlsx": extended_profile_path,
        "refinitiv_step1_manifest": manifest_path,
}


def _normalize_workbook_scalar(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    text = str(value).strip()
    return text or None


def _normalize_lookup_text(value: str | None) -> str | None:
    if value is None:
        return None
    normalized = str(value).strip()
    return normalized or None


def _read_refinitiv_ric_lookup_sheet(
    workbook_path: Path | str,
    *,
    sheet_name: str = "ric_lookup",
) -> pl.DataFrame:
    from openpyxl import load_workbook

    workbook = load_workbook(Path(workbook_path), read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"lookup workbook missing sheet: {sheet_name}")
        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            raise ValueError(f"lookup workbook sheet {sheet_name!r} is empty")
        column_names = tuple(_normalize_workbook_scalar(name) for name in header)
        if any(name is None for name in column_names):
            raise ValueError(f"lookup workbook sheet {sheet_name!r} has blank headers")
        missing = [name for name in RIC_LOOKUP_COLUMNS if name not in column_names]
        if missing:
            raise ValueError(f"lookup workbook sheet {sheet_name!r} missing required columns: {missing}")

        records: list[dict[str, str | None]] = []
        for row in rows:
            record = {
                str(column_names[idx]): _normalize_workbook_scalar(row[idx]) if idx < len(row) else None
                for idx in range(len(column_names))
            }
            records.append(record)
    finally:
        workbook.close()

    if not records:
        return pl.DataFrame(schema={name: pl.Utf8 for name in RIC_LOOKUP_COLUMNS}).with_columns(
            pl.col("first_seen_caldt").cast(pl.Date, strict=False),
            pl.col("last_seen_caldt").cast(pl.Date, strict=False),
        )

    df = pl.DataFrame(records).select(
        [
            pl.col(name).cast(pl.Utf8, strict=False).alias(name)
            if name not in ("first_seen_caldt", "last_seen_caldt")
            else pl.col(name).cast(pl.Utf8, strict=False).str.strptime(pl.Date, strict=False).alias(name)
            for name in RIC_LOOKUP_COLUMNS
        ]
    )
    return df


def _normalize_extended_workbook_value(column_name: str, value: Any) -> str | bool | None:
    if column_name in RIC_LOOKUP_EXTENDED_BOOLEAN_COLUMNS:
        if value is None:
            return None
        if isinstance(value, bool):
            return value
        normalized = str(value).strip().lower()
        if normalized == "true":
            return True
        if normalized == "false":
            return False
        return None
    return _normalize_workbook_scalar(value)


def _read_refinitiv_ric_lookup_extended_sheet(
    workbook_path: Path | str,
    *,
    sheet_name: str = "lookup_diagnostics",
) -> pl.DataFrame:
    from openpyxl import load_workbook

    workbook = load_workbook(Path(workbook_path), read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"lookup workbook missing sheet: {sheet_name}")
        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            raise ValueError(f"lookup workbook sheet {sheet_name!r} is empty")
        column_names = tuple(_normalize_workbook_scalar(name) for name in header)
        if any(name is None for name in column_names):
            raise ValueError(f"lookup workbook sheet {sheet_name!r} has blank headers")
        missing = [name for name in RIC_LOOKUP_EXTENDED_COLUMNS if name not in column_names]
        if missing:
            raise ValueError(f"lookup workbook sheet {sheet_name!r} missing required columns: {missing}")

        records: list[dict[str, str | bool | None]] = []
        for row in rows:
            record = {
                str(column_names[idx]): (
                    _normalize_extended_workbook_value(str(column_names[idx]), row[idx]) if idx < len(row) else None
                )
                for idx in range(len(column_names))
            }
            records.append(record)
    finally:
        workbook.close()

    if not records:
        return _cast_df_to_schema(
            pl.DataFrame(schema={name: dtype for name, dtype in _extended_lookup_schema().items()}),
            _extended_lookup_schema(),
        )

    return _cast_df_to_schema(pl.DataFrame(records), _extended_lookup_schema()).select(
        RIC_LOOKUP_EXTENDED_COLUMNS
    )


def _normalized_lookup_result_value(value: Any) -> str | None:
    normalized = _normalize_lookup_text(value)
    if normalized is None:
        return None
    if normalized.upper() in LOOKUP_FAILURE_MARKERS:
        return None
    if "invalid identifier" in normalized.lower():
        return None
    return normalized


def _parse_bridge_row_id_liid(bridge_row_id: str | None) -> str | None:
    normalized = _normalize_lookup_text(bridge_row_id)
    if normalized is None:
        return None
    parts = normalized.split(":", 4)
    if len(parts) != 5:
        raise ValueError(f"bridge_row_id has unexpected format: {normalized!r}")
    return None if parts[1] == "-" else parts[1]


def _lookup_candidate_from_record(record: dict[str, Any], source: str) -> LookupIdentityCandidate | None:
    returned_ric = _normalized_lookup_result_value(record.get(f"{source}_returned_ric"))
    if returned_ric is None:
        return None
    return LookupIdentityCandidate(
        source=source,
        returned_ric=returned_ric,
        returned_isin=_normalized_lookup_result_value(record.get(f"{source}_returned_isin")),
        returned_cusip=_normalized_lookup_result_value(record.get(f"{source}_returned_cusip")),
    )


def _accepted_candidate_from_record(record: dict[str, Any]) -> LookupIdentityCandidate | None:
    accepted_ric = _normalized_lookup_result_value(record.get("accepted_ric"))
    if accepted_ric is None:
        return None
    return LookupIdentityCandidate(
        source=_normalize_lookup_text(record.get("accepted_ric_source")) or "CONVENTIONAL",
        returned_ric=accepted_ric,
        returned_isin=_normalized_lookup_result_value(record.get("accepted_identity_returned_isin")),
        returned_cusip=_normalized_lookup_result_value(record.get("accepted_identity_returned_cusip")),
    )


def _identity_candidates_agree(
    left: LookupIdentityCandidate,
    right: LookupIdentityCandidate,
) -> bool:
    if left.returned_isin is not None and right.returned_isin is not None and left.returned_isin != right.returned_isin:
        return False
    if (
        left.returned_cusip is not None
        and right.returned_cusip is not None
        and left.returned_cusip != right.returned_cusip
    ):
        return False
    if left.returned_isin is not None and right.returned_isin is not None:
        return True
    if left.returned_cusip is not None and right.returned_cusip is not None:
        return True
    if left.returned_isin is None and right.returned_isin is None and left.returned_cusip is None and right.returned_cusip is None:
        return left.returned_ric == right.returned_ric
    return False


def _merge_identity_fields(
    left: LookupIdentityCandidate,
    right: LookupIdentityCandidate | None = None,
) -> tuple[str | None, str | None]:
    if right is None:
        return left.returned_isin, left.returned_cusip
    return (
        left.returned_isin or right.returned_isin,
        left.returned_cusip or right.returned_cusip,
    )


def _candidates_materially_conflict(
    left: LookupIdentityCandidate,
    right: LookupIdentityCandidate,
) -> bool:
    return not _identity_candidates_agree(left, right)


def _derive_accepted_resolution(record: dict[str, Any]) -> dict[str, Any]:
    isin_candidate = _lookup_candidate_from_record(record, "ISIN")
    cusip_candidate = _lookup_candidate_from_record(record, "CUSIP")
    ticker_candidate = _lookup_candidate_from_record(record, "TICKER")

    accepted_identity_returned_isin: str | None = None
    accepted_identity_returned_cusip: str | None = None
    accepted_ric: str | None = None
    accepted_ric_source: str | None = None
    accepted_resolution_status = "unresolved_after_isin_cusip"
    conventional_identity_conflict = False

    if isin_candidate is not None and cusip_candidate is None:
        accepted_identity_returned_isin, accepted_identity_returned_cusip = _merge_identity_fields(
            isin_candidate
        )
        accepted_ric = isin_candidate.returned_ric
        accepted_ric_source = "ISIN"
        accepted_resolution_status = "resolved_from_isin"
    elif cusip_candidate is not None and isin_candidate is None:
        accepted_identity_returned_isin, accepted_identity_returned_cusip = _merge_identity_fields(
            cusip_candidate
        )
        accepted_ric = cusip_candidate.returned_ric
        accepted_ric_source = "CUSIP"
        accepted_resolution_status = "resolved_from_cusip"
    elif isin_candidate is not None and cusip_candidate is not None:
        if _identity_candidates_agree(isin_candidate, cusip_candidate):
            accepted_identity_returned_isin, accepted_identity_returned_cusip = _merge_identity_fields(
                isin_candidate,
                cusip_candidate,
            )
            accepted_ric = isin_candidate.returned_ric
            accepted_ric_source = "ISIN"
            accepted_resolution_status = "resolved_conventional_agree"
        else:
            conventional_identity_conflict = True
            accepted_resolution_status = "unresolved_conventional_conflict"

    accepted_candidate = None if accepted_ric is None else LookupIdentityCandidate(
        source=accepted_ric_source or "CONVENTIONAL",
        returned_ric=accepted_ric,
        returned_isin=accepted_identity_returned_isin,
        returned_cusip=accepted_identity_returned_cusip,
    )

    ticker_candidate_conflicts_with_conventional: bool | None
    if ticker_candidate is None or accepted_candidate is None:
        ticker_candidate_conflicts_with_conventional = None
    else:
        ticker_candidate_conflicts_with_conventional = _candidates_materially_conflict(
            accepted_candidate,
            ticker_candidate,
        )

    return {
        "accepted_identity_returned_isin": accepted_identity_returned_isin,
        "accepted_identity_returned_cusip": accepted_identity_returned_cusip,
        "accepted_ric": accepted_ric,
        "accepted_ric_source": accepted_ric_source,
        "accepted_resolution_status": accepted_resolution_status,
        "conventional_identity_conflict": conventional_identity_conflict,
        "ticker_candidate_ric": None if ticker_candidate is None else ticker_candidate.returned_ric,
        "ticker_candidate_available": ticker_candidate is not None,
        "ticker_candidate_conflicts_with_conventional": ticker_candidate_conflicts_with_conventional,
    }


def _accepted_source_matches_target(
    target_record: dict[str, Any],
    source_candidate: LookupIdentityCandidate,
) -> bool:
    target_isin = _normalize_lookup_text(target_record.get("ISIN"))
    target_cusip = _normalize_lookup_text(target_record.get("CUSIP"))
    if target_isin is not None:
        return source_candidate.returned_isin is not None and source_candidate.returned_isin == target_isin
    if target_cusip is not None:
        return source_candidate.returned_cusip is not None and source_candidate.returned_cusip == target_cusip
    return False


def _choose_adjacent_extension_source(
    prior_candidate: LookupIdentityCandidate,
    next_candidate: LookupIdentityCandidate,
    *,
    prior_record: dict[str, Any],
    next_record: dict[str, Any],
) -> tuple[dict[str, Any], str]:
    if prior_candidate.source != "ISIN" and next_candidate.source == "ISIN":
        return next_record, "ADJACENT"
    return prior_record, "ADJACENT"


def build_refinitiv_step1_resolution_frame(lookup_df: pl.DataFrame) -> pl.DataFrame:
    missing = [name for name in RIC_LOOKUP_EXTENDED_COLUMNS if name not in lookup_df.columns]
    if missing:
        raise ValueError(f"resolution input missing required columns: {missing}")

    normalized_df = _normalize_resolution_input_df(lookup_df)
    if normalized_df.height == 0:
        return _empty_resolution_df()

    records = normalized_df.to_dicts()
    grouped_records: dict[tuple[str | None, str | None], list[dict[str, Any]]] = {}

    for record in records:
        record.update(_derive_accepted_resolution(record))
        record["_liid"] = _parse_bridge_row_id_liid(_normalize_lookup_text(record.get("bridge_row_id")))
        grouped_records.setdefault(
            (
                _normalize_lookup_text(record.get("KYPERMNO")),
                record["_liid"],
            ),
            [],
        ).append(record)

    for group_records in grouped_records.values():
        ordered_records = sorted(
            group_records,
            key=lambda row: (
                row.get("first_seen_caldt") or dt.date.min,
                row.get("last_seen_caldt") or dt.date.min,
                _normalize_lookup_text(row.get("bridge_row_id")) or "",
            ),
        )
        for idx, record in enumerate(ordered_records):
            record["extended_ric"] = None
            record["extended_from_bridge_row_id"] = None
            record["extended_from_span_start"] = None
            record["extended_from_span_end"] = None
            record["extension_direction"] = None
            record["extension_status"] = "not_extended"

            if _normalize_lookup_text(record.get("accepted_ric")) is not None:
                continue
            if bool(record.get("conventional_identity_conflict")):
                record["extension_status"] = "not_extended_due_to_conflict"
                continue

            prior_record = ordered_records[idx - 1] if idx > 0 else None
            next_record = ordered_records[idx + 1] if idx + 1 < len(ordered_records) else None
            prior_candidate = None if prior_record is None else _accepted_candidate_from_record(prior_record)
            next_candidate = None if next_record is None else _accepted_candidate_from_record(next_record)
            target_has_raw_conventional = (
                _normalize_lookup_text(record.get("ISIN")) is not None
                or _normalize_lookup_text(record.get("CUSIP")) is not None
            )

            if prior_candidate is not None and next_candidate is not None and not _identity_candidates_agree(
                prior_candidate,
                next_candidate,
            ):
                record["extension_status"] = "not_extended_due_to_conflict"
                continue

            compatible_prior = (
                prior_candidate is not None
                and target_has_raw_conventional
                and _accepted_source_matches_target(record, prior_candidate)
            )
            compatible_next = (
                next_candidate is not None
                and target_has_raw_conventional
                and _accepted_source_matches_target(record, next_candidate)
            )

            chosen_source_record: dict[str, Any] | None = None
            extension_direction: str | None = None
            extension_status = "not_extended_no_adjacent_conventional_source"

            if not target_has_raw_conventional:
                if (
                    prior_candidate is not None
                    and next_candidate is not None
                    and _identity_candidates_agree(prior_candidate, next_candidate)
                ):
                    chosen_source_record, extension_direction = _choose_adjacent_extension_source(
                        prior_candidate,
                        next_candidate,
                        prior_record=prior_record,
                        next_record=next_record,
                    )
                    extension_status = "extended_from_adjacent_conventional_span"
            elif compatible_prior and compatible_next:
                chosen_source_record, extension_direction = _choose_adjacent_extension_source(
                    prior_candidate,
                    next_candidate,
                    prior_record=prior_record,
                    next_record=next_record,
                )
                extension_status = "extended_from_adjacent_conventional_span"
            elif compatible_prior:
                chosen_source_record = prior_record
                extension_direction = "PRIOR"
                extension_status = "extended_from_prior_conventional_span"
            elif compatible_next:
                chosen_source_record = next_record
                extension_direction = "NEXT"
                extension_status = "extended_from_next_conventional_span"

            if chosen_source_record is not None:
                record["extended_ric"] = chosen_source_record.get("accepted_ric")
                record["extended_from_bridge_row_id"] = chosen_source_record.get("bridge_row_id")
                record["extended_from_span_start"] = chosen_source_record.get("first_seen_caldt")
                record["extended_from_span_end"] = chosen_source_record.get("last_seen_caldt")
                record["extension_direction"] = extension_direction
                record["extension_status"] = extension_status
            else:
                record["extension_status"] = extension_status

    for record in records:
        accepted_ric = _normalize_lookup_text(record.get("accepted_ric"))
        extended_ric = _normalize_lookup_text(record.get("extended_ric"))
        if accepted_ric is not None:
            record["effective_collection_ric"] = accepted_ric
            record["effective_collection_ric_source"] = _normalize_lookup_text(record.get("accepted_ric_source"))
            record["effective_resolution_status"] = "effective_from_accepted_ric"
        elif extended_ric is not None:
            extension_direction = _normalize_lookup_text(record.get("extension_direction")) or "ADJACENT"
            chosen_source_source = None
            if _normalize_lookup_text(record.get("extended_from_bridge_row_id")) is not None:
                source_lookup = next(
                    (
                        candidate_record
                        for candidate_record in grouped_records.get(
                            (
                                _normalize_lookup_text(record.get("KYPERMNO")),
                                record.get("_liid"),
                            ),
                            [],
                        )
                        if candidate_record.get("bridge_row_id") == record.get("extended_from_bridge_row_id")
                    ),
                    None,
                )
                chosen_source_source = (
                    None if source_lookup is None else _normalize_lookup_text(source_lookup.get("accepted_ric_source"))
                )
            record["effective_collection_ric"] = extended_ric
            record["effective_collection_ric_source"] = (
                None
                if chosen_source_source is None
                else f"EXTENDED_FROM_{extension_direction}_{chosen_source_source}"
            )
            record["effective_resolution_status"] = "effective_from_extended_ric"
        else:
            record["effective_collection_ric"] = None
            record["effective_collection_ric_source"] = None
            record["effective_resolution_status"] = "unresolved_after_accept_and_extend"
        record.pop("_liid", None)

    resolution_df = _cast_df_to_schema(pl.DataFrame(records), _resolution_output_schema())
    return resolution_df.select(RIC_LOOKUP_RESOLUTION_OUTPUT_COLUMNS)


def _value_counts(records: list[dict[str, Any]], field_name: str) -> dict[str, int]:
    counts = Counter(
        _normalize_lookup_text(record.get(field_name))
        for record in records
        if _normalize_lookup_text(record.get(field_name)) is not None
    )
    return dict(sorted(counts.items()))


def _summarize_resolution_frame(
    resolution_df: pl.DataFrame,
    *,
    input_workbook_path: Path,
) -> dict[str, Any]:
    records = resolution_df.to_dicts()
    accepted_ric_rows = int(sum(1 for record in records if _normalize_lookup_text(record.get("accepted_ric")) is not None))
    extended_ric_rows = int(sum(1 for record in records if _normalize_lookup_text(record.get("extended_ric")) is not None))
    effective_ric_rows = int(
        sum(1 for record in records if _normalize_lookup_text(record.get("effective_collection_ric")) is not None)
    )
    ticker_only_candidates_without_effective = int(
        sum(
            1
            for record in records
            if bool(record.get("ticker_candidate_available"))
            and _normalize_lookup_text(record.get("accepted_ric")) is None
            and _normalize_lookup_text(record.get("extended_ric")) is None
            and _normalize_lookup_text(record.get("effective_collection_ric")) is None
        )
    )
    return {
        "pipeline_name": "refinitiv_step1_resolution",
        "artifact_version": "v1",
        "generated_at_utc": dt.datetime.now(dt.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "input_filled_lookup_workbook": str(input_workbook_path),
        "source_sheet_name": "lookup_diagnostics",
        "resolution_output_rows": int(resolution_df.height),
        "rows_with_accepted_ric": accepted_ric_rows,
        "rows_with_extended_ric": extended_ric_rows,
        "rows_with_effective_collection_ric": effective_ric_rows,
        "rows_unresolved_after_accept_and_extend": int(
            sum(
                1
                for record in records
                if _normalize_lookup_text(record.get("effective_resolution_status"))
                == "unresolved_after_accept_and_extend"
            )
        ),
        "rows_with_ticker_only_candidates_but_no_effective_collection_ric": ticker_only_candidates_without_effective,
        "rows_blocked_from_extension_due_conventional_conflicts": int(
            sum(
                1
                for record in records
                if _normalize_lookup_text(record.get("extension_status")) == "not_extended_due_to_conflict"
            )
        ),
        "accepted_resolution_status_counts": _value_counts(records, "accepted_resolution_status"),
        "extension_status_counts": _value_counts(records, "extension_status"),
        "effective_resolution_status_counts": _value_counts(records, "effective_resolution_status"),
        "extension_direction_counts": _value_counts(records, "extension_direction"),
        "effective_collection_ric_source_counts": _value_counts(records, "effective_collection_ric_source"),
        "ticker_candidate_available_rows": int(
            sum(1 for record in records if bool(record.get("ticker_candidate_available")))
        ),
        "ticker_candidate_conflicts_with_conventional_rows": int(
            sum(1 for record in records if record.get("ticker_candidate_conflicts_with_conventional") is True)
        ),
        "conventional_identity_conflict_rows": int(
            sum(1 for record in records if bool(record.get("conventional_identity_conflict")))
        ),
    }


def _read_refinitiv_ric_lookup_extended_artifact(
    artifact_path: Path | str,
) -> pl.DataFrame:
    artifact_path = Path(artifact_path)
    if artifact_path.suffix.lower() == ".parquet":
        return _cast_df_to_schema(
            pl.read_parquet(artifact_path).select(RIC_LOOKUP_EXTENDED_COLUMNS),
            _extended_lookup_schema(),
        ).select(RIC_LOOKUP_EXTENDED_COLUMNS)
    return _read_refinitiv_ric_lookup_extended_sheet(artifact_path)


def run_refinitiv_step1_resolution_pipeline(
    filled_lookup_workbook_path: Path | str,
    output_dir: Path | str,
) -> dict[str, Path]:
    filled_lookup_workbook_path = Path(filled_lookup_workbook_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    resolution_input_df = _read_refinitiv_ric_lookup_extended_artifact(filled_lookup_workbook_path)
    resolution_df = build_refinitiv_step1_resolution_frame(resolution_input_df)

    parquet_path = output_dir / "refinitiv_ric_resolution_common_stock.parquet"

    resolution_df.write_parquet(parquet_path, compression="zstd")

    return {
        "refinitiv_ric_resolution_common_stock_parquet": parquet_path,
    }


def _resolution_diagnostic_target_schema() -> dict[str, pl.DataType]:
    schema = _resolution_output_schema()
    schema.update(
        {
            "LIID_inferred": pl.Utf8,
            "logical_bridge_group_id": pl.Utf8,
            "diagnostic_case_id": pl.Utf8,
            "case_target_bridge_row_id": pl.Utf8,
            "target_class": pl.Utf8,
            "group_sequence_index": pl.Int64,
            "group_row_count": pl.Int64,
            "case_previous_row_available": pl.Boolean,
            "case_next_row_available": pl.Boolean,
            "case_any_adjacent_row_available": pl.Boolean,
            "case_previous_bridge_row_id": pl.Utf8,
            "case_next_bridge_row_id": pl.Utf8,
            "case_previous_effective_available": pl.Boolean,
            "case_next_effective_available": pl.Boolean,
            "case_any_adjacent_effective_available": pl.Boolean,
            "case_both_adjacent_effective_available": pl.Boolean,
            "case_previous_effective_collection_ric": pl.Utf8,
            "case_next_effective_collection_ric": pl.Utf8,
            "case_previous_effective_collection_ric_source": pl.Utf8,
            "case_next_effective_collection_ric_source": pl.Utf8,
            "case_previous_accepted_identity_returned_isin": pl.Utf8,
            "case_next_accepted_identity_returned_isin": pl.Utf8,
            "case_previous_accepted_identity_returned_cusip": pl.Utf8,
            "case_next_accepted_identity_returned_cusip": pl.Utf8,
            "isin_candidate_matches_case_previous_effective_ric": pl.Boolean,
            "isin_candidate_matches_case_next_effective_ric": pl.Boolean,
            "cusip_candidate_matches_case_previous_effective_ric": pl.Boolean,
            "cusip_candidate_matches_case_next_effective_ric": pl.Boolean,
            "ticker_candidate_matches_case_previous_effective_ric": pl.Boolean,
            "ticker_candidate_matches_case_next_effective_ric": pl.Boolean,
            "raw_isin_matches_case_previous_accepted_identity": pl.Boolean,
            "raw_isin_matches_case_next_accepted_identity": pl.Boolean,
            "raw_cusip_matches_case_previous_accepted_identity": pl.Boolean,
            "raw_cusip_matches_case_next_accepted_identity": pl.Boolean,
            "ticker_returned_isin_matches_case_previous_accepted_identity": pl.Boolean,
            "ticker_returned_isin_matches_case_next_accepted_identity": pl.Boolean,
            "ticker_returned_cusip_matches_case_previous_accepted_identity": pl.Boolean,
            "ticker_returned_cusip_matches_case_next_accepted_identity": pl.Boolean,
        }
    )
    return schema


def _resolution_diagnostic_context_schema() -> dict[str, pl.DataType]:
    schema = _resolution_diagnostic_target_schema()
    schema.update(
        {
            "diagnostic_role": pl.Utf8,
            "context_offset": pl.Int64,
        }
    )
    return schema


def _resolution_diagnostic_handoff_schema() -> dict[str, pl.DataType]:
    return {
        "diagnostic_case_id": pl.Utf8,
        "case_target_bridge_row_id": pl.Utf8,
        "target_class": pl.Utf8,
        "retrieval_sequence_index": pl.Int64,
        "retrieval_role": pl.Utf8,
        "diagnostic_role": pl.Utf8,
        "bridge_row_id": pl.Utf8,
        "KYPERMNO": pl.Utf8,
        "CUSIP": pl.Utf8,
        "ISIN": pl.Utf8,
        "TICKER": pl.Utf8,
        "lookup_input": pl.Utf8,
        "lookup_input_source": pl.Utf8,
        "request_start_date": pl.Utf8,
        "request_end_date": pl.Utf8,
        "effective_collection_ric": pl.Utf8,
        "effective_collection_ric_source": pl.Utf8,
        "accepted_ric": pl.Utf8,
        "accepted_ric_source": pl.Utf8,
        "ISIN_returned_ric": pl.Utf8,
        "CUSIP_returned_ric": pl.Utf8,
        "ticker_candidate_ric": pl.Utf8,
        "case_previous_effective_collection_ric": pl.Utf8,
        "case_next_effective_collection_ric": pl.Utf8,
    }


def _empty_resolution_diagnostic_targets_df() -> pl.DataFrame:
    schema = _resolution_diagnostic_target_schema()
    return _cast_df_to_schema(pl.DataFrame(schema={name: dtype for name, dtype in schema.items()}), schema)


def _empty_resolution_diagnostic_context_df() -> pl.DataFrame:
    schema = _resolution_diagnostic_context_schema()
    return _cast_df_to_schema(pl.DataFrame(schema={name: dtype for name, dtype in schema.items()}), schema)


def _empty_resolution_diagnostic_handoff_df() -> pl.DataFrame:
    schema = _resolution_diagnostic_handoff_schema()
    return _cast_df_to_schema(pl.DataFrame(schema={name: dtype for name, dtype in schema.items()}), schema)


def _resolution_diagnostic_target_class(record: dict[str, Any]) -> str | None:
    if bool(record.get("conventional_identity_conflict")):
        return "conventional_conflict"
    if (
        bool(record.get("ticker_candidate_available"))
        and _normalized_lookup_result_value(record.get("effective_collection_ric")) is None
    ):
        return "unresolved_ticker_only_candidate"
    return None


def _match_with_normalizer(
    left: Any,
    right: Any,
    *,
    normalizer: Callable[[Any], str | None],
) -> bool | None:
    left_normalized = normalizer(left)
    right_normalized = normalizer(right)
    if left_normalized is None or right_normalized is None:
        return None
    return left_normalized == right_normalized


def _count_true_records(records: list[dict[str, Any]], field_name: str) -> int:
    return int(sum(1 for record in records if record.get(field_name) is True))


def _resolution_diagnostic_class_summary(target_records: list[dict[str, Any]]) -> dict[str, dict[str, int]]:
    grouped: dict[str, list[dict[str, Any]]] = {}
    for record in target_records:
        target_class = _normalize_lookup_text(record.get("target_class"))
        if target_class is None:
            continue
        grouped.setdefault(target_class, []).append(record)

    summary: dict[str, dict[str, int]] = {}
    for target_class, records in sorted(grouped.items()):
        summary[target_class] = {
            "targets": int(len(records)),
            "isolated_targets": int(sum(1 for record in records if not bool(record.get("case_any_adjacent_row_available")))),
            "targets_with_previous_effective": int(
                sum(1 for record in records if bool(record.get("case_previous_effective_available")))
            ),
            "targets_with_next_effective": int(
                sum(1 for record in records if bool(record.get("case_next_effective_available")))
            ),
            "targets_with_any_adjacent_effective": int(
                sum(1 for record in records if bool(record.get("case_any_adjacent_effective_available")))
            ),
            "targets_with_both_adjacent_effective": int(
                sum(1 for record in records if bool(record.get("case_both_adjacent_effective_available")))
            ),
        }
    return summary


def _summarize_refinitiv_step1_resolution_diagnostic_artifacts(
    *,
    resolution_df: pl.DataFrame,
    targets_df: pl.DataFrame,
    context_df: pl.DataFrame,
    handoff_df: pl.DataFrame,
) -> dict[str, Any]:
    target_records = targets_df.to_dicts()
    context_records = context_df.to_dicts()
    handoff_records = handoff_df.to_dicts()

    candidate_match_fields = (
        "isin_candidate_matches_case_previous_effective_ric",
        "isin_candidate_matches_case_next_effective_ric",
        "cusip_candidate_matches_case_previous_effective_ric",
        "cusip_candidate_matches_case_next_effective_ric",
        "ticker_candidate_matches_case_previous_effective_ric",
        "ticker_candidate_matches_case_next_effective_ric",
    )
    raw_identity_match_fields = (
        "raw_isin_matches_case_previous_accepted_identity",
        "raw_isin_matches_case_next_accepted_identity",
        "raw_cusip_matches_case_previous_accepted_identity",
        "raw_cusip_matches_case_next_accepted_identity",
        "ticker_returned_isin_matches_case_previous_accepted_identity",
        "ticker_returned_isin_matches_case_next_accepted_identity",
        "ticker_returned_cusip_matches_case_previous_accepted_identity",
        "ticker_returned_cusip_matches_case_next_accepted_identity",
    )

    return {
        "source_resolution_rows": int(resolution_df.height),
        "diagnostic_case_count": int(targets_df.height),
        "target_row_count": int(targets_df.height),
        "context_row_count": int(context_df.height),
        "handoff_row_count": int(handoff_df.height),
        "target_class_counts": _value_counts(target_records, "target_class"),
        "diagnostic_role_counts": _value_counts(context_records, "diagnostic_role"),
        "handoff_retrieval_role_counts": _value_counts(handoff_records, "retrieval_role"),
        "handoff_lookup_input_source_counts": _value_counts(handoff_records, "lookup_input_source"),
        "targets_with_previous_row": _count_true_records(target_records, "case_previous_row_available"),
        "targets_with_next_row": _count_true_records(target_records, "case_next_row_available"),
        "targets_with_any_adjacent_row": _count_true_records(target_records, "case_any_adjacent_row_available"),
        "targets_with_previous_effective": _count_true_records(target_records, "case_previous_effective_available"),
        "targets_with_next_effective": _count_true_records(target_records, "case_next_effective_available"),
        "targets_with_any_adjacent_effective": _count_true_records(
            target_records,
            "case_any_adjacent_effective_available",
        ),
        "targets_with_both_adjacent_effective": _count_true_records(
            target_records,
            "case_both_adjacent_effective_available",
        ),
        "candidate_vs_adjacent_match_counts": {
            field_name: _count_true_records(target_records, field_name) for field_name in candidate_match_fields
        },
        "raw_identity_match_counts": {
            field_name: _count_true_records(target_records, field_name) for field_name in raw_identity_match_fields
        },
        "target_class_summary": _resolution_diagnostic_class_summary(target_records),
    }


def _append_resolution_diagnostic_handoff_row(
    handoff_rows: list[dict[str, Any]],
    *,
    case_record: dict[str, Any],
    source_record: dict[str, Any],
    retrieval_sequence_index: int,
    retrieval_role: str,
    diagnostic_role: str,
    lookup_input: Any,
    lookup_input_source: str,
) -> int:
    normalized_lookup_input = _normalized_lookup_result_value(lookup_input)
    if normalized_lookup_input is None:
        return retrieval_sequence_index

    handoff_rows.append(
        {
            "diagnostic_case_id": case_record["diagnostic_case_id"],
            "case_target_bridge_row_id": case_record["case_target_bridge_row_id"],
            "target_class": case_record["target_class"],
            "retrieval_sequence_index": retrieval_sequence_index,
            "retrieval_role": retrieval_role,
            "diagnostic_role": diagnostic_role,
            "bridge_row_id": _normalize_lookup_text(source_record.get("bridge_row_id")),
            "KYPERMNO": _normalize_lookup_text(source_record.get("KYPERMNO")),
            "CUSIP": _normalize_lookup_text(source_record.get("CUSIP")),
            "ISIN": _normalize_lookup_text(source_record.get("ISIN")),
            "TICKER": _normalize_lookup_text(source_record.get("TICKER")),
            "lookup_input": normalized_lookup_input,
            "lookup_input_source": lookup_input_source,
            "request_start_date": _date_to_text(source_record.get("first_seen_caldt")),
            "request_end_date": _date_to_text(source_record.get("last_seen_caldt")),
            "effective_collection_ric": _normalized_lookup_result_value(source_record.get("effective_collection_ric")),
            "effective_collection_ric_source": _normalize_lookup_text(source_record.get("effective_collection_ric_source")),
            "accepted_ric": _normalized_lookup_result_value(source_record.get("accepted_ric")),
            "accepted_ric_source": _normalize_lookup_text(source_record.get("accepted_ric_source")),
            "ISIN_returned_ric": _normalized_lookup_result_value(case_record.get("ISIN_returned_ric")),
            "CUSIP_returned_ric": _normalized_lookup_result_value(case_record.get("CUSIP_returned_ric")),
            "ticker_candidate_ric": _normalized_lookup_result_value(case_record.get("ticker_candidate_ric")),
            "case_previous_effective_collection_ric": _normalized_lookup_result_value(
                case_record.get("case_previous_effective_collection_ric")
            ),
            "case_next_effective_collection_ric": _normalized_lookup_result_value(
                case_record.get("case_next_effective_collection_ric")
            ),
        }
    )
    return retrieval_sequence_index + 1


def build_refinitiv_step1_resolution_diagnostic_artifacts(
    resolution_df: pl.DataFrame,
) -> tuple[pl.DataFrame, pl.DataFrame, pl.DataFrame, dict[str, Any]]:
    missing = [name for name in RIC_LOOKUP_RESOLUTION_OUTPUT_COLUMNS if name not in resolution_df.columns]
    if missing:
        raise ValueError(f"resolution diagnostic input missing required columns: {missing}")

    if resolution_df.height == 0:
        empty_targets = _empty_resolution_diagnostic_targets_df().select(RESOLUTION_DIAGNOSTIC_TARGET_COLUMNS)
        empty_context = _empty_resolution_diagnostic_context_df().select(RESOLUTION_DIAGNOSTIC_CONTEXT_COLUMNS)
        empty_handoff = _empty_resolution_diagnostic_handoff_df().select(RESOLUTION_DIAGNOSTIC_HANDOFF_COLUMNS)
        return (
            empty_targets,
            empty_context,
            empty_handoff,
            _summarize_refinitiv_step1_resolution_diagnostic_artifacts(
                resolution_df=resolution_df,
                targets_df=empty_targets,
                context_df=empty_context,
                handoff_df=empty_handoff,
            ),
        )

    records = resolution_df.to_dicts()
    grouped_records: dict[tuple[str | None, str | None], list[dict[str, Any]]] = {}
    for record in records:
        liid_inferred = _parse_bridge_row_id_liid(_normalize_lookup_text(record.get("bridge_row_id")))
        grouped_records.setdefault(
            (
                _normalize_lookup_text(record.get("KYPERMNO")),
                liid_inferred,
            ),
            [],
        ).append(record)

    target_rows: list[dict[str, Any]] = []
    context_rows: list[dict[str, Any]] = []
    handoff_rows: list[dict[str, Any]] = []

    for (kypermno, liid_inferred), group_records in grouped_records.items():
        ordered_records = sorted(
            group_records,
            key=lambda row: (
                row.get("first_seen_caldt") or dt.date.min,
                row.get("last_seen_caldt") or dt.date.min,
                _normalize_lookup_text(row.get("bridge_row_id")) or "",
            ),
        )
        logical_bridge_group_id = f"{kypermno or '-'}:{liid_inferred or '-'}"

        for idx, record in enumerate(ordered_records):
            target_class = _resolution_diagnostic_target_class(record)
            if target_class is None:
                continue

            prior_record = ordered_records[idx - 1] if idx > 0 else None
            next_record = ordered_records[idx + 1] if idx + 1 < len(ordered_records) else None

            case_previous_effective_ric = None if prior_record is None else _normalized_lookup_result_value(
                prior_record.get("effective_collection_ric")
            )
            case_next_effective_ric = None if next_record is None else _normalized_lookup_result_value(
                next_record.get("effective_collection_ric")
            )
            case_record_shared = {
                "LIID_inferred": liid_inferred,
                "logical_bridge_group_id": logical_bridge_group_id,
                "diagnostic_case_id": f"{target_class}:{_normalize_lookup_text(record.get('bridge_row_id')) or idx}",
                "case_target_bridge_row_id": _normalize_lookup_text(record.get("bridge_row_id")),
                "target_class": target_class,
                "group_row_count": len(ordered_records),
                "case_previous_row_available": prior_record is not None,
                "case_next_row_available": next_record is not None,
                "case_any_adjacent_row_available": prior_record is not None or next_record is not None,
                "case_previous_bridge_row_id": None if prior_record is None else _normalize_lookup_text(
                    prior_record.get("bridge_row_id")
                ),
                "case_next_bridge_row_id": None if next_record is None else _normalize_lookup_text(
                    next_record.get("bridge_row_id")
                ),
                "case_previous_effective_available": case_previous_effective_ric is not None,
                "case_next_effective_available": case_next_effective_ric is not None,
                "case_any_adjacent_effective_available": (
                    case_previous_effective_ric is not None or case_next_effective_ric is not None
                ),
                "case_both_adjacent_effective_available": (
                    case_previous_effective_ric is not None and case_next_effective_ric is not None
                ),
                "case_previous_effective_collection_ric": case_previous_effective_ric,
                "case_next_effective_collection_ric": case_next_effective_ric,
                "case_previous_effective_collection_ric_source": None
                if prior_record is None
                else _normalize_lookup_text(prior_record.get("effective_collection_ric_source")),
                "case_next_effective_collection_ric_source": None
                if next_record is None
                else _normalize_lookup_text(next_record.get("effective_collection_ric_source")),
                "case_previous_accepted_identity_returned_isin": None
                if prior_record is None
                else _normalized_lookup_result_value(prior_record.get("accepted_identity_returned_isin")),
                "case_next_accepted_identity_returned_isin": None
                if next_record is None
                else _normalized_lookup_result_value(next_record.get("accepted_identity_returned_isin")),
                "case_previous_accepted_identity_returned_cusip": None
                if prior_record is None
                else _normalized_lookup_result_value(prior_record.get("accepted_identity_returned_cusip")),
                "case_next_accepted_identity_returned_cusip": None
                if next_record is None
                else _normalized_lookup_result_value(next_record.get("accepted_identity_returned_cusip")),
                "isin_candidate_matches_case_previous_effective_ric": _match_with_normalizer(
                    record.get("ISIN_returned_ric"),
                    case_previous_effective_ric,
                    normalizer=_normalized_lookup_result_value,
                ),
                "isin_candidate_matches_case_next_effective_ric": _match_with_normalizer(
                    record.get("ISIN_returned_ric"),
                    case_next_effective_ric,
                    normalizer=_normalized_lookup_result_value,
                ),
                "cusip_candidate_matches_case_previous_effective_ric": _match_with_normalizer(
                    record.get("CUSIP_returned_ric"),
                    case_previous_effective_ric,
                    normalizer=_normalized_lookup_result_value,
                ),
                "cusip_candidate_matches_case_next_effective_ric": _match_with_normalizer(
                    record.get("CUSIP_returned_ric"),
                    case_next_effective_ric,
                    normalizer=_normalized_lookup_result_value,
                ),
                "ticker_candidate_matches_case_previous_effective_ric": _match_with_normalizer(
                    record.get("ticker_candidate_ric"),
                    case_previous_effective_ric,
                    normalizer=_normalized_lookup_result_value,
                ),
                "ticker_candidate_matches_case_next_effective_ric": _match_with_normalizer(
                    record.get("ticker_candidate_ric"),
                    case_next_effective_ric,
                    normalizer=_normalized_lookup_result_value,
                ),
                "raw_isin_matches_case_previous_accepted_identity": _match_with_normalizer(
                    record.get("ISIN"),
                    None if prior_record is None else prior_record.get("accepted_identity_returned_isin"),
                    normalizer=_normalize_lookup_text,
                ),
                "raw_isin_matches_case_next_accepted_identity": _match_with_normalizer(
                    record.get("ISIN"),
                    None if next_record is None else next_record.get("accepted_identity_returned_isin"),
                    normalizer=_normalize_lookup_text,
                ),
                "raw_cusip_matches_case_previous_accepted_identity": _match_with_normalizer(
                    record.get("CUSIP"),
                    None if prior_record is None else prior_record.get("accepted_identity_returned_cusip"),
                    normalizer=_normalize_lookup_text,
                ),
                "raw_cusip_matches_case_next_accepted_identity": _match_with_normalizer(
                    record.get("CUSIP"),
                    None if next_record is None else next_record.get("accepted_identity_returned_cusip"),
                    normalizer=_normalize_lookup_text,
                ),
                "ticker_returned_isin_matches_case_previous_accepted_identity": _match_with_normalizer(
                    record.get("TICKER_returned_isin"),
                    None if prior_record is None else prior_record.get("accepted_identity_returned_isin"),
                    normalizer=_normalized_lookup_result_value,
                ),
                "ticker_returned_isin_matches_case_next_accepted_identity": _match_with_normalizer(
                    record.get("TICKER_returned_isin"),
                    None if next_record is None else next_record.get("accepted_identity_returned_isin"),
                    normalizer=_normalized_lookup_result_value,
                ),
                "ticker_returned_cusip_matches_case_previous_accepted_identity": _match_with_normalizer(
                    record.get("TICKER_returned_cusip"),
                    None if prior_record is None else prior_record.get("accepted_identity_returned_cusip"),
                    normalizer=_normalized_lookup_result_value,
                ),
                "ticker_returned_cusip_matches_case_next_accepted_identity": _match_with_normalizer(
                    record.get("TICKER_returned_cusip"),
                    None if next_record is None else next_record.get("accepted_identity_returned_cusip"),
                    normalizer=_normalized_lookup_result_value,
                ),
            }

            target_record = {
                **{name: record.get(name) for name in RIC_LOOKUP_RESOLUTION_OUTPUT_COLUMNS},
                **case_record_shared,
                "group_sequence_index": idx + 1,
            }
            target_rows.append(target_record)

            for context_offset, diagnostic_role, context_record in (
                (-1, "PREVIOUS", prior_record),
                (0, "TARGET", record),
                (1, "NEXT", next_record),
            ):
                if context_record is None:
                    continue
                context_rows.append(
                    {
                        **{name: context_record.get(name) for name in RIC_LOOKUP_RESOLUTION_OUTPUT_COLUMNS},
                        **case_record_shared,
                        "diagnostic_role": diagnostic_role,
                        "context_offset": context_offset,
                        "group_sequence_index": ordered_records.index(context_record) + 1,
                    }
                )

            retrieval_sequence_index = 1
            if prior_record is not None:
                retrieval_sequence_index = _append_resolution_diagnostic_handoff_row(
                    handoff_rows,
                    case_record=target_record,
                    source_record=prior_record,
                    retrieval_sequence_index=retrieval_sequence_index,
                    retrieval_role="PREVIOUS_EFFECTIVE",
                    diagnostic_role="PREVIOUS",
                    lookup_input=prior_record.get("effective_collection_ric"),
                    lookup_input_source="effective_collection_ric",
                )
            retrieval_sequence_index = _append_resolution_diagnostic_handoff_row(
                handoff_rows,
                case_record=target_record,
                source_record=record,
                retrieval_sequence_index=retrieval_sequence_index,
                retrieval_role="TARGET_ISIN_CANDIDATE",
                diagnostic_role="TARGET",
                lookup_input=record.get("ISIN_returned_ric"),
                lookup_input_source="ISIN_returned_ric",
            )
            retrieval_sequence_index = _append_resolution_diagnostic_handoff_row(
                handoff_rows,
                case_record=target_record,
                source_record=record,
                retrieval_sequence_index=retrieval_sequence_index,
                retrieval_role="TARGET_CUSIP_CANDIDATE",
                diagnostic_role="TARGET",
                lookup_input=record.get("CUSIP_returned_ric"),
                lookup_input_source="CUSIP_returned_ric",
            )
            retrieval_sequence_index = _append_resolution_diagnostic_handoff_row(
                handoff_rows,
                case_record=target_record,
                source_record=record,
                retrieval_sequence_index=retrieval_sequence_index,
                retrieval_role="TARGET_TICKER_CANDIDATE",
                diagnostic_role="TARGET",
                lookup_input=record.get("ticker_candidate_ric"),
                lookup_input_source="ticker_candidate_ric",
            )
            if next_record is not None:
                _append_resolution_diagnostic_handoff_row(
                    handoff_rows,
                    case_record=target_record,
                    source_record=next_record,
                    retrieval_sequence_index=retrieval_sequence_index,
                    retrieval_role="NEXT_EFFECTIVE",
                    diagnostic_role="NEXT",
                    lookup_input=next_record.get("effective_collection_ric"),
                    lookup_input_source="effective_collection_ric",
                )

    targets_df = (
        _cast_df_to_schema(pl.DataFrame(target_rows), _resolution_diagnostic_target_schema()).select(
            RESOLUTION_DIAGNOSTIC_TARGET_COLUMNS
        )
        if target_rows
        else _empty_resolution_diagnostic_targets_df().select(RESOLUTION_DIAGNOSTIC_TARGET_COLUMNS)
    ).sort("diagnostic_case_id")

    context_df = (
        _cast_df_to_schema(pl.DataFrame(context_rows), _resolution_diagnostic_context_schema()).select(
            RESOLUTION_DIAGNOSTIC_CONTEXT_COLUMNS
        )
        if context_rows
        else _empty_resolution_diagnostic_context_df().select(RESOLUTION_DIAGNOSTIC_CONTEXT_COLUMNS)
    ).sort("diagnostic_case_id", "context_offset", "bridge_row_id")

    handoff_df = (
        _cast_df_to_schema(pl.DataFrame(handoff_rows), _resolution_diagnostic_handoff_schema()).select(
            RESOLUTION_DIAGNOSTIC_HANDOFF_COLUMNS
        )
        if handoff_rows
        else _empty_resolution_diagnostic_handoff_df().select(RESOLUTION_DIAGNOSTIC_HANDOFF_COLUMNS)
    ).sort("diagnostic_case_id", "retrieval_sequence_index", "bridge_row_id")

    summary = _summarize_refinitiv_step1_resolution_diagnostic_artifacts(
        resolution_df=resolution_df,
        targets_df=targets_df,
        context_df=context_df,
        handoff_df=handoff_df,
    )
    return targets_df, context_df, handoff_df, summary


def run_refinitiv_step1_resolution_diagnostic_pipeline(
    resolution_artifact_path: Path | str,
    output_dir: Path | str,
) -> dict[str, Path]:
    resolution_artifact_path = Path(resolution_artifact_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    resolution_df = pl.read_parquet(resolution_artifact_path)
    targets_df, context_df, handoff_df, summary = build_refinitiv_step1_resolution_diagnostic_artifacts(
        resolution_df
    )

    targets_parquet_path = output_dir / "refinitiv_ric_resolution_diagnostic_targets.parquet"
    targets_csv_path = output_dir / "refinitiv_ric_resolution_diagnostic_targets.csv"
    context_parquet_path = output_dir / "refinitiv_ric_resolution_diagnostic_context.parquet"
    context_csv_path = output_dir / "refinitiv_ric_resolution_diagnostic_context.csv"
    handoff_csv_path = output_dir / "refinitiv_ric_resolution_diagnostic_handoff.csv"
    handoff_xlsx_path = output_dir / "refinitiv_ric_resolution_diagnostic_handoff.xlsx"
    summary_path = output_dir / "refinitiv_ric_resolution_diagnostic_summary.json"
    manifest_path = output_dir / "refinitiv_ric_resolution_diagnostic_manifest.json"

    targets_df.write_parquet(targets_parquet_path, compression="zstd")
    targets_df.write_csv(targets_csv_path)
    context_df.write_parquet(context_parquet_path, compression="zstd")
    context_df.write_csv(context_csv_path)
    handoff_df.write_csv(handoff_csv_path)

    source_resolution_summary_path = resolution_artifact_path.with_name(
        "refinitiv_ric_resolution_common_stock_summary.json"
    )
    summary_payload: dict[str, Any] = {
        "pipeline_name": "refinitiv_step1_resolution_diagnostic_handoff",
        "artifact_version": "v1",
        "generated_at_utc": dt.datetime.now(dt.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "source_resolution_artifact": str(resolution_artifact_path),
        "source_resolution_summary_path": str(source_resolution_summary_path)
        if source_resolution_summary_path.exists()
        else None,
        "retrieval_handoff_input_row_order": [
            "diagnostic_case_id",
            "target_class",
            "retrieval_role",
            "diagnostic_role",
            "lookup_input_source",
            "case_target_bridge_row_id",
            "bridge_row_id",
            "lookup_input",
            "request_start_date",
            "request_end_date",
            "KYPERMNO",
            "TICKER",
            "CUSIP",
            "ISIN",
        ],
        **summary,
    }
    _write_json(summary_path, summary_payload)

    manifest_payload: dict[str, Any] = {
        **summary_payload,
        "target_columns": list(RESOLUTION_DIAGNOSTIC_TARGET_COLUMNS),
        "context_columns": list(RESOLUTION_DIAGNOSTIC_CONTEXT_COLUMNS),
        "handoff_columns": list(RESOLUTION_DIAGNOSTIC_HANDOFF_COLUMNS),
        "artifacts": {
            "refinitiv_ric_resolution_diagnostic_targets_parquet": str(targets_parquet_path),
            "refinitiv_ric_resolution_diagnostic_targets_csv": str(targets_csv_path),
            "refinitiv_ric_resolution_diagnostic_context_parquet": str(context_parquet_path),
            "refinitiv_ric_resolution_diagnostic_context_csv": str(context_csv_path),
            "refinitiv_ric_resolution_diagnostic_handoff_csv": str(handoff_csv_path),
            "refinitiv_ric_resolution_diagnostic_summary": str(summary_path),
            "refinitiv_ric_resolution_diagnostic_manifest": str(manifest_path),
        },
    }
    manifest_payload["artifacts"]["refinitiv_ric_resolution_diagnostic_handoff_xlsx"] = str(handoff_xlsx_path)

    write_refinitiv_resolution_diagnostic_workbook(
        _coerce_text_columns(context_df, RESOLUTION_DIAGNOSTIC_CONTEXT_TEXT_COLUMNS),
        _coerce_text_columns(targets_df, RESOLUTION_DIAGNOSTIC_TARGET_TEXT_COLUMNS),
        handoff_df,
        handoff_xlsx_path,
        readme_payload=manifest_payload,
        context_text_columns=RESOLUTION_DIAGNOSTIC_CONTEXT_TEXT_COLUMNS,
        target_text_columns=RESOLUTION_DIAGNOSTIC_TARGET_TEXT_COLUMNS,
    )
    _write_json(manifest_path, manifest_payload)

    return {
        "refinitiv_ric_resolution_diagnostic_targets_parquet": targets_parquet_path,
        "refinitiv_ric_resolution_diagnostic_targets_csv": targets_csv_path,
        "refinitiv_ric_resolution_diagnostic_context_parquet": context_parquet_path,
        "refinitiv_ric_resolution_diagnostic_context_csv": context_csv_path,
        "refinitiv_ric_resolution_diagnostic_handoff_csv": handoff_csv_path,
        "refinitiv_ric_resolution_diagnostic_handoff_xlsx": handoff_xlsx_path,
        "refinitiv_ric_resolution_diagnostic_summary": summary_path,
        "refinitiv_ric_resolution_diagnostic_manifest": manifest_path,
    }


def _ownership_validation_handoff_schema() -> dict[str, pl.DataType]:
    schema: dict[str, pl.DataType] = {}
    for name in OWNERSHIP_VALIDATION_HANDOFF_COLUMNS:
        schema[name] = pl.Int64 if name in OWNERSHIP_VALIDATION_HANDOFF_INT_COLUMNS else pl.Utf8
    return schema


def _ownership_validation_results_schema() -> dict[str, pl.DataType]:
    schema = _ownership_validation_handoff_schema()
    schema.update(
        {
            "returned_ric": pl.Utf8,
            "returned_date": pl.Date,
            "returned_category": pl.Utf8,
            "returned_value": pl.Float64,
        }
    )
    return schema


def _ownership_universe_handoff_schema() -> dict[str, pl.DataType]:
    schema = _resolution_output_schema()
    schema.update(
        {
            "diagnostic_case_id": pl.Utf8,
            "candidate_slot": pl.Utf8,
            "candidate_ric": pl.Utf8,
            "ownership_lookup_row_id": pl.Utf8,
            "ownership_lookup_role": pl.Utf8,
            "lookup_input": pl.Utf8,
            "lookup_input_source": pl.Utf8,
            "request_start_date": pl.Utf8,
            "request_end_date": pl.Utf8,
            "retrieval_eligible": pl.Boolean,
            "retrieval_exclusion_reason": pl.Utf8,
        }
    )
    return schema


def _ownership_universe_results_schema() -> dict[str, pl.DataType]:
    schema = _ownership_universe_handoff_schema()
    schema.update(
        {
            "returned_ric": pl.Utf8,
            "returned_date": pl.Date,
            "returned_category": pl.Utf8,
            "returned_value": pl.Float64,
        }
    )
    return schema


def _ownership_universe_row_summary_schema() -> dict[str, pl.DataType]:
    schema = _ownership_universe_handoff_schema()
    schema.update(
        {
            "retrieval_row_present": pl.Boolean,
            "ownership_rows_returned": pl.Int64,
            "ownership_first_date": pl.Date,
            "ownership_last_date": pl.Date,
            "ownership_distinct_categories": pl.Int64,
            "ownership_nonnull_value_count": pl.Int64,
            "ownership_single_returned_ric": pl.Boolean,
            "ownership_returned_ric_nunique": pl.Int64,
        }
    )
    return schema


def _ownership_validation_retrieval_summary_schema() -> dict[str, pl.DataType]:
    schema = _ownership_validation_handoff_schema()
    schema.update(
        {
            "ownership_rows_returned": pl.Int64,
            "ownership_first_date": pl.Date,
            "ownership_last_date": pl.Date,
            "ownership_distinct_categories": pl.Int64,
            "ownership_nonnull_value_count": pl.Int64,
            "ownership_single_returned_ric": pl.Boolean,
            "ownership_returned_ric_nunique": pl.Int64,
        }
    )
    return schema


def _ownership_validation_pairwise_schema() -> dict[str, pl.DataType]:
    schema: dict[str, pl.DataType] = {}
    for name in OWNERSHIP_VALIDATION_PAIRWISE_COMPARISON_COLUMNS:
        if name in {
            "overlap_date_count",
            "overlap_category_count",
            "overlap_date_category_pair_count",
            "matched_value_pair_count",
        }:
            schema[name] = pl.Int64
        elif name in {
            "same_returned_ric",
            "same_category_set_on_overlap",
            "pair_has_useful_overlap",
            "pair_supports_corrobation",
            "pair_supports_same_identity_ric_variant",
            "pair_conflicts",
        }:
            schema[name] = pl.Boolean
        elif name in {"mean_abs_value_diff", "median_abs_value_diff", "max_abs_value_diff"}:
            schema[name] = pl.Float64
        else:
            schema[name] = pl.Utf8
    return schema


def _ownership_validation_case_summary_schema() -> dict[str, pl.DataType]:
    schema: dict[str, pl.DataType] = {}
    for name in OWNERSHIP_VALIDATION_CASE_SUMMARY_COLUMNS:
        if name in {
            "candidate_has_ownership_data",
            "any_adjacent_effective_has_ownership_data",
            "candidate_matches_previous_effective_ownership",
            "candidate_matches_next_effective_ownership",
            "candidate_matches_any_adjacent_effective_ownership",
        }:
            schema[name] = pl.Boolean
        elif name in {
            "candidate_retrieval_rows_with_data",
            "adjacent_effective_retrieval_rows_with_data",
            "pair_supports_corrobation_count",
            "pair_supports_same_identity_ric_variant_count",
            "pair_conflicts_count",
        }:
            schema[name] = pl.Int64
        else:
            schema[name] = pl.Utf8
    return schema


def _empty_ownership_validation_handoff_df() -> pl.DataFrame:
    schema = _ownership_validation_handoff_schema()
    return _cast_df_to_schema(pl.DataFrame(schema={name: dtype for name, dtype in schema.items()}), schema)


def _empty_ownership_validation_results_df() -> pl.DataFrame:
    schema = _ownership_validation_results_schema()
    return _cast_df_to_schema(pl.DataFrame(schema={name: dtype for name, dtype in schema.items()}), schema)


def _empty_ownership_universe_handoff_df() -> pl.DataFrame:
    schema = _ownership_universe_handoff_schema()
    return _cast_df_to_schema(pl.DataFrame(schema={name: dtype for name, dtype in schema.items()}), schema)


def _empty_ownership_universe_results_df() -> pl.DataFrame:
    schema = _ownership_universe_results_schema()
    return _cast_df_to_schema(pl.DataFrame(schema={name: dtype for name, dtype in schema.items()}), schema)


def _empty_ownership_universe_row_summary_df() -> pl.DataFrame:
    schema = _ownership_universe_row_summary_schema()
    return _cast_df_to_schema(pl.DataFrame(schema={name: dtype for name, dtype in schema.items()}), schema)


def _empty_ownership_validation_retrieval_summary_df() -> pl.DataFrame:
    schema = _ownership_validation_retrieval_summary_schema()
    return _cast_df_to_schema(pl.DataFrame(schema={name: dtype for name, dtype in schema.items()}), schema)


def _empty_ownership_validation_pairwise_df() -> pl.DataFrame:
    schema = _ownership_validation_pairwise_schema()
    return _cast_df_to_schema(pl.DataFrame(schema={name: dtype for name, dtype in schema.items()}), schema)


def _empty_ownership_validation_case_summary_df() -> pl.DataFrame:
    schema = _ownership_validation_case_summary_schema()
    return _cast_df_to_schema(pl.DataFrame(schema={name: dtype for name, dtype in schema.items()}), schema)


def _ownership_validation_role_sort_key(retrieval_role: Any) -> tuple[int, str]:
    normalized = _normalize_lookup_text(retrieval_role)
    if normalized is None:
        return (len(OWNERSHIP_VALIDATION_RETRIEVAL_ROLE_ORDER), "")
    try:
        return (OWNERSHIP_VALIDATION_RETRIEVAL_ROLE_ORDER.index(normalized), normalized)
    except ValueError:
        return (len(OWNERSHIP_VALIDATION_RETRIEVAL_ROLE_ORDER), normalized)


def _read_resolution_diagnostic_handoff_csv(csv_path: Path | str) -> pl.DataFrame:
    csv_path = Path(csv_path)
    if not csv_path.exists():
        raise FileNotFoundError(f"resolution diagnostic handoff CSV not found: {csv_path}")
    df = pl.read_csv(csv_path)
    missing = [name for name in RESOLUTION_DIAGNOSTIC_HANDOFF_COLUMNS if name not in df.columns]
    if missing:
        raise ValueError(f"resolution diagnostic handoff CSV missing required columns: {missing}")
    schema = {
        name: (pl.Int64 if name == "retrieval_sequence_index" else pl.Utf8)
        for name in RESOLUTION_DIAGNOSTIC_HANDOFF_COLUMNS
    }
    return _cast_df_to_schema(df.select(RESOLUTION_DIAGNOSTIC_HANDOFF_COLUMNS), schema).select(
        RESOLUTION_DIAGNOSTIC_HANDOFF_COLUMNS
    )


def _summarize_refinitiv_step1_ownership_validation_handoff(
    handoff_df: pl.DataFrame,
) -> dict[str, Any]:
    records = handoff_df.to_dicts()
    case_rows = handoff_df.select("sheet_name", "diagnostic_case_id").unique().to_dicts()
    return {
        "diagnostic_case_count": int(handoff_df.select(pl.col("diagnostic_case_id").n_unique()).item())
        if handoff_df.height
        else 0,
        "retrieval_input_row_count": int(handoff_df.height),
        "target_class_counts": _value_counts(records, "target_class"),
        "retrieval_role_counts": _value_counts(records, "retrieval_role"),
        "lookup_input_source_counts": _value_counts(records, "lookup_input_source"),
        "sheet_case_counts": _value_counts(case_rows, "sheet_name"),
        "cases_per_sheet": OWNERSHIP_VALIDATION_CASES_PER_SHEET,
        "case_band_height": OWNERSHIP_VALIDATION_CASE_BAND_HEIGHT,
        "retrieval_role_order": list(OWNERSHIP_VALIDATION_RETRIEVAL_ROLE_ORDER),
    }


def build_refinitiv_step1_ownership_validation_handoff(
    resolution_diagnostic_handoff_df: pl.DataFrame,
) -> tuple[pl.DataFrame, dict[str, Any]]:
    missing = [name for name in RESOLUTION_DIAGNOSTIC_HANDOFF_COLUMNS if name not in resolution_diagnostic_handoff_df.columns]
    if missing:
        raise ValueError(f"ownership validation handoff input missing required columns: {missing}")

    if resolution_diagnostic_handoff_df.height == 0:
        empty_df = _empty_ownership_validation_handoff_df().select(OWNERSHIP_VALIDATION_HANDOFF_COLUMNS)
        return empty_df, _summarize_refinitiv_step1_ownership_validation_handoff(empty_df)

    case_groups: dict[str, list[dict[str, Any]]] = {}
    for record in resolution_diagnostic_handoff_df.select(RESOLUTION_DIAGNOSTIC_HANDOFF_COLUMNS).to_dicts():
        case_id = _normalize_lookup_text(record.get("diagnostic_case_id"))
        if case_id is None:
            continue
        case_groups.setdefault(case_id, []).append(record)

    rows: list[dict[str, Any]] = []
    ordered_case_ids = sorted(case_groups)
    for case_ordinal, diagnostic_case_id in enumerate(ordered_case_ids, start=1):
        case_records = sorted(
            case_groups[diagnostic_case_id],
            key=lambda row: (
                _ownership_validation_role_sort_key(row.get("retrieval_role")),
                int(row.get("retrieval_sequence_index") or 0),
                _normalize_lookup_text(row.get("bridge_row_id")) or "",
            ),
        )
        sheet_number = ((case_ordinal - 1) // OWNERSHIP_VALIDATION_CASES_PER_SHEET) + 1
        sheet_name = f"ownership_validation_{sheet_number:03d}"
        sheet_case_index = ((case_ordinal - 1) % OWNERSHIP_VALIDATION_CASES_PER_SHEET) + 1
        case_band_row_start = 1 + ((sheet_case_index - 1) * OWNERSHIP_VALIDATION_CASE_BAND_HEIGHT)
        case_band_row_end = case_band_row_start + OWNERSHIP_VALIDATION_CASE_BAND_HEIGHT - 1

        for record in case_records:
            retrieval_role = _normalize_lookup_text(record.get("retrieval_role"))
            block_slot_index = _ownership_validation_role_sort_key(retrieval_role)[0] + 1
            rows.append(
                {
                    **{name: record.get(name) for name in RESOLUTION_DIAGNOSTIC_HANDOFF_COLUMNS},
                    "sheet_name": sheet_name,
                    "sheet_case_index": sheet_case_index,
                    "case_band_row_start": case_band_row_start,
                    "case_band_row_end": case_band_row_end,
                    "block_slot_index": block_slot_index,
                    "block_slot_role": retrieval_role,
                }
            )

    handoff_df = (
        _cast_df_to_schema(pl.DataFrame(rows), _ownership_validation_handoff_schema())
        .select(OWNERSHIP_VALIDATION_HANDOFF_COLUMNS)
        .sort("sheet_name", "sheet_case_index", "block_slot_index", "retrieval_sequence_index", "bridge_row_id")
    )
    return handoff_df, _summarize_refinitiv_step1_ownership_validation_handoff(handoff_df)


def _normalize_ownership_result_text(value: Any) -> str | None:
    normalized = _normalize_workbook_scalar(value)
    if normalized is None:
        return None
    if normalized.upper() in LOOKUP_FAILURE_MARKERS:
        return None
    return normalized


def _normalize_ownership_result_date(value: Any) -> dt.date | None:
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    normalized = _normalize_workbook_scalar(value)
    if normalized is None:
        return None
    try:
        return dt.date.fromisoformat(normalized)
    except ValueError:
        return None


def _normalize_ownership_result_value(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    normalized = _normalize_workbook_scalar(value)
    if normalized is None:
        return None
    try:
        return float(normalized)
    except ValueError:
        return None


def _build_explicit_schema_df(
    rows: list[dict[str, Any]],
    schema: dict[str, pl.DataType],
) -> pl.DataFrame:
    if not rows:
        return _cast_df_to_schema(pl.DataFrame(schema={name: dtype for name, dtype in schema.items()}), schema)
    return pl.DataFrame(rows, schema=schema, strict=False).select(list(schema))


def _parse_refinitiv_ownership_validation_filled_workbook(
    filled_workbook_path: Path | str,
    handoff_df: pl.DataFrame,
) -> pl.DataFrame:
    from openpyxl import load_workbook

    if handoff_df.height == 0:
        return _empty_ownership_validation_results_df().select(OWNERSHIP_VALIDATION_RESULTS_COLUMNS)

    workbook = load_workbook(Path(filled_workbook_path), read_only=False, data_only=True)
    try:
        missing_sheets = sorted(
            {
                sheet_name
                for sheet_name in handoff_df.select("sheet_name").unique().get_column("sheet_name").to_list()
                if sheet_name not in workbook.sheetnames
            }
        )
        if missing_sheets:
            raise ValueError(f"filled ownership workbook missing sheets: {missing_sheets}")

        lookup_input_row_offset = OWNERSHIP_VALIDATION_VISIBLE_INPUT_FIELDS.index("lookup_input") + 1
        results: list[dict[str, Any]] = []
        for row in handoff_df.iter_rows(named=True):
            worksheet = workbook[str(row["sheet_name"])]
            case_band_row_start = int(row["case_band_row_start"])
            case_band_row_end = int(row["case_band_row_end"])
            block_slot_index = int(row["block_slot_index"])
            base_col = 2 + ((block_slot_index - 1) * 5)
            lookup_input_row = case_band_row_start + lookup_input_row_offset
            if _normalize_ownership_result_text(worksheet.cell(row=lookup_input_row, column=base_col).value) is None:
                continue

            any_results = False
            blank_streak = 0
            max_scan_row = min(case_band_row_end, worksheet.max_row)
            for excel_row in range(case_band_row_start + 1, max_scan_row + 1):
                returned_ric = _normalize_ownership_result_text(
                    worksheet.cell(row=excel_row, column=base_col + 1).value
                )
                returned_date = _normalize_ownership_result_date(
                    worksheet.cell(row=excel_row, column=base_col + 2).value
                )
                returned_value = _normalize_ownership_result_value(
                    worksheet.cell(row=excel_row, column=base_col + 3).value
                )
                returned_category = _normalize_ownership_result_text(
                    worksheet.cell(row=excel_row, column=base_col + 4).value
                )

                if (
                    returned_ric is None
                    and returned_date is None
                    and returned_value is None
                    and returned_category is None
                ):
                    blank_streak += 1
                    if any_results and blank_streak >= OWNERSHIP_VALIDATION_BLANK_RESULT_STREAK:
                        break
                    continue

                any_results = True
                blank_streak = 0
                results.append(
                    {
                        **{name: row.get(name) for name in OWNERSHIP_VALIDATION_HANDOFF_COLUMNS},
                        "returned_ric": returned_ric,
                        "returned_date": returned_date,
                        "returned_category": returned_category,
                        "returned_value": returned_value,
                    }
                )
    finally:
        workbook.close()

    if not results:
        return _empty_ownership_validation_results_df().select(OWNERSHIP_VALIDATION_RESULTS_COLUMNS)

    return (
        _build_explicit_schema_df(results, _ownership_validation_results_schema())
        .select(OWNERSHIP_VALIDATION_RESULTS_COLUMNS)
        .unique(
            subset=[
                "diagnostic_case_id",
                "retrieval_role",
                "returned_date",
                "returned_category",
                "returned_value",
                "returned_ric",
            ],
            maintain_order=True,
        )
    )


def build_refinitiv_ownership_validation_retrieval_summary(
    handoff_df: pl.DataFrame,
    results_df: pl.DataFrame,
) -> pl.DataFrame:
    if handoff_df.height == 0:
        return _empty_ownership_validation_retrieval_summary_df().select(OWNERSHIP_VALIDATION_RETRIEVAL_SUMMARY_COLUMNS)

    key_columns = list(OWNERSHIP_VALIDATION_HANDOFF_COLUMNS)
    base_df = handoff_df.select(key_columns).unique(subset=["diagnostic_case_id", "retrieval_role"], maintain_order=True)

    if results_df.height == 0:
        return (
            base_df.with_columns(
                pl.lit(0, dtype=pl.Int64).alias("ownership_rows_returned"),
                pl.lit(None, dtype=pl.Date).alias("ownership_first_date"),
                pl.lit(None, dtype=pl.Date).alias("ownership_last_date"),
                pl.lit(0, dtype=pl.Int64).alias("ownership_distinct_categories"),
                pl.lit(0, dtype=pl.Int64).alias("ownership_nonnull_value_count"),
                pl.lit(False, dtype=pl.Boolean).alias("ownership_single_returned_ric"),
                pl.lit(0, dtype=pl.Int64).alias("ownership_returned_ric_nunique"),
            )
            .select(OWNERSHIP_VALIDATION_RETRIEVAL_SUMMARY_COLUMNS)
        )

    agg_df = results_df.group_by(key_columns).agg(
        pl.len().alias("ownership_rows_returned"),
        pl.col("returned_date").drop_nulls().min().alias("ownership_first_date"),
        pl.col("returned_date").drop_nulls().max().alias("ownership_last_date"),
        pl.col("returned_category").drop_nulls().n_unique().cast(pl.Int64).alias("ownership_distinct_categories"),
        pl.col("returned_value").drop_nulls().len().cast(pl.Int64).alias("ownership_nonnull_value_count"),
        pl.col("returned_ric").drop_nulls().n_unique().cast(pl.Int64).alias("ownership_returned_ric_nunique"),
    )

    summary_df = (
        base_df.join(agg_df, on=key_columns, how="left")
        .with_columns(
            pl.col("ownership_rows_returned").fill_null(0).cast(pl.Int64),
            pl.col("ownership_distinct_categories").fill_null(0).cast(pl.Int64),
            pl.col("ownership_nonnull_value_count").fill_null(0).cast(pl.Int64),
            pl.col("ownership_returned_ric_nunique").fill_null(0).cast(pl.Int64),
        )
        .with_columns(
            (pl.col("ownership_returned_ric_nunique") == 1).fill_null(False).alias("ownership_single_returned_ric")
        )
        .select(OWNERSHIP_VALIDATION_RETRIEVAL_SUMMARY_COLUMNS)
    )
    return _cast_df_to_schema(summary_df, _ownership_validation_retrieval_summary_schema()).select(
        OWNERSHIP_VALIDATION_RETRIEVAL_SUMMARY_COLUMNS
    )


def _compare_ownership_result_frames(
    target_results: pl.DataFrame,
    adjacent_results: pl.DataFrame,
) -> dict[str, Any]:
    target_dates = {
        value for value in target_results.get_column("returned_date").to_list() if isinstance(value, dt.date)
    }
    adjacent_dates = {
        value for value in adjacent_results.get_column("returned_date").to_list() if isinstance(value, dt.date)
    }
    target_categories = {
        value for value in target_results.get_column("returned_category").to_list() if _normalize_lookup_text(value) is not None
    }
    adjacent_categories = {
        value
        for value in adjacent_results.get_column("returned_category").to_list()
        if _normalize_lookup_text(value) is not None
    }

    overlap_dates = target_dates & adjacent_dates
    overlap_categories = target_categories & adjacent_categories

    target_pairs_df = (
        target_results.select("returned_date", "returned_category", "returned_value")
        .filter(pl.col("returned_date").is_not_null() & pl.col("returned_category").is_not_null())
        .group_by("returned_date", "returned_category")
        .agg(pl.col("returned_value").drop_nulls().first().alias("target_value"))
    )
    adjacent_pairs_df = (
        adjacent_results.select("returned_date", "returned_category", "returned_value")
        .filter(pl.col("returned_date").is_not_null() & pl.col("returned_category").is_not_null())
        .group_by("returned_date", "returned_category")
        .agg(pl.col("returned_value").drop_nulls().first().alias("adjacent_value"))
    )
    overlap_pairs_df = target_pairs_df.join(adjacent_pairs_df, on=["returned_date", "returned_category"], how="inner")

    target_ric_set = {
        value for value in target_results.get_column("returned_ric").to_list() if _normalize_lookup_text(value) is not None
    }
    adjacent_ric_set = {
        value for value in adjacent_results.get_column("returned_ric").to_list() if _normalize_lookup_text(value) is not None
    }
    same_returned_ric: bool | None = None
    if target_ric_set and adjacent_ric_set:
        same_returned_ric = target_ric_set == adjacent_ric_set

    same_category_set_on_overlap: bool | None = None
    if overlap_dates:
        target_overlap_categories = {
            category
            for row_date, category in target_pairs_df.select("returned_date", "returned_category").iter_rows()
            if row_date in overlap_dates
        }
        adjacent_overlap_categories = {
            category
            for row_date, category in adjacent_pairs_df.select("returned_date", "returned_category").iter_rows()
            if row_date in overlap_dates
        }
        same_category_set_on_overlap = target_overlap_categories == adjacent_overlap_categories

    matched_value_pairs_df = overlap_pairs_df.filter(
        pl.col("target_value").is_not_null() & pl.col("adjacent_value").is_not_null()
    ).with_columns((pl.col("target_value") - pl.col("adjacent_value")).abs().alias("abs_value_diff"))

    matched_value_pair_count = int(matched_value_pairs_df.height)
    mean_abs_value_diff = (
        float(matched_value_pairs_df.select(pl.col("abs_value_diff").mean()).item())
        if matched_value_pair_count
        else None
    )
    median_abs_value_diff = (
        float(matched_value_pairs_df.select(pl.col("abs_value_diff").median()).item())
        if matched_value_pair_count
        else None
    )
    max_abs_value_diff = (
        float(matched_value_pairs_df.select(pl.col("abs_value_diff").max()).item())
        if matched_value_pair_count
        else None
    )

    pair_has_useful_overlap = (
        matched_value_pair_count >= OWNERSHIP_VALIDATION_USEFUL_OVERLAP_MIN_MATCHED_VALUE_PAIRS
        and len(overlap_dates) >= OWNERSHIP_VALIDATION_USEFUL_OVERLAP_MIN_DATES
        and len(overlap_categories) >= OWNERSHIP_VALIDATION_USEFUL_OVERLAP_MIN_CATEGORIES
    )
    pair_supports_corrobation = (
        pair_has_useful_overlap
        and same_returned_ric is True
        and same_category_set_on_overlap is True
        and median_abs_value_diff is not None
        and mean_abs_value_diff is not None
        and max_abs_value_diff is not None
        and median_abs_value_diff <= OWNERSHIP_VALIDATION_SUPPORT_MEDIAN_ABS_DIFF_MAX
        and mean_abs_value_diff <= OWNERSHIP_VALIDATION_SUPPORT_MEAN_ABS_DIFF_MAX
        and max_abs_value_diff <= OWNERSHIP_VALIDATION_SUPPORT_MAX_ABS_DIFF_MAX
    )
    pair_supports_same_identity_ric_variant = (
        pair_has_useful_overlap
        and same_returned_ric is False
        and same_category_set_on_overlap is True
        and median_abs_value_diff is not None
        and mean_abs_value_diff is not None
        and max_abs_value_diff is not None
        and median_abs_value_diff <= OWNERSHIP_VALIDATION_SUPPORT_MEDIAN_ABS_DIFF_MAX
        and mean_abs_value_diff <= OWNERSHIP_VALIDATION_SUPPORT_MEAN_ABS_DIFF_MAX
        and max_abs_value_diff <= OWNERSHIP_VALIDATION_SUPPORT_MAX_ABS_DIFF_MAX
    )
    pair_conflicts = bool(
        pair_has_useful_overlap
        and (
            same_category_set_on_overlap is False
            or (
                mean_abs_value_diff is not None
                and mean_abs_value_diff > OWNERSHIP_VALIDATION_CONFLICT_MEAN_ABS_DIFF_MIN
            )
            or (
                max_abs_value_diff is not None
                and max_abs_value_diff > OWNERSHIP_VALIDATION_CONFLICT_MAX_ABS_DIFF_MIN
            )
        )
    )

    return {
        "overlap_date_count": int(len(overlap_dates)),
        "overlap_category_count": int(len(overlap_categories)),
        "overlap_date_category_pair_count": int(overlap_pairs_df.height),
        "same_returned_ric": same_returned_ric,
        "same_category_set_on_overlap": same_category_set_on_overlap,
        "matched_value_pair_count": matched_value_pair_count,
        "mean_abs_value_diff": mean_abs_value_diff,
        "median_abs_value_diff": median_abs_value_diff,
        "max_abs_value_diff": max_abs_value_diff,
        "pair_has_useful_overlap": pair_has_useful_overlap,
        "pair_supports_corrobation": pair_supports_corrobation,
        "pair_supports_same_identity_ric_variant": pair_supports_same_identity_ric_variant,
        "pair_conflicts": pair_conflicts,
    }


def build_refinitiv_ownership_validation_pairwise_comparisons(
    handoff_df: pl.DataFrame,
    results_df: pl.DataFrame,
) -> pl.DataFrame:
    if handoff_df.height == 0:
        return _empty_ownership_validation_pairwise_df().select(OWNERSHIP_VALIDATION_PAIRWISE_COMPARISON_COLUMNS)

    results_by_key: dict[tuple[str, str], pl.DataFrame] = {}
    if results_df.height > 0:
        for diagnostic_case_id in sorted(
            results_df.select("diagnostic_case_id").unique().get_column("diagnostic_case_id").drop_nulls().to_list()
        ):
            case_results = results_df.filter(pl.col("diagnostic_case_id") == diagnostic_case_id)
            for retrieval_role in sorted(
                case_results.select("retrieval_role").unique().get_column("retrieval_role").drop_nulls().to_list()
            ):
                results_by_key[(diagnostic_case_id, retrieval_role)] = case_results.filter(
                    pl.col("retrieval_role") == retrieval_role
                )

    pair_rows: list[dict[str, Any]] = []
    for diagnostic_case_id in sorted(
        handoff_df.select("diagnostic_case_id").unique().get_column("diagnostic_case_id").drop_nulls().to_list()
    ):
        case_df = handoff_df.filter(pl.col("diagnostic_case_id") == diagnostic_case_id).sort(
            "block_slot_index", "retrieval_sequence_index"
        )
        case_rows = case_df.to_dicts()
        by_role = {
            _normalize_lookup_text(row.get("retrieval_role")): row
            for row in case_rows
            if _normalize_lookup_text(row.get("retrieval_role")) is not None
        }
        for target_role in ("TARGET_ISIN_CANDIDATE", "TARGET_CUSIP_CANDIDATE", "TARGET_TICKER_CANDIDATE"):
            target_row = by_role.get(target_role)
            if target_row is None:
                continue
            target_results = results_by_key.get(
                (diagnostic_case_id, target_role),
                _empty_ownership_validation_results_df().select(OWNERSHIP_VALIDATION_RESULTS_COLUMNS),
            )
            for adjacent_role, adjacent_direction in (
                ("PREVIOUS_EFFECTIVE", "PREVIOUS"),
                ("NEXT_EFFECTIVE", "NEXT"),
            ):
                adjacent_row = by_role.get(adjacent_role)
                if adjacent_row is None:
                    continue
                adjacent_results = results_by_key.get(
                    (diagnostic_case_id, adjacent_role),
                    _empty_ownership_validation_results_df().select(OWNERSHIP_VALIDATION_RESULTS_COLUMNS),
                )
                pair_rows.append(
                    {
                        "diagnostic_case_id": diagnostic_case_id,
                        "target_class": _normalize_lookup_text(target_row.get("target_class")),
                        "case_target_bridge_row_id": _normalize_lookup_text(target_row.get("case_target_bridge_row_id")),
                        "target_retrieval_role": target_role,
                        "target_lookup_input": _normalize_lookup_text(target_row.get("lookup_input")),
                        "target_lookup_input_source": _normalize_lookup_text(target_row.get("lookup_input_source")),
                        "target_bridge_row_id": _normalize_lookup_text(target_row.get("bridge_row_id")),
                        "adjacent_retrieval_role": adjacent_role,
                        "adjacent_lookup_input": _normalize_lookup_text(adjacent_row.get("lookup_input")),
                        "adjacent_lookup_input_source": _normalize_lookup_text(adjacent_row.get("lookup_input_source")),
                        "adjacent_bridge_row_id": _normalize_lookup_text(adjacent_row.get("bridge_row_id")),
                        "adjacent_direction": adjacent_direction,
                        **_compare_ownership_result_frames(target_results, adjacent_results),
                    }
                )

    if not pair_rows:
        return _empty_ownership_validation_pairwise_df().select(OWNERSHIP_VALIDATION_PAIRWISE_COMPARISON_COLUMNS)

    return _cast_df_to_schema(pl.DataFrame(pair_rows), _ownership_validation_pairwise_schema()).select(
        OWNERSHIP_VALIDATION_PAIRWISE_COMPARISON_COLUMNS
    )


def build_refinitiv_ownership_validation_case_summary(
    handoff_df: pl.DataFrame,
    retrieval_summary_df: pl.DataFrame,
    pairwise_df: pl.DataFrame,
) -> tuple[pl.DataFrame, dict[str, Any]]:
    if handoff_df.height == 0:
        empty_df = _empty_ownership_validation_case_summary_df().select(OWNERSHIP_VALIDATION_CASE_SUMMARY_COLUMNS)
        summary_payload = {
            "diagnostic_case_count": 0,
            "candidate_retrieval_input_rows": 0,
            "retrieval_rows_with_any_returned_data": 0,
            "diagnostic_cases_with_candidate_ownership_data": 0,
            "diagnostic_cases_matching_previous_effective_ownership": 0,
            "diagnostic_cases_matching_next_effective_ownership": 0,
            "diagnostic_cases_matching_any_adjacent_effective_ownership": 0,
            "ownership_validation_bucket_counts": {},
        }
        return empty_df, summary_payload

    handoff_records = handoff_df.to_dicts()
    retrieval_records = retrieval_summary_df.to_dicts()
    pair_records = pairwise_df.to_dicts()
    cases = sorted(
        handoff_df.select("diagnostic_case_id").unique().get_column("diagnostic_case_id").drop_nulls().to_list()
    )

    case_rows: list[dict[str, Any]] = []
    for diagnostic_case_id in cases:
        case_handoff_rows = [row for row in handoff_records if row.get("diagnostic_case_id") == diagnostic_case_id]
        case_retrieval_rows = [row for row in retrieval_records if row.get("diagnostic_case_id") == diagnostic_case_id]
        case_pair_rows = [row for row in pair_records if row.get("diagnostic_case_id") == diagnostic_case_id]

        candidate_rows = [row for row in case_retrieval_rows if str(row.get("retrieval_role", "")).startswith("TARGET_")]
        adjacent_rows = [
            row
            for row in case_retrieval_rows
            if row.get("retrieval_role") in {"PREVIOUS_EFFECTIVE", "NEXT_EFFECTIVE"}
        ]
        candidate_has_ownership_data = any(int(row.get("ownership_rows_returned") or 0) > 0 for row in candidate_rows)
        any_adjacent_effective_has_ownership_data = any(
            int(row.get("ownership_rows_returned") or 0) > 0 for row in adjacent_rows
        )
        candidate_matches_previous_effective_ownership = any(
            row.get("adjacent_direction") == "PREVIOUS"
            and (bool(row.get("pair_supports_corrobation")) or bool(row.get("pair_supports_same_identity_ric_variant")))
            for row in case_pair_rows
        )
        candidate_matches_next_effective_ownership = any(
            row.get("adjacent_direction") == "NEXT"
            and (bool(row.get("pair_supports_corrobation")) or bool(row.get("pair_supports_same_identity_ric_variant")))
            for row in case_pair_rows
        )
        candidate_matches_any_adjacent_effective_ownership = (
            candidate_matches_previous_effective_ownership or candidate_matches_next_effective_ownership
        )
        pair_supports_corrobation_count = int(sum(1 for row in case_pair_rows if bool(row.get("pair_supports_corrobation"))))
        pair_supports_same_identity_ric_variant_count = int(
            sum(1 for row in case_pair_rows if bool(row.get("pair_supports_same_identity_ric_variant")))
        )
        pair_conflicts_count = int(sum(1 for row in case_pair_rows if bool(row.get("pair_conflicts"))))

        if not candidate_has_ownership_data or not any_adjacent_effective_has_ownership_data:
            ownership_validation_bucket = "ownership_no_useful_data"
        elif pair_conflicts_count > 0:
            ownership_validation_bucket = "ownership_conflicts_with_adjacent_identity"
        elif pair_supports_corrobation_count > 0:
            ownership_validation_bucket = "ownership_corrobates_candidate"
        elif pair_supports_same_identity_ric_variant_count > 0:
            ownership_validation_bucket = "ownership_supports_same_identity_ric_variant"
        else:
            ownership_validation_bucket = "ownership_inconclusive_sparse"

        anchor_row = case_handoff_rows[0]
        case_rows.append(
            {
                "diagnostic_case_id": diagnostic_case_id,
                "target_class": _normalize_lookup_text(anchor_row.get("target_class")),
                "case_target_bridge_row_id": _normalize_lookup_text(anchor_row.get("case_target_bridge_row_id")),
                "candidate_has_ownership_data": candidate_has_ownership_data,
                "any_adjacent_effective_has_ownership_data": any_adjacent_effective_has_ownership_data,
                "candidate_matches_previous_effective_ownership": candidate_matches_previous_effective_ownership,
                "candidate_matches_next_effective_ownership": candidate_matches_next_effective_ownership,
                "candidate_matches_any_adjacent_effective_ownership": candidate_matches_any_adjacent_effective_ownership,
                "candidate_retrieval_rows_with_data": int(
                    sum(1 for row in candidate_rows if int(row.get("ownership_rows_returned") or 0) > 0)
                ),
                "adjacent_effective_retrieval_rows_with_data": int(
                    sum(1 for row in adjacent_rows if int(row.get("ownership_rows_returned") or 0) > 0)
                ),
                "pair_supports_corrobation_count": pair_supports_corrobation_count,
                "pair_supports_same_identity_ric_variant_count": pair_supports_same_identity_ric_variant_count,
                "pair_conflicts_count": pair_conflicts_count,
                "ownership_validation_bucket": ownership_validation_bucket,
            }
        )

    case_summary_df = _cast_df_to_schema(
        pl.DataFrame(case_rows),
        _ownership_validation_case_summary_schema(),
    ).select(OWNERSHIP_VALIDATION_CASE_SUMMARY_COLUMNS)

    summary_payload = {
        "diagnostic_case_count": int(case_summary_df.height),
        "candidate_retrieval_input_rows": int(
            retrieval_summary_df.filter(pl.col("retrieval_role").str.starts_with("TARGET_")).height
        ),
        "retrieval_rows_with_any_returned_data": int(
            retrieval_summary_df.filter(pl.col("ownership_rows_returned") > 0).height
        ),
        "diagnostic_cases_with_candidate_ownership_data": int(
            case_summary_df.filter(pl.col("candidate_has_ownership_data")).height
        ),
        "diagnostic_cases_matching_previous_effective_ownership": int(
            case_summary_df.filter(pl.col("candidate_matches_previous_effective_ownership")).height
        ),
        "diagnostic_cases_matching_next_effective_ownership": int(
            case_summary_df.filter(pl.col("candidate_matches_next_effective_ownership")).height
        ),
        "diagnostic_cases_matching_any_adjacent_effective_ownership": int(
            case_summary_df.filter(pl.col("candidate_matches_any_adjacent_effective_ownership")).height
        ),
        "ownership_validation_bucket_counts": _value_counts(case_summary_df.to_dicts(), "ownership_validation_bucket"),
    }
    return case_summary_df, summary_payload


def run_refinitiv_step1_ownership_validation_handoff_pipeline(
    resolution_diagnostic_handoff_csv_path: Path | str,
    output_dir: Path | str,
) -> dict[str, Path]:
    resolution_diagnostic_handoff_csv_path = Path(resolution_diagnostic_handoff_csv_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    resolution_handoff_df = _read_resolution_diagnostic_handoff_csv(resolution_diagnostic_handoff_csv_path)
    handoff_df, summary = build_refinitiv_step1_ownership_validation_handoff(resolution_handoff_df)

    handoff_parquet_path = output_dir / "refinitiv_ownership_validation_handoff_common_stock.parquet"
    handoff_csv_path = output_dir / "refinitiv_ownership_validation_handoff_common_stock.csv"
    handoff_xlsx_path = output_dir / "refinitiv_ownership_validation_handoff_common_stock.xlsx"
    summary_path = output_dir / "refinitiv_ownership_validation_handoff_common_stock_summary.json"
    manifest_path = output_dir / "refinitiv_ownership_validation_handoff_common_stock_manifest.json"

    handoff_df.write_parquet(handoff_parquet_path, compression="zstd")
    handoff_df.write_csv(handoff_csv_path)

    summary_payload: dict[str, Any] = {
        "pipeline_name": "refinitiv_step1_ownership_validation_handoff",
        "artifact_version": "v1",
        "generated_at_utc": dt.datetime.now(dt.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "source_resolution_diagnostic_handoff_csv": str(resolution_diagnostic_handoff_csv_path),
        "visible_input_field_order": list(OWNERSHIP_VALIDATION_VISIBLE_INPUT_FIELDS),
        **summary,
    }
    _write_json(summary_path, summary_payload)

    manifest_payload = {
        **summary_payload,
        "handoff_columns": list(OWNERSHIP_VALIDATION_HANDOFF_COLUMNS),
        "artifacts": {
            "refinitiv_ownership_validation_handoff_common_stock_parquet": str(handoff_parquet_path),
            "refinitiv_ownership_validation_handoff_common_stock_csv": str(handoff_csv_path),
            "refinitiv_ownership_validation_handoff_common_stock_xlsx": str(handoff_xlsx_path),
            "refinitiv_ownership_validation_handoff_common_stock_summary": str(summary_path),
            "refinitiv_ownership_validation_handoff_common_stock_manifest": str(manifest_path),
        },
    }
    write_refinitiv_ownership_validation_workbook(
        handoff_df,
        handoff_xlsx_path,
        readme_payload=manifest_payload,
        input_field_order=OWNERSHIP_VALIDATION_VISIBLE_INPUT_FIELDS,
        block_slot_roles=OWNERSHIP_VALIDATION_RETRIEVAL_ROLE_ORDER,
    )
    _write_json(manifest_path, manifest_payload)

    return {
        "refinitiv_ownership_validation_handoff_common_stock_parquet": handoff_parquet_path,
        "refinitiv_ownership_validation_handoff_common_stock_csv": handoff_csv_path,
        "refinitiv_ownership_validation_handoff_common_stock_xlsx": handoff_xlsx_path,
        "refinitiv_ownership_validation_handoff_common_stock_summary": summary_path,
        "refinitiv_ownership_validation_handoff_common_stock_manifest": manifest_path,
    }


def run_refinitiv_step1_ownership_validation_results_pipeline(
    filled_workbook_path: Path | str,
    output_dir: Path | str,
    *,
    handoff_csv_path: Path | str | None = None,
) -> dict[str, Path]:
    filled_workbook_path = Path(filled_workbook_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    if handoff_csv_path is None:
        handoff_csv_path = output_dir / "refinitiv_ownership_validation_handoff_common_stock.csv"
    handoff_csv_path = Path(handoff_csv_path)

    handoff_df = _cast_df_to_schema(
        pl.read_csv(handoff_csv_path).select(OWNERSHIP_VALIDATION_HANDOFF_COLUMNS),
        _ownership_validation_handoff_schema(),
    ).select(OWNERSHIP_VALIDATION_HANDOFF_COLUMNS)
    results_df = _parse_refinitiv_ownership_validation_filled_workbook(filled_workbook_path, handoff_df).select(
        OWNERSHIP_VALIDATION_RESULTS_COLUMNS
    )
    retrieval_summary_df = build_refinitiv_ownership_validation_retrieval_summary(handoff_df, results_df)
    pairwise_df = build_refinitiv_ownership_validation_pairwise_comparisons(handoff_df, results_df)
    case_summary_df, case_summary_payload = build_refinitiv_ownership_validation_case_summary(
        handoff_df,
        retrieval_summary_df,
        pairwise_df,
    )

    results_parquet_path = output_dir / "refinitiv_ownership_validation_results.parquet"
    results_csv_path = output_dir / "refinitiv_ownership_validation_results.csv"
    retrieval_summary_parquet_path = output_dir / "refinitiv_ownership_validation_retrieval_summary.parquet"
    retrieval_summary_csv_path = output_dir / "refinitiv_ownership_validation_retrieval_summary.csv"
    pairwise_parquet_path = output_dir / "refinitiv_ownership_validation_pairwise_comparisons.parquet"
    pairwise_csv_path = output_dir / "refinitiv_ownership_validation_pairwise_comparisons.csv"
    case_summary_parquet_path = output_dir / "refinitiv_ownership_validation_case_summary.parquet"
    case_summary_csv_path = output_dir / "refinitiv_ownership_validation_case_summary.csv"
    case_summary_json_path = output_dir / "refinitiv_ownership_validation_case_summary.json"
    manifest_path = output_dir / "refinitiv_ownership_validation_results_manifest.json"

    results_df.write_parquet(results_parquet_path, compression="zstd")
    results_df.write_csv(results_csv_path)
    retrieval_summary_df.write_parquet(retrieval_summary_parquet_path, compression="zstd")
    retrieval_summary_df.write_csv(retrieval_summary_csv_path)
    pairwise_df.write_parquet(pairwise_parquet_path, compression="zstd")
    pairwise_df.write_csv(pairwise_csv_path)
    case_summary_df.write_parquet(case_summary_parquet_path, compression="zstd")
    case_summary_df.write_csv(case_summary_csv_path)

    case_summary_json_payload: dict[str, Any] = {
        "pipeline_name": "refinitiv_step1_ownership_validation_results",
        "artifact_version": "v1",
        "generated_at_utc": dt.datetime.now(dt.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "source_handoff_csv": str(handoff_csv_path),
        "source_filled_workbook": str(filled_workbook_path),
        "ownership_results_rows": int(results_df.height),
        **case_summary_payload,
    }
    _write_json(case_summary_json_path, case_summary_json_payload)

    manifest_payload = {
        **case_summary_json_payload,
        "results_columns": list(OWNERSHIP_VALIDATION_RESULTS_COLUMNS),
        "retrieval_summary_columns": list(OWNERSHIP_VALIDATION_RETRIEVAL_SUMMARY_COLUMNS),
        "pairwise_comparison_columns": list(OWNERSHIP_VALIDATION_PAIRWISE_COMPARISON_COLUMNS),
        "case_summary_columns": list(OWNERSHIP_VALIDATION_CASE_SUMMARY_COLUMNS),
        "artifacts": {
            "refinitiv_ownership_validation_results_parquet": str(results_parquet_path),
            "refinitiv_ownership_validation_results_csv": str(results_csv_path),
            "refinitiv_ownership_validation_retrieval_summary_parquet": str(retrieval_summary_parquet_path),
            "refinitiv_ownership_validation_retrieval_summary_csv": str(retrieval_summary_csv_path),
            "refinitiv_ownership_validation_pairwise_comparisons_parquet": str(pairwise_parquet_path),
            "refinitiv_ownership_validation_pairwise_comparisons_csv": str(pairwise_csv_path),
            "refinitiv_ownership_validation_case_summary_parquet": str(case_summary_parquet_path),
            "refinitiv_ownership_validation_case_summary_csv": str(case_summary_csv_path),
            "refinitiv_ownership_validation_case_summary_json": str(case_summary_json_path),
            "refinitiv_ownership_validation_results_manifest": str(manifest_path),
        },
    }
    _write_json(manifest_path, manifest_payload)

    return {
        "refinitiv_ownership_validation_results_parquet": results_parquet_path,
        "refinitiv_ownership_validation_results_csv": results_csv_path,
        "refinitiv_ownership_validation_retrieval_summary_parquet": retrieval_summary_parquet_path,
        "refinitiv_ownership_validation_retrieval_summary_csv": retrieval_summary_csv_path,
        "refinitiv_ownership_validation_pairwise_comparisons_parquet": pairwise_parquet_path,
        "refinitiv_ownership_validation_pairwise_comparisons_csv": pairwise_csv_path,
        "refinitiv_ownership_validation_case_summary_parquet": case_summary_parquet_path,
        "refinitiv_ownership_validation_case_summary_csv": case_summary_csv_path,
        "refinitiv_ownership_validation_case_summary_json": case_summary_json_path,
        "refinitiv_ownership_validation_results_manifest": manifest_path,
    }


def _read_resolution_artifact_parquet(parquet_path: Path | str) -> pl.DataFrame:
    parquet_path = Path(parquet_path)
    if not parquet_path.exists():
        raise FileNotFoundError(f"resolved Refinitiv artifact not found: {parquet_path}")
    df = pl.read_parquet(parquet_path)
    missing = [name for name in RIC_LOOKUP_RESOLUTION_OUTPUT_COLUMNS if name not in df.columns]
    if missing:
        raise ValueError(f"resolved Refinitiv artifact missing required columns: {missing}")
    return _cast_df_to_schema(
        df.select(RIC_LOOKUP_RESOLUTION_OUTPUT_COLUMNS),
        _resolution_output_schema(),
    ).select(RIC_LOOKUP_RESOLUTION_OUTPUT_COLUMNS)


def _ownership_universe_request_date(value: Any) -> str | None:
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    normalized = _normalize_lookup_text(value)
    return normalized


def _ownership_universe_snapshot_row(
    record: dict[str, Any],
    *,
    ownership_lookup_role: str,
    lookup_input: str | None,
    lookup_input_source: str | None,
    retrieval_eligible: bool,
    retrieval_exclusion_reason: str | None,
) -> dict[str, Any]:
    bridge_row_id = _normalize_lookup_text(record.get("bridge_row_id"))
    request_start_date = _ownership_universe_request_date(record.get("first_seen_caldt"))
    request_end_date = _ownership_universe_request_date(record.get("last_seen_caldt"))
    candidate_ric = _normalize_lookup_text(lookup_input)
    ownership_lookup_row_id = (
        f"{bridge_row_id}|{ownership_lookup_role}" if bridge_row_id is not None else ownership_lookup_role
    )
    return {
        **{name: record.get(name) for name in RIC_LOOKUP_RESOLUTION_OUTPUT_COLUMNS},
        "diagnostic_case_id": bridge_row_id,
        "candidate_slot": ownership_lookup_role,
        "candidate_ric": candidate_ric,
        "ownership_lookup_row_id": ownership_lookup_row_id,
        "ownership_lookup_role": ownership_lookup_role,
        "lookup_input": lookup_input,
        "lookup_input_source": lookup_input_source,
        "request_start_date": request_start_date,
        "request_end_date": request_end_date,
        "retrieval_eligible": retrieval_eligible,
        "retrieval_exclusion_reason": retrieval_exclusion_reason,
    }


def _summarize_refinitiv_step1_ownership_universe_handoff(
    handoff_df: pl.DataFrame,
    *,
    source_resolution_row_count: int,
) -> dict[str, Any]:
    eligible_df = handoff_df.filter(pl.col("retrieval_eligible").fill_null(False))
    noneligible_df = handoff_df.filter(~pl.col("retrieval_eligible").fill_null(False))
    eligible_records = eligible_df.to_dicts()
    noneligible_records = noneligible_df.to_dicts()
    return {
        "resolved_universe_row_count": int(source_resolution_row_count),
        "handoff_snapshot_row_count": int(handoff_df.height),
        "retrieval_eligible_row_count": int(eligible_df.height),
        "retrieval_eligible_bridge_row_count": int(
            eligible_df.select(pl.col("bridge_row_id").drop_nulls().n_unique()).item()
        )
        if eligible_df.height
        else 0,
        "retrieval_eligible_kypermno_count": int(
            eligible_df.select(pl.col("KYPERMNO").drop_nulls().n_unique()).item()
        )
        if eligible_df.height
        else 0,
        "non_retrieval_row_count": int(noneligible_df.height),
        "retrieval_role_counts": _value_counts(eligible_records, "ownership_lookup_role"),
        "retrieval_exclusion_reason_counts": _value_counts(noneligible_records, "retrieval_exclusion_reason"),
        "retrieval_sheet_name": OWNERSHIP_UNIVERSE_RETRIEVAL_SHEET_NAME,
        "request_block_headers": list(OWNERSHIP_UNIVERSE_BLOCK_HEADERS),
        "visible_input_field_order": list(OWNERSHIP_UNIVERSE_VISIBLE_INPUT_FIELDS),
    }


def build_refinitiv_step1_ownership_universe_handoff(
    resolution_df: pl.DataFrame,
) -> tuple[pl.DataFrame, dict[str, Any]]:
    missing = [name for name in RIC_LOOKUP_RESOLUTION_OUTPUT_COLUMNS if name not in resolution_df.columns]
    if missing:
        raise ValueError(f"ownership universe handoff input missing required columns: {missing}")

    normalized_df = _cast_df_to_schema(
        resolution_df.select(RIC_LOOKUP_RESOLUTION_OUTPUT_COLUMNS),
        _resolution_output_schema(),
    ).sort(
        pl.col("KYPERMNO").cast(pl.Int64, strict=False),
        "KYPERMNO",
        "first_seen_caldt",
        "last_seen_caldt",
        "bridge_row_id",
    )

    if normalized_df.height == 0:
        empty_df = _empty_ownership_universe_handoff_df().select(OWNERSHIP_UNIVERSE_HANDOFF_COLUMNS)
        return empty_df, _summarize_refinitiv_step1_ownership_universe_handoff(
            empty_df,
            source_resolution_row_count=0,
        )

    rows: list[dict[str, Any]] = []
    for record in normalized_df.to_dicts():
        effective_collection_ric = _normalize_lookup_text(record.get("effective_collection_ric"))
        conventional_identity_conflict = bool(record.get("conventional_identity_conflict"))
        ticker_candidate_available = bool(record.get("ticker_candidate_available"))
        ticker_candidate_ric = _normalize_lookup_text(record.get("ticker_candidate_ric"))

        if effective_collection_ric is not None:
            rows.append(
                _ownership_universe_snapshot_row(
                    record,
                    ownership_lookup_role="UNIVERSE_EFFECTIVE",
                    lookup_input=effective_collection_ric,
                    lookup_input_source="effective_collection_ric",
                    retrieval_eligible=True,
                    retrieval_exclusion_reason=None,
                )
            )
            continue

        if conventional_identity_conflict:
            emitted = False
            isin_candidate_ric = _normalize_lookup_text(record.get("ISIN_returned_ric"))
            cusip_candidate_ric = _normalize_lookup_text(record.get("CUSIP_returned_ric"))
            if isin_candidate_ric is not None:
                rows.append(
                    _ownership_universe_snapshot_row(
                        record,
                        ownership_lookup_role="UNIVERSE_TARGET_ISIN_CANDIDATE",
                        lookup_input=isin_candidate_ric,
                        lookup_input_source="ISIN_returned_ric",
                        retrieval_eligible=True,
                        retrieval_exclusion_reason=None,
                    )
                )
                emitted = True
            if cusip_candidate_ric is not None:
                rows.append(
                    _ownership_universe_snapshot_row(
                        record,
                        ownership_lookup_role="UNIVERSE_TARGET_CUSIP_CANDIDATE",
                        lookup_input=cusip_candidate_ric,
                        lookup_input_source="CUSIP_returned_ric",
                        retrieval_eligible=True,
                        retrieval_exclusion_reason=None,
                    )
                )
                emitted = True
            if emitted:
                continue

        if ticker_candidate_available and ticker_candidate_ric is not None and not conventional_identity_conflict:
            rows.append(
                _ownership_universe_snapshot_row(
                    record,
                    ownership_lookup_role="UNIVERSE_TARGET_TICKER_CANDIDATE",
                    lookup_input=ticker_candidate_ric,
                    lookup_input_source="ticker_candidate_ric",
                    retrieval_eligible=True,
                    retrieval_exclusion_reason=None,
                )
            )
            continue

        rows.append(
            _ownership_universe_snapshot_row(
                record,
                ownership_lookup_role="UNIVERSE_NOT_RETRIEVABLE",
                lookup_input=None,
                lookup_input_source=None,
                retrieval_eligible=False,
                retrieval_exclusion_reason="no_usable_lookup_input",
            )
        )

    handoff_df = (
        _build_explicit_schema_df(rows, _ownership_universe_handoff_schema())
        .select(OWNERSHIP_UNIVERSE_HANDOFF_COLUMNS)
    )
    return handoff_df, _summarize_refinitiv_step1_ownership_universe_handoff(
        handoff_df,
        source_resolution_row_count=normalized_df.height,
    )


def _parse_refinitiv_ownership_universe_filled_workbook(
    filled_workbook_path: Path | str,
    handoff_df: pl.DataFrame,
) -> pl.DataFrame:
    from openpyxl import load_workbook

    eligible_handoff_df = handoff_df.filter(pl.col("retrieval_eligible").fill_null(False))
    if eligible_handoff_df.height == 0:
        return _empty_ownership_universe_results_df().select(OWNERSHIP_UNIVERSE_RESULTS_COLUMNS)

    workbook = load_workbook(Path(filled_workbook_path), read_only=False, data_only=True)
    try:
        if OWNERSHIP_UNIVERSE_RETRIEVAL_SHEET_NAME not in workbook.sheetnames:
            raise ValueError(
                f"filled ownership universe workbook missing sheet: {OWNERSHIP_UNIVERSE_RETRIEVAL_SHEET_NAME}"
            )

        worksheet = workbook[OWNERSHIP_UNIVERSE_RETRIEVAL_SHEET_NAME]
        block_width = len(OWNERSHIP_UNIVERSE_BLOCK_HEADERS)
        expected_headers = list(OWNERSHIP_UNIVERSE_BLOCK_HEADERS)
        handoff_by_candidate_key = {
            _ownership_universe_candidate_key(row): row for row in eligible_handoff_df.to_dicts()
        }
        results: list[dict[str, Any]] = []

        for base_col in range(1, worksheet.max_column + 1, block_width):
            header_values = [
                _normalize_workbook_scalar(worksheet.cell(row=1, column=base_col + offset).value)
                for offset in range(block_width)
            ]
            if all(value is None for value in header_values):
                continue
            if header_values != expected_headers:
                raise ValueError(
                    "ownership_retrieval contains an unexpected request block header at "
                    f"column {base_col}: {header_values}"
                )

            block_values = {
                field_name: _normalize_workbook_scalar(
                    worksheet.cell(row=2 + field_offset, column=base_col).value
                )
                for field_offset, field_name in enumerate(OWNERSHIP_UNIVERSE_VISIBLE_INPUT_FIELDS)
            }
            candidate_ric = _normalize_lookup_text(block_values.get("candidate_ric"))
            if candidate_ric is None:
                continue
            matched_handoff = handoff_by_candidate_key.get(
                (
                    _normalize_lookup_text(block_values.get("bridge_row_id")),
                    _normalize_lookup_text(block_values.get("bridge_row_id")),
                    _normalize_lookup_text(block_values.get("candidate_slot")),
                    candidate_ric,
                )
            )
            if matched_handoff is None:
                raise ValueError(
                    "ownership_retrieval contains a block that cannot be matched to the handoff snapshot: "
                    f"{block_values}"
                )

            for excel_row in range(3, worksheet.max_row + 1):
                returned_ric = _normalize_ownership_result_text(
                    worksheet.cell(row=excel_row, column=base_col + 1).value
                )
                returned_date = _normalize_ownership_result_date(
                    worksheet.cell(row=excel_row, column=base_col + 2).value
                )
                returned_value = _normalize_ownership_result_value(
                    worksheet.cell(row=excel_row, column=base_col + 3).value
                )
                returned_category = _normalize_ownership_result_text(
                    worksheet.cell(row=excel_row, column=base_col + 4).value
                )

                if (
                    returned_ric is None
                    and returned_date is None
                    and returned_value is None
                    and returned_category is None
                ):
                    break
                results.append(
                    {
                        **{name: matched_handoff.get(name) for name in OWNERSHIP_UNIVERSE_HANDOFF_COLUMNS},
                        "returned_ric": returned_ric,
                        "returned_date": returned_date,
                        "returned_category": returned_category,
                        "returned_value": returned_value,
                    }
                )
    finally:
        workbook.close()

    if not results:
        return _empty_ownership_universe_results_df().select(OWNERSHIP_UNIVERSE_RESULTS_COLUMNS)

    return (
        _build_explicit_schema_df(results, _ownership_universe_results_schema())
        .select(OWNERSHIP_UNIVERSE_RESULTS_COLUMNS)
        .unique(
            subset=[
                "ownership_lookup_row_id",
                "returned_date",
                "returned_category",
                "returned_value",
                "returned_ric",
            ],
            maintain_order=True,
        )
    )


def _ownership_universe_candidate_key(record: dict[str, Any]) -> tuple[str | None, str | None, str | None, str | None]:
    return (
        _normalize_lookup_text(record.get("diagnostic_case_id")),
        _normalize_lookup_text(record.get("bridge_row_id")),
        _normalize_lookup_text(record.get("candidate_slot")),
        _normalize_lookup_text(record.get("candidate_ric")),
    )


def build_refinitiv_ownership_universe_row_summary(
    handoff_df: pl.DataFrame,
    results_df: pl.DataFrame,
) -> pl.DataFrame:
    if handoff_df.height == 0:
        return _empty_ownership_universe_row_summary_df().select(OWNERSHIP_UNIVERSE_ROW_SUMMARY_COLUMNS)

    key_columns = list(OWNERSHIP_UNIVERSE_HANDOFF_COLUMNS)
    base_df = handoff_df.select(key_columns)

    if results_df.height == 0:
        return (
            base_df.with_columns(
                pl.col("retrieval_eligible").fill_null(False).alias("retrieval_row_present"),
                pl.lit(0, dtype=pl.Int64).alias("ownership_rows_returned"),
                pl.lit(None, dtype=pl.Date).alias("ownership_first_date"),
                pl.lit(None, dtype=pl.Date).alias("ownership_last_date"),
                pl.lit(0, dtype=pl.Int64).alias("ownership_distinct_categories"),
                pl.lit(0, dtype=pl.Int64).alias("ownership_nonnull_value_count"),
                pl.lit(False, dtype=pl.Boolean).alias("ownership_single_returned_ric"),
                pl.lit(0, dtype=pl.Int64).alias("ownership_returned_ric_nunique"),
            )
            .select(OWNERSHIP_UNIVERSE_ROW_SUMMARY_COLUMNS)
        )

    agg_df = results_df.group_by(key_columns).agg(
        pl.len().alias("ownership_rows_returned"),
        pl.col("returned_date").drop_nulls().min().alias("ownership_first_date"),
        pl.col("returned_date").drop_nulls().max().alias("ownership_last_date"),
        pl.col("returned_category").drop_nulls().n_unique().cast(pl.Int64).alias("ownership_distinct_categories"),
        pl.col("returned_value").drop_nulls().len().cast(pl.Int64).alias("ownership_nonnull_value_count"),
        pl.col("returned_ric").drop_nulls().n_unique().cast(pl.Int64).alias("ownership_returned_ric_nunique"),
    )

    summary_df = (
        base_df.join(agg_df, on=key_columns, how="left")
        .with_columns(
            pl.col("retrieval_eligible").fill_null(False).alias("retrieval_row_present"),
            pl.col("ownership_rows_returned").fill_null(0).cast(pl.Int64),
            pl.col("ownership_distinct_categories").fill_null(0).cast(pl.Int64),
            pl.col("ownership_nonnull_value_count").fill_null(0).cast(pl.Int64),
            pl.col("ownership_returned_ric_nunique").fill_null(0).cast(pl.Int64),
        )
        .with_columns(
            (pl.col("ownership_returned_ric_nunique") == 1).fill_null(False).alias("ownership_single_returned_ric")
        )
        .select(OWNERSHIP_UNIVERSE_ROW_SUMMARY_COLUMNS)
    )
    return _cast_df_to_schema(summary_df, _ownership_universe_row_summary_schema()).select(
        OWNERSHIP_UNIVERSE_ROW_SUMMARY_COLUMNS
    )


def run_refinitiv_step1_ownership_universe_handoff_pipeline(
    resolution_artifact_path: Path | str,
    output_dir: Path | str,
) -> dict[str, Path]:
    resolution_artifact_path = Path(resolution_artifact_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    resolution_df = _read_resolution_artifact_parquet(resolution_artifact_path)
    handoff_df, summary = build_refinitiv_step1_ownership_universe_handoff(resolution_df)

    handoff_parquet_path = output_dir / "refinitiv_ownership_universe_handoff_common_stock.parquet"
    handoff_xlsx_path = output_dir / "refinitiv_ownership_universe_handoff_common_stock.xlsx"

    handoff_df.write_parquet(handoff_parquet_path, compression="zstd")
    readme_payload = {
        "pipeline_name": "refinitiv_step1_ownership_universe_handoff",
        "artifact_version": "v1",
        "generated_at_utc": dt.datetime.now(dt.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "source_resolution_artifact": str(resolution_artifact_path),
        **summary,
        "handoff_columns": list(OWNERSHIP_UNIVERSE_HANDOFF_COLUMNS),
    }
    _write_workbook_or_reuse_locked_output(
        write_refinitiv_ownership_universe_workbook,
        handoff_df,
        handoff_xlsx_path,
        readme_payload=readme_payload,
        input_field_order=OWNERSHIP_UNIVERSE_VISIBLE_INPUT_FIELDS,
        block_headers=OWNERSHIP_UNIVERSE_BLOCK_HEADERS,
    )

    return {
        "refinitiv_ownership_universe_handoff_common_stock_parquet": handoff_parquet_path,
        "refinitiv_ownership_universe_handoff_common_stock_xlsx": handoff_xlsx_path,
    }


def run_refinitiv_step1_ownership_universe_results_pipeline(
    filled_workbook_path: Path | str,
    output_dir: Path | str,
    *,
    handoff_parquet_path: Path | str | None = None,
) -> dict[str, Path]:
    filled_workbook_path = Path(filled_workbook_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    if handoff_parquet_path is None:
        handoff_parquet_path = output_dir / "refinitiv_ownership_universe_handoff_common_stock.parquet"
    handoff_parquet_path = Path(handoff_parquet_path)

    handoff_df = _cast_df_to_schema(
        pl.read_parquet(handoff_parquet_path).select(OWNERSHIP_UNIVERSE_HANDOFF_COLUMNS),
        _ownership_universe_handoff_schema(),
    ).select(OWNERSHIP_UNIVERSE_HANDOFF_COLUMNS)
    results_df = _parse_refinitiv_ownership_universe_filled_workbook(filled_workbook_path, handoff_df).select(
        OWNERSHIP_UNIVERSE_RESULTS_COLUMNS
    )
    row_summary_df = build_refinitiv_ownership_universe_row_summary(handoff_df, results_df)

    results_parquet_path = output_dir / "refinitiv_ownership_universe_results.parquet"
    row_summary_parquet_path = output_dir / "refinitiv_ownership_universe_row_summary.parquet"

    results_df.write_parquet(results_parquet_path, compression="zstd")
    row_summary_df.write_parquet(row_summary_parquet_path, compression="zstd")

    return {
        "refinitiv_ownership_universe_results_parquet": results_parquet_path,
        "refinitiv_ownership_universe_row_summary_parquet": row_summary_parquet_path,
    }


def build_refinitiv_step1_ownership_authority_tables(
    resolution_df: pl.DataFrame,
    ownership_results_df: pl.DataFrame,
    ownership_row_summary_df: pl.DataFrame,
    *,
    reviewed_ticker_allowlist_df: pl.DataFrame | None = None,
) -> tuple[dict[str, pl.DataFrame], dict[str, Any]]:
    from thesis_pkg.pipelines.refinitiv.authority import (
        build_refinitiv_step1_ownership_authority_tables as _impl,
    )

    return _impl(
        resolution_df,
        ownership_results_df,
        ownership_row_summary_df,
        reviewed_ticker_allowlist_df=reviewed_ticker_allowlist_df,
    )


def run_refinitiv_step1_ownership_authority_pipeline(
    *,
    resolution_artifact_path: Path | str,
    ownership_results_artifact_path: Path | str,
    ownership_row_summary_artifact_path: Path | str,
    output_dir: Path | str,
    reviewed_ticker_allowlist_path: Path | str | None = None,
) -> dict[str, Path]:
    from thesis_pkg.pipelines.refinitiv.authority import (
        run_refinitiv_step1_ownership_authority_pipeline as _impl,
    )

    return _impl(
        resolution_artifact_path=resolution_artifact_path,
        ownership_results_artifact_path=ownership_results_artifact_path,
        ownership_row_summary_artifact_path=ownership_row_summary_artifact_path,
        output_dir=output_dir,
        reviewed_ticker_allowlist_path=reviewed_ticker_allowlist_path,
    )


def build_refinitiv_lm2011_doc_ownership_requests(
    doc_filing_df: pl.DataFrame,
    authority_decisions_df: pl.DataFrame,
    authority_exceptions_df: pl.DataFrame,
) -> pl.DataFrame:
    from thesis_pkg.pipelines.refinitiv.doc_ownership import (
        build_refinitiv_lm2011_doc_ownership_requests as _impl,
    )

    return _impl(
        doc_filing_df,
        authority_decisions_df,
        authority_exceptions_df,
    )


def run_refinitiv_lm2011_doc_ownership_exact_handoff_pipeline(
    *,
    doc_filing_artifact_path: Path | str,
    authority_decisions_artifact_path: Path | str,
    authority_exceptions_artifact_path: Path | str,
    output_dir: Path | str,
) -> dict[str, Path]:
    from thesis_pkg.pipelines.refinitiv.doc_ownership import (
        run_refinitiv_lm2011_doc_ownership_exact_handoff_pipeline as _impl,
    )

    return _impl(
        doc_filing_artifact_path=doc_filing_artifact_path,
        authority_decisions_artifact_path=authority_decisions_artifact_path,
        authority_exceptions_artifact_path=authority_exceptions_artifact_path,
        output_dir=output_dir,
    )


def run_refinitiv_lm2011_doc_ownership_fallback_handoff_pipeline(
    *,
    exact_filled_workbook_path: Path | str,
    output_dir: Path | str,
) -> dict[str, Path]:
    from thesis_pkg.pipelines.refinitiv.doc_ownership import (
        run_refinitiv_lm2011_doc_ownership_fallback_handoff_pipeline as _impl,
    )

    return _impl(
        exact_filled_workbook_path=exact_filled_workbook_path,
        output_dir=output_dir,
    )


def run_refinitiv_lm2011_doc_ownership_finalize_pipeline(
    *,
    output_dir: Path | str,
    fallback_filled_workbook_path: Path | str | None = None,
) -> dict[str, Path]:
    from thesis_pkg.pipelines.refinitiv.doc_ownership import (
        run_refinitiv_lm2011_doc_ownership_finalize_pipeline as _impl,
    )

    return _impl(
        output_dir=output_dir,
        fallback_filled_workbook_path=fallback_filled_workbook_path,
    )


def _is_failed_lookup_record(record: dict[str, Any]) -> tuple[bool, str, bool]:
    vendor_primary_ric = _normalize_lookup_text(record.get("vendor_primary_ric"))
    returned_fields = (
        _normalize_lookup_text(record.get("vendor_returned_name")),
        _normalize_lookup_text(record.get("vendor_returned_cusip")),
        _normalize_lookup_text(record.get("vendor_returned_isin")),
    )
    ric_missing = vendor_primary_ric is None or vendor_primary_ric.upper() == "NULL"
    invalid_identifier_signal = any(
        value is not None and "invalid identifier" in value.lower()
        for value in returned_fields
    )
    failed_lookup = ric_missing or invalid_identifier_signal
    if ric_missing and invalid_identifier_signal:
        failed_reason = "null_ric_and_invalid_identifier"
    elif ric_missing:
        failed_reason = "null_ric"
    elif invalid_identifier_signal:
        failed_reason = "invalid_identifier"
    else:
        failed_reason = "successful_lookup"
    return failed_lookup, failed_reason, invalid_identifier_signal


def _alternative_identifier(record: dict[str, Any]) -> tuple[str | None, str | None]:
    preferred_lookup_type = _normalize_lookup_text(record.get("preferred_lookup_type"))
    preferred_lookup_id = _normalize_lookup_text(record.get("preferred_lookup_id"))
    if preferred_lookup_type == "ISIN":
        candidates = (
            (_normalize_lookup_text(record.get("CUSIP")), "CUSIP"),
            (_normalize_lookup_text(record.get("TICKER")), "TICKER"),
        )
    elif preferred_lookup_type == "CUSIP":
        candidates = (
            (_normalize_lookup_text(record.get("ISIN")), "ISIN"),
            (_normalize_lookup_text(record.get("TICKER")), "TICKER"),
        )
    elif preferred_lookup_type == "TICKER":
        candidates = (
            (_normalize_lookup_text(record.get("ISIN")), "ISIN"),
            (_normalize_lookup_text(record.get("CUSIP")), "CUSIP"),
        )
    else:
        return None, None

    for candidate, candidate_type in candidates:
        if candidate is not None and candidate != preferred_lookup_id:
            return candidate, candidate_type
    return None, None


def build_refinitiv_null_ric_rescue_candidates(
    lookup_df: pl.DataFrame,
) -> tuple[pl.DataFrame, dict[str, Any]]:
    required_columns = list(RIC_LOOKUP_COLUMNS)
    missing = [name for name in required_columns if name not in lookup_df.columns]
    if missing:
        raise ValueError(f"lookup diagnostics input missing required columns: {missing}")

    normalized_df = lookup_df.select(
        pl.col("bridge_row_id").cast(pl.Utf8, strict=False),
        pl.col("KYPERMNO").cast(pl.Utf8, strict=False),
        pl.col("CUSIP").cast(pl.Utf8, strict=False),
        pl.col("ISIN").cast(pl.Utf8, strict=False),
        pl.col("TICKER").cast(pl.Utf8, strict=False),
        pl.col("first_seen_caldt").cast(pl.Date, strict=False),
        pl.col("last_seen_caldt").cast(pl.Date, strict=False),
        pl.col("preferred_lookup_id").cast(pl.Utf8, strict=False),
        pl.col("preferred_lookup_type").cast(pl.Utf8, strict=False),
        pl.col("vendor_primary_ric").cast(pl.Utf8, strict=False),
        pl.col("vendor_returned_name").cast(pl.Utf8, strict=False),
        pl.col("vendor_returned_cusip").cast(pl.Utf8, strict=False),
        pl.col("vendor_returned_isin").cast(pl.Utf8, strict=False),
        pl.col("vendor_match_status").cast(pl.Utf8, strict=False),
        pl.col("vendor_notes").cast(pl.Utf8, strict=False),
    )

    all_records = normalized_df.to_dicts()
    successful_by_kypermno: dict[str, list[dict[str, Any]]] = {}
    failure_reason_counts: Counter[str] = Counter()
    successful_rows = 0

    for record in all_records:
        failed_lookup, failed_reason, invalid_identifier_signal = _is_failed_lookup_record(record)
        record["failed_lookup_flag"] = failed_lookup
        record["failed_lookup_reason"] = failed_reason
        record["invalid_identifier_signal"] = invalid_identifier_signal
        if failed_lookup:
            failure_reason_counts[failed_reason] += 1
        else:
            successful_rows += 1
            kypermno = _normalize_lookup_text(record.get("KYPERMNO"))
            if kypermno is not None:
                successful_by_kypermno.setdefault(kypermno, []).append(record)

    diagnostic_rows: list[dict[str, Any]] = []

    for record in all_records:
        if not bool(record["failed_lookup_flag"]):
            continue

        kypermno = _normalize_lookup_text(record.get("KYPERMNO"))
        successful_rows_same_permno = successful_by_kypermno.get(kypermno or "", [])
        current_first_seen = record.get("first_seen_caldt")
        current_last_seen = record.get("last_seen_caldt")

        before_gaps: list[int] = []
        after_gaps: list[int] = []
        overlap_exists = False
        successful_identifier_pairs: set[tuple[str | None, str | None]] = set()
        successful_rics: set[str] = set()

        for successful in successful_rows_same_permno:
            successful_identifier_pairs.add(
                (
                    _normalize_lookup_text(successful.get("CUSIP")),
                    _normalize_lookup_text(successful.get("ISIN")),
                )
            )
            successful_ric = _normalize_lookup_text(successful.get("vendor_primary_ric"))
            if successful_ric is not None:
                successful_rics.add(successful_ric)

            successful_first_seen = successful.get("first_seen_caldt")
            successful_last_seen = successful.get("last_seen_caldt")
            if (
                isinstance(current_first_seen, dt.date)
                and isinstance(current_last_seen, dt.date)
                and isinstance(successful_first_seen, dt.date)
                and isinstance(successful_last_seen, dt.date)
            ):
                if successful_last_seen < current_first_seen:
                    before_gaps.append((current_first_seen - successful_last_seen).days)
                elif successful_first_seen > current_last_seen:
                    after_gaps.append((successful_first_seen - current_last_seen).days)
                else:
                    overlap_exists = True

        nearest_before = min(before_gaps) if before_gaps else None
        nearest_after = min(after_gaps) if after_gaps else None
        failed_pair = (
            _normalize_lookup_text(record.get("CUSIP")),
            _normalize_lookup_text(record.get("ISIN")),
        )
        failed_pair_matches_success = failed_pair in successful_identifier_pairs
        alt_identifier, alt_identifier_type = _alternative_identifier(record)

        unique_successful_pairs = {
            pair for pair in successful_identifier_pairs if pair != (None, None)
        }
        candidate_successful_ric = next(iter(successful_rics)) if len(successful_rics) == 1 else None
        candidate_successful_ric_available = candidate_successful_ric is not None
        if len(unique_successful_pairs) == 1:
            candidate_successful_cusip, candidate_successful_isin = next(iter(unique_successful_pairs))
        else:
            candidate_successful_cusip, candidate_successful_isin = None, None

        has_success = bool(successful_rows_same_permno)
        alternative_identifier_available = alt_identifier is not None
        diagnostic_rows.append(
            {
                **{name: record.get(name) for name in RIC_LOOKUP_COLUMNS},
                "failed_lookup_flag": True,
                "failed_lookup_reason": record["failed_lookup_reason"],
                "invalid_identifier_signal": bool(record["invalid_identifier_signal"]),
                "successful_row_exists_for_kypermno": has_success,
                "successful_row_exists_before_span": bool(before_gaps),
                "successful_row_exists_after_span": bool(after_gaps),
                "successful_row_overlap_exists": overlap_exists,
                "nearest_successful_gap_days_before": nearest_before,
                "nearest_successful_gap_days_after": nearest_after,
                "unique_successful_identifier_pair_count": len(unique_successful_pairs),
                "unique_successful_ric_count": len(successful_rics),
                "failed_identifier_pair_matches_success": failed_pair_matches_success,
                "alternative_identifier_available": alternative_identifier_available,
                "alternative_identifier": alt_identifier,
                "alternative_identifier_type": alt_identifier_type,
                "candidate_successful_ric_available": candidate_successful_ric_available,
                "candidate_successful_ric": candidate_successful_ric,
                "candidate_successful_cusip": candidate_successful_cusip,
                "candidate_successful_isin": candidate_successful_isin,
            }
        )

    diagnostics_df = pl.DataFrame(diagnostic_rows) if diagnostic_rows else pl.DataFrame(schema={
        name: pl.Utf8 for name in NULL_RIC_DIAGNOSTIC_COLUMNS
    })
    if diagnostics_df.height:
        diagnostics_df = diagnostics_df.select(
            [
                pl.col(name).cast(pl.Date, strict=False).alias(name)
                if name in ("first_seen_caldt", "last_seen_caldt")
                else pl.col(name)
                for name in diagnostics_df.columns
            ]
        ).select(NULL_RIC_DIAGNOSTIC_COLUMNS)
    else:
        diagnostics_df = diagnostics_df.with_columns(
            pl.col("first_seen_caldt").cast(pl.Date, strict=False),
            pl.col("last_seen_caldt").cast(pl.Date, strict=False),
        )

    summary: dict[str, Any] = {
        "total_lookup_rows": int(len(all_records)),
        "successful_lookup_rows": int(successful_rows),
        "failed_lookup_rows": int(len(diagnostic_rows)),
        "failure_reason_counts": dict(sorted(failure_reason_counts.items())),
        "diagnostic_flag_counts": {
            "successful_row_exists_before_span": int(
                diagnostics_df.select(pl.col("successful_row_exists_before_span").cast(pl.Int64).sum()).item()
            )
            if diagnostics_df.height
            else 0,
            "successful_row_exists_after_span": int(
                diagnostics_df.select(pl.col("successful_row_exists_after_span").cast(pl.Int64).sum()).item()
            )
            if diagnostics_df.height
            else 0,
            "successful_row_overlap_exists": int(
                diagnostics_df.select(pl.col("successful_row_overlap_exists").cast(pl.Int64).sum()).item()
            )
            if diagnostics_df.height
            else 0,
            "alternative_identifier_available": int(
                diagnostics_df.select(pl.col("alternative_identifier_available").cast(pl.Int64).sum()).item()
            )
            if diagnostics_df.height
            else 0,
            "candidate_successful_ric_available": int(
                diagnostics_df.select(pl.col("candidate_successful_ric_available").cast(pl.Int64).sum()).item()
            )
            if diagnostics_df.height
            else 0,
            "no_successful_row_for_kypermno": int(
                diagnostics_df.select((~pl.col("successful_row_exists_for_kypermno")).cast(pl.Int64).sum()).item()
            )
            if diagnostics_df.height
            else 0,
            "multiple_successful_identifier_pairs_or_rics": int(
                diagnostics_df.select(
                    (
                        (pl.col("unique_successful_identifier_pair_count") > 1)
                        | (pl.col("unique_successful_ric_count") > 1)
                    )
                    .cast(pl.Int64)
                    .sum()
                ).item()
            )
            if diagnostics_df.height
            else 0,
        },
        "rows_with_alternative_identifier": int(
            diagnostics_df.select(pl.col("alternative_identifier").is_not_null().sum()).item()
        )
        if diagnostics_df.height
        else 0,
    }
    return diagnostics_df, summary


def _build_null_ric_review_frame(diagnostics_df: pl.DataFrame) -> pl.DataFrame:
    review_df = diagnostics_df.select(
        [
            *[pl.col(name) for name in NULL_RIC_REVIEW_COLUMNS if name in diagnostics_df.columns],
            pl.lit(None, dtype=pl.Utf8).alias("test_category"),
            pl.lit(None, dtype=pl.Utf8).alias("test_method"),
            pl.lit(None, dtype=pl.Utf8).alias("test_result"),
            pl.lit(None, dtype=pl.Utf8).alias("test_notes"),
        ]
    )
    return review_df.select(NULL_RIC_REVIEW_COLUMNS)


def _date_to_text(value: Any) -> str | None:
    if isinstance(value, dt.date):
        return value.isoformat()
    normalized = _normalize_lookup_text(value)
    return normalized


def _explicit_sample_category_field(review_df: pl.DataFrame) -> str | None:
    for field_name in ("sample_category", "test_category"):
        if field_name not in review_df.columns:
            continue
        if review_df.select(pl.col(field_name).drop_nulls().len()).item() > 0:
            non_blank_count = review_df.select(
                pl.col(field_name)
                .cast(pl.Utf8, strict=False)
                .str.strip_chars()
                .replace("", None)
                .drop_nulls()
                .len()
            ).item()
            if int(non_blank_count) > 0:
                return field_name
    return None


def _resolve_ownership_lookup_input(row: dict[str, Any]) -> tuple[str | None, str]:
    for field_name, source_name in (
        ("candidate_successful_ric", "candidate_successful_ric"),
        ("alternative_identifier", "alternative_identifier"),
        ("TICKER", "TICKER"),
        ("preferred_lookup_id", "preferred_lookup_id"),
    ):
        value = _normalize_lookup_text(row.get(field_name))
        if value is not None:
            return value, source_name
    return None, "preferred_lookup_id"


def _ownership_smoke_sample_row(sample_category: str, row: dict[str, Any]) -> dict[str, Any]:
    lookup_input, lookup_input_source = _resolve_ownership_lookup_input(row)
    return {
        "sample_category": sample_category,
        "bridge_row_id": _normalize_lookup_text(row.get("bridge_row_id")),
        "KYPERMNO": _normalize_lookup_text(row.get("KYPERMNO")),
        "TICKER": _normalize_lookup_text(row.get("TICKER")),
        "lookup_input": lookup_input,
        "lookup_input_source": lookup_input_source,
        "request_start_date": _date_to_text(row.get("first_seen_caldt")),
        "request_end_date": _date_to_text(row.get("last_seen_caldt")),
        "preferred_lookup_id": _normalize_lookup_text(row.get("preferred_lookup_id")),
        "preferred_lookup_type": _normalize_lookup_text(row.get("preferred_lookup_type")),
        "alternative_identifier": _normalize_lookup_text(row.get("alternative_identifier")),
        "alternative_identifier_type": _normalize_lookup_text(row.get("alternative_identifier_type")),
        "candidate_successful_ric": _normalize_lookup_text(row.get("candidate_successful_ric")),
        "successful_row_exists_before_span": bool(row.get("successful_row_exists_before_span")),
        "successful_row_exists_after_span": bool(row.get("successful_row_exists_after_span")),
        "successful_row_overlap_exists": bool(row.get("successful_row_overlap_exists")),
        "alternative_identifier_available": bool(row.get("alternative_identifier_available")),
        "candidate_successful_ric_available": bool(row.get("candidate_successful_ric_available")),
        "unique_successful_identifier_pair_count": int(
            row.get("unique_successful_identifier_pair_count") or 0
        ),
        "unique_successful_ric_count": int(row.get("unique_successful_ric_count") or 0),
    }


def build_refinitiv_ownership_smoke_sample(
    review_df: pl.DataFrame,
    *,
    target_block_count: int = 10,
    min_per_category: int = 2,
) -> tuple[pl.DataFrame, dict[str, Any]]:
    sorted_rows = sorted(review_df.to_dicts(), key=lambda row: str(row.get("bridge_row_id") or ""))
    explicit_category_field = _explicit_sample_category_field(review_df)

    category_pools: list[tuple[str, list[dict[str, Any]]]] = []
    if explicit_category_field is not None:
        grouped_rows: dict[str, list[dict[str, Any]]] = {}
        for row in sorted_rows:
            category_value = _normalize_lookup_text(row.get(explicit_category_field))
            if category_value is None:
                continue
            grouped_rows.setdefault(category_value, []).append(row)
        category_pools = [(name, grouped_rows[name]) for name in sorted(grouped_rows)]
        category_source = explicit_category_field
    else:
        category_pools = [
            (category_name, [row for row in sorted_rows if predicate(row)])
            for category_name, predicate in OWNERSHIP_SMOKE_CATEGORY_SPECS
        ]
        category_source = "derived_from_review_flags"

    category_pools = [(name, rows) for name, rows in category_pools if rows]
    sampled_rows: list[dict[str, Any]] = []
    sample_counts: dict[str, int] = {name: 0 for name, _ in category_pools}

    for category_name, rows in category_pools:
        for row in rows[:min(min_per_category, len(rows))]:
            sampled_rows.append(_ownership_smoke_sample_row(category_name, row))
            sample_counts[category_name] += 1

    if len(sampled_rows) < target_block_count:
        for category_name, rows in category_pools:
            start_idx = sample_counts[category_name]
            for row in rows[start_idx:]:
                if len(sampled_rows) >= target_block_count:
                    break
                sampled_rows.append(_ownership_smoke_sample_row(category_name, row))
                sample_counts[category_name] += 1
            if len(sampled_rows) >= target_block_count:
                break

    if sampled_rows:
        sample_df = pl.DataFrame(sampled_rows).select(OWNERSHIP_SMOKE_SAMPLE_COLUMNS)
    else:
        sample_df = pl.DataFrame(schema=OWNERSHIP_SMOKE_SAMPLE_SCHEMA)

    metadata: dict[str, Any] = {
        "sample_category_counts": {name: count for name, count in sample_counts.items() if count > 0},
        "available_category_counts": {name: len(rows) for name, rows in category_pools},
        "category_order": [name for name, _ in category_pools],
        "category_field_used": "sample_category",
        "category_source": category_source,
        "sample_source_name": "failed_lookup_review",
        "target_block_count": int(target_block_count),
        "min_per_category": int(min_per_category),
        "block_count": int(sample_df.height),
        "lookup_input_priority": [
            "candidate_successful_ric",
            "alternative_identifier",
            "TICKER",
            "preferred_lookup_id",
        ],
    }
    return sample_df, metadata


def run_refinitiv_null_ric_diagnostics_pipeline(
    filled_lookup_workbook_path: Path | str,
    output_dir: Path | str,
    *,
    emit_review_workbook: bool = True,
    emit_ownership_smoke_workbook: bool = True,
    ownership_target_block_count: int = 10,
    ownership_min_per_category: int = 2,
) -> dict[str, Path]:
    filled_lookup_workbook_path = Path(filled_lookup_workbook_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    lookup_df = _read_refinitiv_ric_lookup_sheet(filled_lookup_workbook_path)
    diagnostics_df, summary = build_refinitiv_null_ric_rescue_candidates(lookup_df)
    review_df = _build_null_ric_review_frame(diagnostics_df)
    ownership_smoke_df, ownership_smoke_metadata = build_refinitiv_ownership_smoke_sample(
        review_df,
        target_block_count=ownership_target_block_count,
        min_per_category=ownership_min_per_category,
    )

    summary_path = output_dir / "refinitiv_null_ric_diagnostics_summary.json"
    manifest_path = output_dir / "refinitiv_null_ric_diagnostics_manifest.json"
    rescue_parquet_path = output_dir / "refinitiv_null_ric_rescue_candidates.parquet"
    rescue_csv_path = output_dir / "refinitiv_null_ric_rescue_candidates.csv"
    review_workbook_path = output_dir / "refinitiv_null_ric_rescue_candidates_review.xlsx"
    ownership_smoke_parquet_path = output_dir / "refinitiv_ownership_smoke_testing.parquet"
    ownership_smoke_csv_path = output_dir / "refinitiv_ownership_smoke_testing.csv"
    ownership_smoke_workbook_path = output_dir / "refinitiv_ownership_smoke_testing.xlsx"

    diagnostics_df.write_parquet(rescue_parquet_path, compression="zstd")
    diagnostics_df.write_csv(rescue_csv_path)
    ownership_smoke_df.write_parquet(ownership_smoke_parquet_path, compression="zstd")
    ownership_smoke_df.write_csv(ownership_smoke_csv_path)

    summary_payload: dict[str, Any] = {
        "pipeline_name": "refinitiv_null_ric_diagnostics",
        "artifact_version": "v2",
        "generated_at_utc": dt.datetime.now(dt.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "input_filled_lookup_workbook": str(filled_lookup_workbook_path),
        "source_sheet_name": "ric_lookup",
        "failed_lookup_definition": [
            "vendor_primary_ric is null/blank/NULL",
            "or vendor_returned_* contains an invalid identifier signal",
        ],
        "ownership_smoke_sample_source": ownership_smoke_metadata["sample_source_name"],
        "ownership_smoke_category_field_used": ownership_smoke_metadata["category_field_used"],
        "ownership_smoke_category_source": ownership_smoke_metadata["category_source"],
        "ownership_smoke_target_block_count": ownership_smoke_metadata["target_block_count"],
        "ownership_smoke_min_per_category": ownership_smoke_metadata["min_per_category"],
        "ownership_smoke_block_count": ownership_smoke_metadata["block_count"],
        "ownership_smoke_category_counts": ownership_smoke_metadata["sample_category_counts"],
        "ownership_smoke_available_category_counts": ownership_smoke_metadata["available_category_counts"],
        "ownership_smoke_category_order": ownership_smoke_metadata["category_order"],
        "ownership_smoke_lookup_input_priority": ownership_smoke_metadata["lookup_input_priority"],
        **summary,
    }
    _write_json(summary_path, summary_payload)

    manifest_payload: dict[str, Any] = {
        **summary_payload,
        "diagnostic_columns": list(NULL_RIC_DIAGNOSTIC_COLUMNS),
        "review_columns": list(NULL_RIC_REVIEW_COLUMNS),
        "ownership_smoke_columns": list(OWNERSHIP_SMOKE_SAMPLE_COLUMNS),
        "artifacts": {
            "refinitiv_null_ric_diagnostics_summary": str(summary_path),
            "refinitiv_null_ric_rescue_candidates_parquet": str(rescue_parquet_path),
            "refinitiv_null_ric_rescue_candidates_csv": str(rescue_csv_path),
            "refinitiv_ownership_smoke_testing_parquet": str(ownership_smoke_parquet_path),
            "refinitiv_ownership_smoke_testing_csv": str(ownership_smoke_csv_path),
            "refinitiv_null_ric_diagnostics_manifest": str(manifest_path),
        },
    }

    if emit_review_workbook:
        write_refinitiv_null_ric_diagnostics_workbook(
            _coerce_text_columns(review_df, NULL_RIC_REVIEW_TEXT_COLUMNS),
            review_workbook_path,
            readme_payload=manifest_payload,
            text_columns=NULL_RIC_REVIEW_TEXT_COLUMNS,
            sheet_name="failed_lookup_review",
        )
        manifest_payload["artifacts"]["refinitiv_null_ric_rescue_candidates_review_xlsx"] = str(
            review_workbook_path
        )
    if emit_ownership_smoke_workbook:
        write_refinitiv_ownership_smoke_testing_workbook(
            ownership_smoke_df,
            ownership_smoke_workbook_path,
            readme_payload=manifest_payload,
        )
        manifest_payload["artifacts"]["refinitiv_ownership_smoke_testing_xlsx"] = str(
            ownership_smoke_workbook_path
        )

    _write_json(manifest_path, manifest_payload)

    paths: dict[str, Path] = {
        "refinitiv_null_ric_diagnostics_summary": summary_path,
        "refinitiv_null_ric_rescue_candidates_parquet": rescue_parquet_path,
        "refinitiv_null_ric_rescue_candidates_csv": rescue_csv_path,
        "refinitiv_ownership_smoke_testing_parquet": ownership_smoke_parquet_path,
        "refinitiv_ownership_smoke_testing_csv": ownership_smoke_csv_path,
        "refinitiv_null_ric_diagnostics_manifest": manifest_path,
    }
    if emit_review_workbook:
        paths["refinitiv_null_ric_rescue_candidates_review_xlsx"] = review_workbook_path
    if emit_ownership_smoke_workbook:
        paths["refinitiv_ownership_smoke_testing_xlsx"] = ownership_smoke_workbook_path
    return paths
