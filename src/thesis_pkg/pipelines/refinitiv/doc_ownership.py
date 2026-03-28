from __future__ import annotations

from collections import defaultdict
import datetime as dt
from pathlib import Path
from typing import Any

import polars as pl

from thesis_pkg.io.excel import write_refinitiv_lm2011_doc_ownership_workbook
from thesis_pkg.pipelines.refinitiv_bridge_pipeline import (
    _build_explicit_schema_df,
    _cast_df_to_schema,
    _normalize_lookup_text,
    _normalize_workbook_scalar,
    _write_workbook_or_reuse_locked_output,
)


DOC_OWNERSHIP_REQUEST_COLUMNS: tuple[str, ...] = (
    "doc_id",
    "filing_date",
    "KYPERMNO",
    "authoritative_ric",
    "authority_decision_status",
    "target_quarter_end",
    "target_effective_date",
    "fallback_window_start",
    "fallback_window_end",
    "retrieval_eligible",
    "retrieval_exclusion_reason",
)

DOC_OWNERSHIP_RAW_COLUMNS: tuple[str, ...] = (
    "doc_id",
    "filing_date",
    "KYPERMNO",
    "authoritative_ric",
    "authority_decision_status",
    "target_quarter_end",
    "target_effective_date",
    "request_stage",
    "response_date",
    "response_date_is_imputed",
    "returned_category",
    "returned_category_normalized",
    "returned_value",
    "is_institutional_category",
)

DOC_OWNERSHIP_FINAL_COLUMNS: tuple[str, ...] = (
    "doc_id",
    "filing_date",
    "KYPERMNO",
    "authoritative_ric",
    "authority_decision_status",
    "target_quarter_end",
    "target_effective_date",
    "selected_response_date",
    "returned_category",
    "institutional_ownership_pct",
    "retrieval_status",
    "fallback_used",
)

DOC_OWNERSHIP_INPUT_FIELDS: tuple[str, ...] = (
    "doc_id",
    "authoritative_ric",
    "target_quarter_end",
    "target_effective_date",
    "fallback_window_start",
    "fallback_window_end",
    "filing_date",
    "KYPERMNO",
    "authority_decision_status",
)

DOC_OWNERSHIP_BLOCK_HEADERS: tuple[str, ...] = (
    "input_data",
    "returned_ric",
    "returned_date",
    "returned_value",
    "returned_category",
)

DOC_OWNERSHIP_RETRIEVAL_SHEET_PREFIX = "ownership_retrieval"
DOC_OWNERSHIP_EXACT_STAGE = "EXACT"
DOC_OWNERSHIP_FALLBACK_STAGE = "FALLBACK"
DOC_OWNERSHIP_MAX_FALLBACK_DAYS = 45


def _doc_ownership_request_schema() -> dict[str, pl.DataType]:
    return {
        "doc_id": pl.Utf8,
        "filing_date": pl.Date,
        "KYPERMNO": pl.Utf8,
        "authoritative_ric": pl.Utf8,
        "authority_decision_status": pl.Utf8,
        "target_quarter_end": pl.Date,
        "target_effective_date": pl.Date,
        "fallback_window_start": pl.Date,
        "fallback_window_end": pl.Date,
        "retrieval_eligible": pl.Boolean,
        "retrieval_exclusion_reason": pl.Utf8,
    }


def _doc_ownership_raw_schema() -> dict[str, pl.DataType]:
    return {
        "doc_id": pl.Utf8,
        "filing_date": pl.Date,
        "KYPERMNO": pl.Utf8,
        "authoritative_ric": pl.Utf8,
        "authority_decision_status": pl.Utf8,
        "target_quarter_end": pl.Date,
        "target_effective_date": pl.Date,
        "request_stage": pl.Utf8,
        "response_date": pl.Date,
        "response_date_is_imputed": pl.Boolean,
        "returned_category": pl.Utf8,
        "returned_category_normalized": pl.Utf8,
        "returned_value": pl.Float64,
        "is_institutional_category": pl.Boolean,
    }


def _doc_ownership_final_schema() -> dict[str, pl.DataType]:
    return {
        "doc_id": pl.Utf8,
        "filing_date": pl.Date,
        "KYPERMNO": pl.Utf8,
        "authoritative_ric": pl.Utf8,
        "authority_decision_status": pl.Utf8,
        "target_quarter_end": pl.Date,
        "target_effective_date": pl.Date,
        "selected_response_date": pl.Date,
        "returned_category": pl.Utf8,
        "institutional_ownership_pct": pl.Float64,
        "retrieval_status": pl.Utf8,
        "fallback_used": pl.Boolean,
    }


def _empty_df(columns: tuple[str, ...], schema: dict[str, pl.DataType]) -> pl.DataFrame:
    return _cast_df_to_schema(pl.DataFrame(schema={name: dtype for name, dtype in schema.items()}), schema).select(
        list(columns)
    )


def _normalize_date_value(value: Any) -> dt.date | None:
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    normalized = _normalize_workbook_scalar(value)
    if normalized is None:
        return None
    try:
        return dt.date.fromisoformat(normalized)
    except ValueError:
        try:
            return dt.datetime.fromisoformat(normalized).date()
        except ValueError:
            pass
    if " " in normalized:
        date_part = normalized.split(" ", 1)[0].strip()
        try:
            return dt.date.fromisoformat(date_part)
        except ValueError:
            return None
    return None


def _normalize_float_value(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    normalized = _normalize_workbook_scalar(value)
    if normalized is None:
        return None
    try:
        return float(normalized)
    except ValueError:
        return None


def _normalize_category(value: Any) -> str | None:
    normalized = _normalize_lookup_text(_normalize_workbook_scalar(value))
    if normalized is None:
        return None
    return " ".join(normalized.split())


def _is_institutional_category(value: str | None) -> bool:
    return value is not None and value.casefold() == "Holdings by Institutions".casefold()


def _clean_institutional_value(value: float | None) -> float | None:
    if value is None:
        return None
    if value < 0:
        return None
    return min(float(value), 100.0)


def _normalize_kypermno(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        return _normalize_lookup_text(value)
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if int(value) == value:
            return str(int(value))
        return str(value)
    return _normalize_lookup_text(str(value))


def _most_recent_quarter_end_before(filing_date: dt.date) -> dt.date:
    quarter_start_month = ((filing_date.month - 1) // 3) * 3 + 1
    quarter_start = dt.date(filing_date.year, quarter_start_month, 1)
    return quarter_start - dt.timedelta(days=1)


def _target_effective_date_for_quarter_end(target_quarter_end: dt.date) -> dt.date:
    return target_quarter_end + dt.timedelta(days=1)


def _clip_date_lower(value: dt.date | None, lower_bound: dt.date | None) -> dt.date | None:
    if value is None or lower_bound is None:
        return value
    return max(value, lower_bound)


def _clip_date_upper(value: dt.date | None, upper_bound: dt.date | None) -> dt.date | None:
    if value is None or upper_bound is None:
        return value
    return min(value, upper_bound)


def _read_doc_filing_artifact(parquet_path: Path | str) -> pl.DataFrame:
    parquet_path = Path(parquet_path)
    if not parquet_path.exists():
        raise FileNotFoundError(f"doc filing artifact not found: {parquet_path}")

    schema = pl.scan_parquet(parquet_path).collect_schema()
    permno_column = "kypermno" if "kypermno" in schema else "KYPERMNO" if "KYPERMNO" in schema else None
    if permno_column is None:
        raise ValueError("doc filing artifact missing both kypermno and KYPERMNO")
    required = {"doc_id", "filing_date", permno_column}
    missing = sorted(name for name in required if name not in schema)
    if missing:
        raise ValueError(f"doc filing artifact missing required columns: {missing}")

    df = (
        pl.read_parquet(parquet_path)
        .select(
            pl.col("doc_id").cast(pl.Utf8, strict=False).alias("doc_id"),
            pl.col("filing_date").cast(pl.Date, strict=False).alias("filing_date"),
            pl.col(permno_column).alias("_kypermno_raw"),
        )
        .with_columns(
            pl.col("_kypermno_raw").map_elements(_normalize_kypermno, return_dtype=pl.Utf8).alias("KYPERMNO")
        )
        .drop("_kypermno_raw")
    )
    duplicate_doc_ids = (
        df.drop_nulls(subset=["doc_id"])
        .group_by("doc_id")
        .len()
        .filter(pl.col("len") > 1)
        .get_column("doc_id")
        .to_list()
    )
    if duplicate_doc_ids:
        raise ValueError(f"doc filing artifact is not unique on doc_id: {duplicate_doc_ids[:10]}")
    return df


def _read_authority_decisions_artifact(parquet_path: Path | str) -> pl.DataFrame:
    parquet_path = Path(parquet_path)
    if not parquet_path.exists():
        raise FileNotFoundError(f"authority decisions artifact not found: {parquet_path}")
    df = pl.read_parquet(parquet_path)
    required_columns = (
        "KYPERMNO",
        "authoritative_ric",
        "authoritative_source_family",
        "authority_decision_status",
        "requires_review",
    )
    missing = [name for name in required_columns if name not in df.columns]
    if missing:
        raise ValueError(f"authority decisions artifact missing required columns: {missing}")
    return df.select(
        pl.col("KYPERMNO").cast(pl.Utf8, strict=False).alias("KYPERMNO"),
        pl.col("authoritative_ric").cast(pl.Utf8, strict=False).alias("authoritative_ric"),
        pl.col("authoritative_source_family").cast(pl.Utf8, strict=False).alias("authoritative_source_family"),
        pl.col("authority_decision_status").cast(pl.Utf8, strict=False).alias("authority_decision_status"),
        pl.col("requires_review").cast(pl.Boolean, strict=False).fill_null(False).alias("requires_review"),
    )


def _read_authority_exceptions_artifact(parquet_path: Path | str) -> pl.DataFrame:
    parquet_path = Path(parquet_path)
    if not parquet_path.exists():
        raise FileNotFoundError(f"authority exceptions artifact not found: {parquet_path}")
    df = pl.read_parquet(parquet_path)
    required_columns = (
        "KYPERMNO",
        "authoritative_ric",
        "authoritative_source_family",
        "authority_window_start_date",
        "authority_window_end_date",
        "authority_exception_status",
    )
    missing = [name for name in required_columns if name not in df.columns]
    if missing:
        raise ValueError(f"authority exceptions artifact missing required columns: {missing}")
    return df.select(
        pl.col("KYPERMNO").cast(pl.Utf8, strict=False).alias("KYPERMNO"),
        pl.col("authoritative_ric").cast(pl.Utf8, strict=False).alias("authoritative_ric"),
        pl.col("authoritative_source_family").cast(pl.Utf8, strict=False).alias("authoritative_source_family"),
        pl.col("authority_window_start_date").cast(pl.Date, strict=False).alias("authority_window_start_date"),
        pl.col("authority_window_end_date").cast(pl.Date, strict=False).alias("authority_window_end_date"),
        pl.col("authority_exception_status").cast(pl.Utf8, strict=False).alias("authority_exception_status"),
    )


def build_refinitiv_lm2011_doc_ownership_requests(
    doc_filing_df: pl.DataFrame,
    authority_decisions_df: pl.DataFrame,
    authority_exceptions_df: pl.DataFrame,
    *,
    request_min_date: dt.date | None = None,
    request_max_date: dt.date | None = None,
) -> pl.DataFrame:
    if (
        request_min_date is not None
        and request_max_date is not None
        and request_min_date > request_max_date
    ):
        raise ValueError("request_min_date must be <= request_max_date")

    decisions_by_permno = {
        _normalize_kypermno(row["KYPERMNO"]): row
        for row in authority_decisions_df.to_dicts()
        if _normalize_kypermno(row.get("KYPERMNO")) is not None
    }
    exceptions_by_permno: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in authority_exceptions_df.to_dicts():
        kypermno = _normalize_kypermno(row.get("KYPERMNO"))
        if kypermno is None:
            continue
        exceptions_by_permno[kypermno].append(row)
    for rows in exceptions_by_permno.values():
        rows.sort(
            key=lambda row: (
                _normalize_date_value(row.get("authority_window_start_date")) or dt.date.min,
                _normalize_date_value(row.get("authority_window_end_date")) or dt.date.max,
                _normalize_lookup_text(row.get("authoritative_ric")) or "",
            )
        )

    rows: list[dict[str, Any]] = []
    for record in doc_filing_df.to_dicts():
        doc_id = _normalize_lookup_text(record.get("doc_id"))
        filing_date = _normalize_date_value(record.get("filing_date"))
        kypermno = _normalize_kypermno(record.get("KYPERMNO") if "KYPERMNO" in record else record.get("kypermno"))
        target_quarter_end = None if filing_date is None else _most_recent_quarter_end_before(filing_date)
        target_effective_date = (
            None if target_quarter_end is None else _target_effective_date_for_quarter_end(target_quarter_end)
        )
        fallback_window_start = target_effective_date
        fallback_window_end = (
            None
            if filing_date is None or target_effective_date is None
            else min(filing_date, target_effective_date + dt.timedelta(days=DOC_OWNERSHIP_MAX_FALLBACK_DAYS))
        )

        authoritative_ric: str | None = None
        retrieval_exclusion_reason: str | None = None
        authority_decision_status: str | None = None
        decision_row = None if kypermno is None else decisions_by_permno.get(kypermno)

        if doc_id is None:
            retrieval_exclusion_reason = "missing_doc_id"
        elif filing_date is None:
            retrieval_exclusion_reason = "missing_filing_date"
        elif kypermno is None:
            retrieval_exclusion_reason = "missing_kypermno"
        elif decision_row is None:
            retrieval_exclusion_reason = "no_authority_decision_for_kypermno"
        else:
            authority_decision_status = _normalize_lookup_text(decision_row.get("authority_decision_status"))
            decision_authoritative_ric = _normalize_lookup_text(decision_row.get("authoritative_ric"))
            if authority_decision_status == "DATE_VARYING_CONVENTIONAL_EXCEPTION":
                matching_rows = [
                    row
                    for row in exceptions_by_permno.get(kypermno, [])
                    if (
                        target_quarter_end is not None
                        and _normalize_date_value(row.get("authority_window_start_date")) is not None
                        and _normalize_date_value(row.get("authority_window_end_date")) is not None
                        and _normalize_date_value(row.get("authority_window_start_date"))
                        <= target_quarter_end
                        <= _normalize_date_value(row.get("authority_window_end_date"))
                    )
                ]
                if len(matching_rows) == 1:
                    authoritative_ric = _normalize_lookup_text(matching_rows[0].get("authoritative_ric"))
                    if authoritative_ric is None:
                        retrieval_exclusion_reason = "matched_exception_window_missing_authoritative_ric"
                elif len(matching_rows) > 1:
                    retrieval_exclusion_reason = "multiple_exception_windows_for_target_quarter"
                else:
                    retrieval_exclusion_reason = "no_exception_window_match_for_target_quarter"
            elif authority_decision_status == "REVIEW_REQUIRED":
                retrieval_exclusion_reason = "authority_review_required"
            elif decision_authoritative_ric is not None:
                authoritative_ric = decision_authoritative_ric
            else:
                retrieval_exclusion_reason = "no_authoritative_ric"

        if retrieval_exclusion_reason is None and target_effective_date is not None:
            if request_min_date is not None and target_effective_date < request_min_date:
                retrieval_exclusion_reason = "target_effective_date_before_request_min_date"
            elif request_max_date is not None and target_effective_date > request_max_date:
                retrieval_exclusion_reason = "target_effective_date_after_request_max_date"

        if retrieval_exclusion_reason is None:
            fallback_window_start = _clip_date_lower(fallback_window_start, request_min_date)
            fallback_window_end = _clip_date_upper(fallback_window_end, request_max_date)

        retrieval_eligible = authoritative_ric is not None and retrieval_exclusion_reason is None
        rows.append(
            {
                "doc_id": doc_id,
                "filing_date": filing_date,
                "KYPERMNO": kypermno,
                "authoritative_ric": authoritative_ric,
                "authority_decision_status": authority_decision_status,
                "target_quarter_end": target_quarter_end,
                "target_effective_date": target_effective_date,
                "fallback_window_start": fallback_window_start,
                "fallback_window_end": fallback_window_end,
                "retrieval_eligible": retrieval_eligible,
                "retrieval_exclusion_reason": retrieval_exclusion_reason,
            }
        )

    return _build_explicit_schema_df(rows, _doc_ownership_request_schema()).select(DOC_OWNERSHIP_REQUEST_COLUMNS)


def _doc_ownership_universe_detail_schema() -> dict[str, pl.DataType]:
    return {
        "doc_id": pl.Utf8,
        "cik_10": pl.Utf8,
        "filing_date": pl.Date,
        "filing_year": pl.Int32,
        "normalized_form": pl.Utf8,
        "KYPERMNO": pl.Utf8,
        "in_backbone": pl.Boolean,
        "in_request": pl.Boolean,
        "in_final": pl.Boolean,
        "request_retrieval_eligible": pl.Boolean,
        "request_retrieval_exclusion_reason": pl.Utf8,
        "final_retrieval_status": pl.Utf8,
        "final_has_nonnull_ownership": pl.Boolean,
        "final_institutional_ownership_pct": pl.Float64,
    }


def _doc_ownership_universe_breakdown_schema() -> dict[str, pl.DataType]:
    return {
        "filing_year": pl.Int32,
        "normalized_form": pl.Utf8,
        "backbone_doc_count": pl.Int64,
        "request_doc_count": pl.Int64,
        "final_doc_count": pl.Int64,
        "final_nonnull_ownership_doc_count": pl.Int64,
        "backbone_only_doc_count": pl.Int64,
        "request_only_doc_count": pl.Int64,
        "final_only_doc_count": pl.Int64,
    }


def _assert_unique_doc_id_membership(df: pl.DataFrame, *, label: str) -> None:
    if "doc_id" not in df.columns:
        raise ValueError(f"{label} missing required column: doc_id")
    duplicate_doc_ids = (
        df.drop_nulls(subset=["doc_id"])
        .group_by("doc_id")
        .len()
        .filter(pl.col("len") > 1)
        .get_column("doc_id")
        .to_list()
    )
    if duplicate_doc_ids:
        raise ValueError(f"{label} must be unique on doc_id: {duplicate_doc_ids[:10]}")


def _normalize_optional_text_expr(column_name: str, *, enabled: bool) -> pl.Expr:
    if not enabled:
        return pl.lit(None, dtype=pl.Utf8)
    return pl.col(column_name).cast(pl.Utf8, strict=False).map_elements(_normalize_lookup_text, return_dtype=pl.Utf8)


def _prepare_backbone_membership_df(backbone_df: pl.DataFrame) -> pl.DataFrame:
    schema = backbone_df.schema
    permno_column = "KYPERMNO" if "KYPERMNO" in schema else "kypermno" if "kypermno" in schema else None
    form_column = "normalized_form" if "normalized_form" in schema else None
    if form_column is None and "document_type_filename" in schema:
        form_column = "document_type_filename"
    prepared = (
        backbone_df.select(
            _normalize_optional_text_expr("doc_id", enabled=True).alias("doc_id"),
            _normalize_optional_text_expr("cik_10", enabled="cik_10" in schema).alias("cik_10"),
            (
                pl.col("filing_date").cast(pl.Date, strict=False)
                if "filing_date" in schema
                else pl.lit(None, dtype=pl.Date)
            ).alias("filing_date"),
            _normalize_optional_text_expr(form_column, enabled=form_column is not None).alias("normalized_form"),
            (
                pl.col(permno_column)
                .map_elements(_normalize_kypermno, return_dtype=pl.Utf8)
                if permno_column is not None
                else pl.lit(None, dtype=pl.Utf8)
            ).alias("KYPERMNO"),
        )
        .drop_nulls(subset=["doc_id"])
    )
    _assert_unique_doc_id_membership(prepared, label="backbone_df")
    return prepared.with_columns(
        pl.col("filing_date").dt.year().cast(pl.Int32, strict=False).alias("filing_year"),
        pl.lit(True).alias("in_backbone"),
    )


def _prepare_request_membership_df(request_df: pl.DataFrame) -> pl.DataFrame:
    schema = request_df.schema
    permno_column = "KYPERMNO" if "KYPERMNO" in schema else "kypermno" if "kypermno" in schema else None
    prepared = (
        request_df.select(
            _normalize_optional_text_expr("doc_id", enabled=True).alias("doc_id"),
            (
                pl.col("filing_date").cast(pl.Date, strict=False)
                if "filing_date" in schema
                else pl.lit(None, dtype=pl.Date)
            ).alias("request_filing_date"),
            (
                pl.col(permno_column)
                .map_elements(_normalize_kypermno, return_dtype=pl.Utf8)
                if permno_column is not None
                else pl.lit(None, dtype=pl.Utf8)
            ).alias("request_KYPERMNO"),
            (
                pl.col("retrieval_eligible").cast(pl.Boolean, strict=False)
                if "retrieval_eligible" in schema
                else pl.lit(None, dtype=pl.Boolean)
            ).alias("request_retrieval_eligible"),
            _normalize_optional_text_expr(
                "retrieval_exclusion_reason",
                enabled="retrieval_exclusion_reason" in schema,
            ).alias("request_retrieval_exclusion_reason"),
        )
        .drop_nulls(subset=["doc_id"])
    )
    _assert_unique_doc_id_membership(prepared, label="request_df")
    return prepared.with_columns(pl.lit(True).alias("in_request"))


def _prepare_final_membership_df(final_df: pl.DataFrame) -> pl.DataFrame:
    schema = final_df.schema
    permno_column = "KYPERMNO" if "KYPERMNO" in schema else "kypermno" if "kypermno" in schema else None
    ownership_value_col = (
        "institutional_ownership_pct"
        if "institutional_ownership_pct" in schema
        else "institutional_ownership"
        if "institutional_ownership" in schema
        else None
    )
    prepared = (
        final_df.select(
            _normalize_optional_text_expr("doc_id", enabled=True).alias("doc_id"),
            (
                pl.col("filing_date").cast(pl.Date, strict=False)
                if "filing_date" in schema
                else pl.lit(None, dtype=pl.Date)
            ).alias("final_filing_date"),
            (
                pl.col(permno_column)
                .map_elements(_normalize_kypermno, return_dtype=pl.Utf8)
                if permno_column is not None
                else pl.lit(None, dtype=pl.Utf8)
            ).alias("final_KYPERMNO"),
            _normalize_optional_text_expr("retrieval_status", enabled="retrieval_status" in schema).alias(
                "final_retrieval_status"
            ),
            (
                pl.col(ownership_value_col).cast(pl.Float64, strict=False)
                if ownership_value_col is not None
                else pl.lit(None, dtype=pl.Float64)
            ).alias("final_institutional_ownership_pct"),
        )
        .drop_nulls(subset=["doc_id"])
    )
    _assert_unique_doc_id_membership(prepared, label="final_df")
    return prepared.with_columns(
        pl.lit(True).alias("in_final"),
        pl.col("final_institutional_ownership_pct").is_not_null().alias("final_has_nonnull_ownership"),
    )


def _safe_rate(numerator: int, denominator: int | None) -> float | None:
    if denominator in (None, 0):
        return None
    return float(numerator) / float(denominator)


def _top_doc_mismatch_cik_counts(
    detail_df: pl.DataFrame,
    *,
    left_flag: str,
    right_flag: str,
    limit: int = 20,
) -> list[dict[str, Any]]:
    if "cik_10" not in detail_df.columns:
        return []
    return (
        detail_df.filter(
            pl.col("cik_10").is_not_null() & (pl.col(left_flag).fill_null(False) != pl.col(right_flag).fill_null(False))
        )
        .group_by("cik_10")
        .len()
        .sort("len", descending=True)
        .head(limit)
        .rename({"len": "doc_count"})
        .to_dicts()
    )


def _build_lm2011_doc_ownership_universe_diagnostics(
    backbone_df: pl.DataFrame,
    request_df: pl.DataFrame,
    final_df: pl.DataFrame | None = None,
) -> tuple[dict[str, pl.DataFrame], dict[str, Any]]:
    backbone_docs = _prepare_backbone_membership_df(backbone_df)
    request_docs = _prepare_request_membership_df(request_df)
    final_docs = (
        _prepare_final_membership_df(final_df)
        if final_df is not None
        else _empty_df(
            (
                "doc_id",
                "final_filing_date",
                "final_KYPERMNO",
                "final_retrieval_status",
                "final_institutional_ownership_pct",
                "in_final",
                "final_has_nonnull_ownership",
            ),
            {
                "doc_id": pl.Utf8,
                "final_filing_date": pl.Date,
                "final_KYPERMNO": pl.Utf8,
                "final_retrieval_status": pl.Utf8,
                "final_institutional_ownership_pct": pl.Float64,
                "in_final": pl.Boolean,
                "final_has_nonnull_ownership": pl.Boolean,
            },
        )
    )
    doc_index = (
        pl.concat(
            [
                backbone_docs.select("doc_id"),
                request_docs.select("doc_id"),
                final_docs.select("doc_id"),
            ],
            how="vertical_relaxed",
        )
        .drop_nulls(subset=["doc_id"])
        .unique(subset=["doc_id"], keep="first", maintain_order=True)
    )
    detail_df = (
        doc_index.join(backbone_docs, on="doc_id", how="left")
        .join(request_docs, on="doc_id", how="left")
        .join(final_docs, on="doc_id", how="left")
        .with_columns(
            pl.coalesce(
                [
                    pl.col("filing_date"),
                    pl.col("request_filing_date"),
                    pl.col("final_filing_date"),
                ]
            ).alias("filing_date"),
            pl.coalesce([pl.col("KYPERMNO"), pl.col("request_KYPERMNO"), pl.col("final_KYPERMNO")]).alias("KYPERMNO"),
            pl.coalesce([pl.col("in_backbone"), pl.lit(False)]).alias("in_backbone"),
            pl.coalesce([pl.col("in_request"), pl.lit(False)]).alias("in_request"),
            pl.coalesce([pl.col("in_final"), pl.lit(False)]).alias("in_final"),
            pl.coalesce([pl.col("final_has_nonnull_ownership"), pl.lit(False)]).alias("final_has_nonnull_ownership"),
        )
        .with_columns(pl.col("filing_date").dt.year().cast(pl.Int32, strict=False).alias("filing_year"))
        .select(
            "doc_id",
            "cik_10",
            "filing_date",
            "filing_year",
            "normalized_form",
            "KYPERMNO",
            "in_backbone",
            "in_request",
            "in_final",
            "request_retrieval_eligible",
            "request_retrieval_exclusion_reason",
            "final_retrieval_status",
            "final_has_nonnull_ownership",
            "final_institutional_ownership_pct",
        )
        .sort("filing_year", "normalized_form", "doc_id")
    )
    detail_df = _build_explicit_schema_df(detail_df.to_dicts(), _doc_ownership_universe_detail_schema()).select(
        tuple(_doc_ownership_universe_detail_schema())
    )

    breakdown_df = (
        detail_df.group_by("filing_year", "normalized_form")
        .agg(
            pl.col("in_backbone").cast(pl.Int64).sum().alias("backbone_doc_count"),
            pl.col("in_request").cast(pl.Int64).sum().alias("request_doc_count"),
            pl.col("in_final").cast(pl.Int64).sum().alias("final_doc_count"),
            pl.col("final_has_nonnull_ownership").cast(pl.Int64).sum().alias("final_nonnull_ownership_doc_count"),
            (pl.col("in_backbone") & pl.col("in_request").not_()).cast(pl.Int64).sum().alias("backbone_only_doc_count"),
            (pl.col("in_request") & pl.col("in_backbone").not_()).cast(pl.Int64).sum().alias("request_only_doc_count"),
            (pl.col("in_final") & pl.col("in_request").not_()).cast(pl.Int64).sum().alias("final_only_doc_count"),
        )
        .sort("filing_year", "normalized_form")
    )
    if final_df is None:
        breakdown_df = breakdown_df.with_columns(
            pl.lit(None, dtype=pl.Int64).alias("final_doc_count"),
            pl.lit(None, dtype=pl.Int64).alias("final_nonnull_ownership_doc_count"),
            pl.lit(None, dtype=pl.Int64).alias("final_only_doc_count"),
        )
    breakdown_df = _build_explicit_schema_df(
        breakdown_df.to_dicts(),
        _doc_ownership_universe_breakdown_schema(),
    ).select(tuple(_doc_ownership_universe_breakdown_schema()))

    backbone_doc_count = int(backbone_docs.height)
    request_doc_count = int(request_docs.height)
    final_doc_count = int(final_docs.height) if final_df is not None else None
    backbone_request_overlap = int(detail_df.filter(pl.col("in_backbone") & pl.col("in_request")).height)
    backbone_only_doc_count = int(detail_df.filter(pl.col("in_backbone") & pl.col("in_request").not_()).height)
    request_only_doc_count = int(detail_df.filter(pl.col("in_request") & pl.col("in_backbone").not_()).height)
    backbone_request_doc_sets_equal = backbone_only_doc_count == 0 and request_only_doc_count == 0

    final_backbone_overlap = None
    final_request_overlap = None
    backbone_final_doc_sets_equal = None
    request_final_doc_sets_equal = None
    final_only_doc_count = None
    request_missing_from_final_doc_count = None
    backbone_missing_from_final_doc_count = None
    final_missing_from_backbone_doc_count = None
    final_nonnull_ownership_doc_count = None
    all_doc_sets_equal = None
    retrieval_status_counts: dict[str, int] | None = None
    if final_df is not None:
        final_backbone_overlap = int(detail_df.filter(pl.col("in_backbone") & pl.col("in_final")).height)
        final_request_overlap = int(detail_df.filter(pl.col("in_request") & pl.col("in_final")).height)
        request_missing_from_final_doc_count = int(
            detail_df.filter(pl.col("in_request") & pl.col("in_final").not_()).height
        )
        final_only_doc_count = int(detail_df.filter(pl.col("in_final") & pl.col("in_request").not_()).height)
        backbone_missing_from_final_doc_count = int(
            detail_df.filter(pl.col("in_backbone") & pl.col("in_final").not_()).height
        )
        final_missing_from_backbone_doc_count = int(
            detail_df.filter(pl.col("in_final") & pl.col("in_backbone").not_()).height
        )
        final_nonnull_ownership_doc_count = int(detail_df.filter(pl.col("final_has_nonnull_ownership")).height)
        backbone_final_doc_sets_equal = (
            backbone_missing_from_final_doc_count == 0 and final_missing_from_backbone_doc_count == 0
        )
        request_final_doc_sets_equal = request_missing_from_final_doc_count == 0 and final_only_doc_count == 0
        all_doc_sets_equal = bool(backbone_request_doc_sets_equal and backbone_final_doc_sets_equal and request_final_doc_sets_equal)
        retrieval_status_counts = {
            str(row["final_retrieval_status"]): int(row["len"])
            for row in detail_df.filter(pl.col("final_retrieval_status").is_not_null())
            .group_by("final_retrieval_status")
            .len()
            .to_dicts()
        }

    summary = {
        "backbone_doc_count": backbone_doc_count,
        "request_doc_count": request_doc_count,
        "final_doc_count": final_doc_count,
        "backbone_request_overlap_doc_count": backbone_request_overlap,
        "backbone_request_overlap_rate_of_backbone": _safe_rate(backbone_request_overlap, backbone_doc_count),
        "backbone_request_overlap_rate_of_request": _safe_rate(backbone_request_overlap, request_doc_count),
        "backbone_only_doc_count": backbone_only_doc_count,
        "request_only_doc_count": request_only_doc_count,
        "backbone_request_doc_sets_equal": backbone_request_doc_sets_equal,
        "final_backbone_overlap_doc_count": final_backbone_overlap,
        "final_request_overlap_doc_count": final_request_overlap,
        "request_missing_from_final_doc_count": request_missing_from_final_doc_count,
        "final_only_doc_count": final_only_doc_count,
        "backbone_missing_from_final_doc_count": backbone_missing_from_final_doc_count,
        "final_missing_from_backbone_doc_count": final_missing_from_backbone_doc_count,
        "final_nonnull_ownership_doc_count": final_nonnull_ownership_doc_count,
        "backbone_final_doc_sets_equal": backbone_final_doc_sets_equal,
        "request_final_doc_sets_equal": request_final_doc_sets_equal,
        "all_doc_sets_equal": all_doc_sets_equal,
        "retrieval_status_counts": retrieval_status_counts,
        "backbone_request_mismatch_cik_counts_top": _top_doc_mismatch_cik_counts(
            detail_df,
            left_flag="in_backbone",
            right_flag="in_request",
        ),
        "request_final_mismatch_cik_counts_top": (
            _top_doc_mismatch_cik_counts(detail_df, left_flag="in_request", right_flag="in_final")
            if final_df is not None
            else []
        ),
    }
    return {"detail": detail_df, "breakdown": breakdown_df}, summary


def _request_readme_payload(request_df: pl.DataFrame, *, request_stage: str) -> dict[str, Any]:
    return {
        "pipeline_name": f"refinitiv_lm2011_doc_ownership_{request_stage.lower()}",
        "request_stage": request_stage,
        "request_count": int(request_df.height),
        "eligible_request_count": int(
            request_df.filter(pl.col("retrieval_eligible").fill_null(False)).select(pl.len()).item()
        )
        if request_df.height
        else 0,
        "request_block_headers": list(DOC_OWNERSHIP_BLOCK_HEADERS),
        "visible_input_field_order": list(DOC_OWNERSHIP_INPUT_FIELDS),
    }


def _matching_retrieval_sheets(sheet_names: list[str]) -> list[str]:
    matching = [
        sheet_name
        for sheet_name in sheet_names
        if sheet_name == DOC_OWNERSHIP_RETRIEVAL_SHEET_PREFIX
        or sheet_name.startswith(f"{DOC_OWNERSHIP_RETRIEVAL_SHEET_PREFIX}_")
    ]
    if not matching:
        raise ValueError("filled workbook missing ownership_retrieval sheet(s)")
    return sorted(matching)


def _parse_doc_ownership_filled_workbook(
    filled_workbook_path: Path | str,
    request_df: pl.DataFrame,
    *,
    request_stage: str,
) -> pl.DataFrame:
    from openpyxl import load_workbook

    eligible_request_df = request_df.filter(pl.col("retrieval_eligible").fill_null(False))
    if eligible_request_df.height == 0:
        return _empty_df(DOC_OWNERSHIP_RAW_COLUMNS, _doc_ownership_raw_schema())

    workbook = load_workbook(Path(filled_workbook_path), read_only=False, data_only=True)
    try:
        sheet_names = _matching_retrieval_sheets(workbook.sheetnames)
        requests_by_doc_id = {
            _normalize_lookup_text(row.get("doc_id")): row for row in eligible_request_df.to_dicts()
        }
        results: list[dict[str, Any]] = []
        for sheet_name in sheet_names:
            worksheet = workbook[sheet_name]
            for base_col in range(1, worksheet.max_column + 1, len(DOC_OWNERSHIP_BLOCK_HEADERS)):
                header_values = [
                    _normalize_workbook_scalar(worksheet.cell(row=1, column=base_col + offset).value)
                    for offset in range(len(DOC_OWNERSHIP_BLOCK_HEADERS))
                ]
                if all(value is None for value in header_values):
                    continue
                if header_values != list(DOC_OWNERSHIP_BLOCK_HEADERS):
                    raise ValueError(
                        "ownership_retrieval contains an unexpected request block header at "
                        f"{sheet_name} column {base_col}: {header_values}"
                    )

                block_values = {
                    field_name: worksheet.cell(row=2 + field_offset, column=base_col).value
                    for field_offset, field_name in enumerate(DOC_OWNERSHIP_INPUT_FIELDS)
                }
                doc_id = _normalize_lookup_text(block_values.get("doc_id"))
                if doc_id is None:
                    continue
                request_row = requests_by_doc_id.get(doc_id)
                if request_row is None:
                    raise ValueError(
                        "ownership_retrieval contains a block that cannot be matched to the request snapshot: "
                        f"{sheet_name} {block_values}"
                    )

                if request_stage == DOC_OWNERSHIP_EXACT_STAGE:
                    for excel_row in range(3, worksheet.max_row + 1):
                        response_date = _normalize_date_value(
                            worksheet.cell(row=excel_row, column=base_col + 2).value
                        )
                        returned_value = _normalize_float_value(
                            worksheet.cell(row=excel_row, column=base_col + 3).value
                        )
                        returned_category = _normalize_category(
                            worksheet.cell(row=excel_row, column=base_col + 4).value
                        )
                        if response_date is None and returned_value is None and returned_category is None:
                            break
                        results.append(
                            {
                                **{name: request_row.get(name) for name in DOC_OWNERSHIP_REQUEST_COLUMNS},
                                "request_stage": DOC_OWNERSHIP_EXACT_STAGE,
                                "response_date": response_date,
                                "response_date_is_imputed": False,
                                "returned_category": returned_category,
                                "returned_category_normalized": returned_category,
                                "returned_value": returned_value,
                                "is_institutional_category": _is_institutional_category(returned_category),
                            }
                        )
                elif request_stage == DOC_OWNERSHIP_FALLBACK_STAGE:
                    for excel_row in range(3, worksheet.max_row + 1):
                        response_date = _normalize_date_value(
                            worksheet.cell(row=excel_row, column=base_col + 2).value
                        )
                        returned_value = _normalize_float_value(
                            worksheet.cell(row=excel_row, column=base_col + 3).value
                        )
                        returned_category = _normalize_category(
                            worksheet.cell(row=excel_row, column=base_col + 4).value
                        )
                        if response_date is None and returned_value is None and returned_category is None:
                            break
                        results.append(
                            {
                                **{name: request_row.get(name) for name in DOC_OWNERSHIP_REQUEST_COLUMNS},
                                "request_stage": DOC_OWNERSHIP_FALLBACK_STAGE,
                                "response_date": response_date,
                                "response_date_is_imputed": False,
                                "returned_category": returned_category,
                                "returned_category_normalized": returned_category,
                                "returned_value": returned_value,
                                "is_institutional_category": _is_institutional_category(returned_category),
                            }
                        )
                else:
                    raise ValueError(f"unsupported request stage: {request_stage}")
    finally:
        workbook.close()

    if not results:
        return _empty_df(DOC_OWNERSHIP_RAW_COLUMNS, _doc_ownership_raw_schema())
    return (
        _build_explicit_schema_df(results, _doc_ownership_raw_schema())
        .select(DOC_OWNERSHIP_RAW_COLUMNS)
        .unique(
            subset=["doc_id", "request_stage", "response_date", "returned_category", "returned_value"],
            maintain_order=True,
        )
    )


def _select_exact_hits(exact_raw_df: pl.DataFrame) -> tuple[dict[str, dict[str, Any]], set[str]]:
    rows_by_doc_id: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in exact_raw_df.to_dicts():
        doc_id = _normalize_lookup_text(row.get("doc_id"))
        if doc_id is None or not bool(row.get("is_institutional_category")):
            continue
        cleaned_value = _clean_institutional_value(_normalize_float_value(row.get("returned_value")))
        if cleaned_value is None:
            continue
        rows_by_doc_id[doc_id].append({**row, "_cleaned_value": cleaned_value})

    selected: dict[str, dict[str, Any]] = {}
    conflict_docs: set[str] = set()
    for doc_id, rows in rows_by_doc_id.items():
        unique_values = {float(row["_cleaned_value"]) for row in rows}
        if len(unique_values) != 1:
            conflict_docs.add(doc_id)
            continue
        selected[doc_id] = rows[0]
    return selected, conflict_docs


def _select_fallback_hits(
    fallback_raw_df: pl.DataFrame,
    request_df: pl.DataFrame,
) -> tuple[dict[str, dict[str, Any]], set[str]]:
    fallback_window_end_by_doc_id = {
        _normalize_lookup_text(row.get("doc_id")): _normalize_date_value(row.get("fallback_window_end"))
        for row in request_df.to_dicts()
        if _normalize_lookup_text(row.get("doc_id")) is not None
    }
    rows_by_doc_id: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in fallback_raw_df.to_dicts():
        doc_id = _normalize_lookup_text(row.get("doc_id"))
        if doc_id is None or not bool(row.get("is_institutional_category")):
            continue
        cleaned_value = _clean_institutional_value(_normalize_float_value(row.get("returned_value")))
        response_date = _normalize_date_value(row.get("response_date"))
        fallback_window_end = fallback_window_end_by_doc_id.get(doc_id)
        if cleaned_value is None or response_date is None:
            continue
        if fallback_window_end is not None and response_date > fallback_window_end:
            continue
        rows_by_doc_id[doc_id].append({**row, "_cleaned_value": cleaned_value})

    selected: dict[str, dict[str, Any]] = {}
    conflict_docs: set[str] = set()
    for doc_id, rows in rows_by_doc_id.items():
        latest_date = max(
            _normalize_date_value(row["response_date"])
            for row in rows
            if row.get("response_date") is not None
        )
        latest_rows = [row for row in rows if _normalize_date_value(row.get("response_date")) == latest_date]
        unique_values = {float(row["_cleaned_value"]) for row in latest_rows}
        if len(unique_values) != 1:
            conflict_docs.add(doc_id)
            continue
        selected[doc_id] = latest_rows[0]
    return selected, conflict_docs


def _build_fallback_request_df(request_df: pl.DataFrame, exact_raw_df: pl.DataFrame) -> pl.DataFrame:
    exact_selected, exact_conflict_docs = _select_exact_hits(exact_raw_df)
    request_rows: list[dict[str, Any]] = []
    for row in request_df.to_dicts():
        doc_id = _normalize_lookup_text(row.get("doc_id"))
        if doc_id is None or not bool(row.get("retrieval_eligible")):
            continue
        if doc_id in exact_selected or doc_id in exact_conflict_docs:
            continue
        request_rows.append({name: row.get(name) for name in DOC_OWNERSHIP_REQUEST_COLUMNS})
    if not request_rows:
        return _empty_df(DOC_OWNERSHIP_REQUEST_COLUMNS, _doc_ownership_request_schema())
    return _build_explicit_schema_df(request_rows, _doc_ownership_request_schema()).select(DOC_OWNERSHIP_REQUEST_COLUMNS)


def _build_doc_ownership_output_tables(
    request_df: pl.DataFrame,
    exact_raw_df: pl.DataFrame,
    fallback_raw_df: pl.DataFrame,
) -> tuple[dict[str, pl.DataFrame], dict[str, Any]]:
    exact_selected, exact_conflict_docs = _select_exact_hits(exact_raw_df)
    fallback_selected, fallback_conflict_docs = _select_fallback_hits(fallback_raw_df, request_df)
    raw_df = (
        pl.concat([df for df in (exact_raw_df, fallback_raw_df) if df.height > 0], how="vertical_relaxed")
        if exact_raw_df.height or fallback_raw_df.height
        else _empty_df(DOC_OWNERSHIP_RAW_COLUMNS, _doc_ownership_raw_schema())
    )

    final_rows: list[dict[str, Any]] = []
    for row in request_df.to_dicts():
        doc_id = _normalize_lookup_text(row.get("doc_id"))
        if doc_id is None:
            continue
        retrieval_status = "NO_USABLE_INSTITUTIONAL_ROW"
        selected_response_date = None
        returned_category = None
        institutional_ownership_pct = None
        fallback_used = False

        if not bool(row.get("retrieval_eligible")):
            if _normalize_lookup_text(row.get("authority_decision_status")) == "REVIEW_REQUIRED":
                retrieval_status = "AUTHORITY_REVIEW_REQUIRED"
            else:
                retrieval_status = "NO_AUTHORITATIVE_RIC"
        elif doc_id in exact_conflict_docs or doc_id in fallback_conflict_docs:
            retrieval_status = "FALLBACK_CONFLICT_REVIEW"
        elif doc_id in exact_selected:
            selected_row = exact_selected[doc_id]
            retrieval_status = "EXACT_TARGET_HIT"
            selected_response_date = _normalize_date_value(selected_row.get("response_date"))
            returned_category = _normalize_category(selected_row.get("returned_category"))
            institutional_ownership_pct = _clean_institutional_value(
                _normalize_float_value(selected_row.get("returned_value"))
            )
        elif doc_id in fallback_selected:
            selected_row = fallback_selected[doc_id]
            retrieval_status = "FALLBACK_WINDOW_HIT"
            fallback_used = True
            selected_response_date = _normalize_date_value(selected_row.get("response_date"))
            returned_category = _normalize_category(selected_row.get("returned_category"))
            institutional_ownership_pct = _clean_institutional_value(
                _normalize_float_value(selected_row.get("returned_value"))
            )

        final_rows.append(
            {
                "doc_id": doc_id,
                "filing_date": _normalize_date_value(row.get("filing_date")),
                "KYPERMNO": _normalize_kypermno(row.get("KYPERMNO")),
                "authoritative_ric": _normalize_lookup_text(row.get("authoritative_ric")),
                "authority_decision_status": _normalize_lookup_text(row.get("authority_decision_status")),
                "target_quarter_end": _normalize_date_value(row.get("target_quarter_end")),
                "target_effective_date": _normalize_date_value(row.get("target_effective_date")),
                "selected_response_date": selected_response_date,
                "returned_category": returned_category,
                "institutional_ownership_pct": institutional_ownership_pct,
                "retrieval_status": retrieval_status,
                "fallback_used": fallback_used,
            }
        )

    final_df = _build_explicit_schema_df(final_rows, _doc_ownership_final_schema()).select(DOC_OWNERSHIP_FINAL_COLUMNS)
    summary = {
        "request_rows": int(request_df.height),
        "exact_raw_rows": int(exact_raw_df.height),
        "fallback_raw_rows": int(fallback_raw_df.height),
        "raw_rows": int(raw_df.height),
        "final_rows": int(final_df.height),
        "nonnull_institutional_ownership_rows": int(
            final_df.filter(pl.col("institutional_ownership_pct").is_not_null()).select(pl.len()).item()
        )
        if final_df.height
        else 0,
        "retrieval_status_counts": {
            row["retrieval_status"]: int(row["len"])
            for row in final_df.group_by("retrieval_status").len().to_dicts()
        }
        if final_df.height
        else {},
    }
    return {"raw": raw_df, "final": final_df}, summary


def run_refinitiv_lm2011_doc_ownership_exact_handoff_pipeline(
    *,
    doc_filing_artifact_path: Path | str,
    authority_decisions_artifact_path: Path | str,
    authority_exceptions_artifact_path: Path | str,
    output_dir: Path | str,
    request_min_date: dt.date | None = None,
    request_max_date: dt.date | None = None,
) -> dict[str, Path]:
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    request_df = build_refinitiv_lm2011_doc_ownership_requests(
        _read_doc_filing_artifact(doc_filing_artifact_path),
        _read_authority_decisions_artifact(authority_decisions_artifact_path),
        _read_authority_exceptions_artifact(authority_exceptions_artifact_path),
        request_min_date=request_min_date,
        request_max_date=request_max_date,
    )

    requests_parquet_path = output_dir / "refinitiv_lm2011_doc_ownership_exact_requests.parquet"
    handoff_xlsx_path = output_dir / "refinitiv_lm2011_doc_ownership_exact_handoff.xlsx"
    request_df.write_parquet(requests_parquet_path, compression="zstd")
    _write_workbook_or_reuse_locked_output(
        write_refinitiv_lm2011_doc_ownership_workbook,
        request_df,
        handoff_xlsx_path,
        readme_payload=_request_readme_payload(request_df, request_stage=DOC_OWNERSHIP_EXACT_STAGE),
        input_field_order=DOC_OWNERSHIP_INPUT_FIELDS,
        block_headers=DOC_OWNERSHIP_BLOCK_HEADERS,
        request_stage=DOC_OWNERSHIP_EXACT_STAGE,
    )
    return {
        "refinitiv_lm2011_doc_ownership_exact_requests_parquet": requests_parquet_path,
        "refinitiv_lm2011_doc_ownership_exact_handoff_xlsx": handoff_xlsx_path,
    }


def run_refinitiv_lm2011_doc_ownership_fallback_handoff_pipeline(
    *,
    exact_filled_workbook_path: Path | str,
    output_dir: Path | str,
) -> dict[str, Path]:
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    exact_requests_path = output_dir / "refinitiv_lm2011_doc_ownership_exact_requests.parquet"
    if not exact_requests_path.exists():
        raise FileNotFoundError(f"exact requests parquet not found: {exact_requests_path}")

    request_df = _cast_df_to_schema(
        pl.read_parquet(exact_requests_path).select(DOC_OWNERSHIP_REQUEST_COLUMNS),
        _doc_ownership_request_schema(),
    ).select(DOC_OWNERSHIP_REQUEST_COLUMNS)
    exact_raw_df = _parse_doc_ownership_filled_workbook(
        exact_filled_workbook_path,
        request_df,
        request_stage=DOC_OWNERSHIP_EXACT_STAGE,
    ).select(DOC_OWNERSHIP_RAW_COLUMNS)
    fallback_request_df = _build_fallback_request_df(request_df, exact_raw_df)

    exact_raw_path = output_dir / "refinitiv_lm2011_doc_ownership_exact_raw.parquet"
    fallback_requests_path = output_dir / "refinitiv_lm2011_doc_ownership_fallback_requests.parquet"
    fallback_handoff_path = output_dir / "refinitiv_lm2011_doc_ownership_fallback_handoff.xlsx"
    exact_raw_df.write_parquet(exact_raw_path, compression="zstd")
    fallback_request_df.write_parquet(fallback_requests_path, compression="zstd")
    _write_workbook_or_reuse_locked_output(
        write_refinitiv_lm2011_doc_ownership_workbook,
        fallback_request_df,
        fallback_handoff_path,
        readme_payload=_request_readme_payload(fallback_request_df, request_stage=DOC_OWNERSHIP_FALLBACK_STAGE),
        input_field_order=DOC_OWNERSHIP_INPUT_FIELDS,
        block_headers=DOC_OWNERSHIP_BLOCK_HEADERS,
        request_stage=DOC_OWNERSHIP_FALLBACK_STAGE,
    )
    return {
        "refinitiv_lm2011_doc_ownership_exact_raw_parquet": exact_raw_path,
        "refinitiv_lm2011_doc_ownership_fallback_requests_parquet": fallback_requests_path,
        "refinitiv_lm2011_doc_ownership_fallback_handoff_xlsx": fallback_handoff_path,
    }


def run_refinitiv_lm2011_doc_ownership_finalize_pipeline(
    *,
    output_dir: Path | str,
    fallback_filled_workbook_path: Path | str | None = None,
) -> dict[str, Path]:
    from thesis_pkg.pipelines.refinitiv.lseg_api_common import write_parquet_atomic

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    exact_requests_path = output_dir / "refinitiv_lm2011_doc_ownership_exact_requests.parquet"
    exact_raw_path = output_dir / "refinitiv_lm2011_doc_ownership_exact_raw.parquet"
    fallback_requests_path = output_dir / "refinitiv_lm2011_doc_ownership_fallback_requests.parquet"
    fallback_raw_path = output_dir / "refinitiv_lm2011_doc_ownership_fallback_raw.parquet"
    if not exact_requests_path.exists():
        raise FileNotFoundError(f"exact requests parquet not found: {exact_requests_path}")
    if not exact_raw_path.exists():
        raise FileNotFoundError(f"exact raw parquet not found: {exact_raw_path}")
    if not fallback_requests_path.exists():
        raise FileNotFoundError(f"fallback requests parquet not found: {fallback_requests_path}")

    request_df = _cast_df_to_schema(
        pl.read_parquet(exact_requests_path).select(DOC_OWNERSHIP_REQUEST_COLUMNS),
        _doc_ownership_request_schema(),
    ).select(DOC_OWNERSHIP_REQUEST_COLUMNS)
    exact_raw_df = _cast_df_to_schema(
        pl.read_parquet(exact_raw_path).select(DOC_OWNERSHIP_RAW_COLUMNS),
        _doc_ownership_raw_schema(),
    ).select(DOC_OWNERSHIP_RAW_COLUMNS)
    fallback_request_df = _cast_df_to_schema(
        pl.read_parquet(fallback_requests_path).select(DOC_OWNERSHIP_REQUEST_COLUMNS),
        _doc_ownership_request_schema(),
    ).select(DOC_OWNERSHIP_REQUEST_COLUMNS)

    fallback_raw_df = _empty_df(DOC_OWNERSHIP_RAW_COLUMNS, _doc_ownership_raw_schema())
    fallback_eligible_count = (
        int(fallback_request_df.filter(pl.col("retrieval_eligible").fill_null(False)).select(pl.len()).item())
        if fallback_request_df.height
        else 0
    )
    if fallback_raw_path.exists():
        fallback_raw_df = _cast_df_to_schema(
            pl.read_parquet(fallback_raw_path).select(DOC_OWNERSHIP_RAW_COLUMNS),
            _doc_ownership_raw_schema(),
        ).select(DOC_OWNERSHIP_RAW_COLUMNS)
    elif fallback_eligible_count > 0:
        if fallback_filled_workbook_path is None:
            raise FileNotFoundError(
                "fallback filled workbook path is required when fallback requests exist and fallback raw parquet is absent"
            )
        fallback_raw_df = _parse_doc_ownership_filled_workbook(
            fallback_filled_workbook_path,
            fallback_request_df,
            request_stage=DOC_OWNERSHIP_FALLBACK_STAGE,
        ).select(DOC_OWNERSHIP_RAW_COLUMNS)

    tables, _ = _build_doc_ownership_output_tables(request_df, exact_raw_df, fallback_raw_df)
    raw_path = output_dir / "refinitiv_lm2011_doc_ownership_raw.parquet"
    final_path = output_dir / "refinitiv_lm2011_doc_ownership.parquet"
    write_parquet_atomic(fallback_raw_df, fallback_raw_path)
    write_parquet_atomic(tables["raw"], raw_path)
    write_parquet_atomic(tables["final"], final_path)
    return {
        "refinitiv_lm2011_doc_ownership_fallback_raw_parquet": fallback_raw_path,
        "refinitiv_lm2011_doc_ownership_raw_parquet": raw_path,
        "refinitiv_lm2011_doc_ownership_parquet": final_path,
    }


__all__ = [
    "DOC_OWNERSHIP_BLOCK_HEADERS",
    "DOC_OWNERSHIP_FINAL_COLUMNS",
    "DOC_OWNERSHIP_INPUT_FIELDS",
    "DOC_OWNERSHIP_RAW_COLUMNS",
    "DOC_OWNERSHIP_REQUEST_COLUMNS",
    "_target_effective_date_for_quarter_end",
    "build_refinitiv_lm2011_doc_ownership_requests",
    "run_refinitiv_lm2011_doc_ownership_exact_handoff_pipeline",
    "run_refinitiv_lm2011_doc_ownership_fallback_handoff_pipeline",
    "run_refinitiv_lm2011_doc_ownership_finalize_pipeline",
]
