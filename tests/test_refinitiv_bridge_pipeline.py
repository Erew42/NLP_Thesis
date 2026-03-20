from __future__ import annotations

import json
import zipfile
from datetime import date
from pathlib import Path

import polars as pl
import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from thesis_pkg.core.ccm.transforms import project_ccm_daily_bridge_surface
from thesis_pkg.io.excel import write_refinitiv_ric_lookup_extended_workbook
from thesis_pkg.pipelines.refinitiv_bridge_pipeline import (
    BRIDGE_OUTPUT_COLUMNS,
    BRIDGE_VENDOR_COLUMNS,
    OWNERSHIP_UNIVERSE_HANDOFF_COLUMNS,
    OWNERSHIP_UNIVERSE_RESULTS_COLUMNS,
    OWNERSHIP_UNIVERSE_ROW_SUMMARY_COLUMNS,
    RIC_LOOKUP_COLUMNS,
    RIC_LOOKUP_EXTENDED_COLUMNS,
    RIC_LOOKUP_RESOLUTION_OUTPUT_COLUMNS,
    RIC_LOOKUP_FILTER_PROFILES,
    _build_filtered_ric_lookup_profile_artifact,
    _build_lookup_profile_bridge_ids,
    _build_ric_lookup_handoff_frames,
    build_refinitiv_ownership_universe_row_summary,
    build_refinitiv_step1_ownership_universe_handoff,
    build_refinitiv_step1_resolution_frame,
    build_refinitiv_lookup_extended_diagnostic_artifact,
    build_refinitiv_step1_bridge_universe,
    run_refinitiv_step1_ownership_universe_handoff_pipeline,
    run_refinitiv_step1_ownership_universe_results_pipeline,
    run_refinitiv_step1_resolution_pipeline,
    run_refinitiv_step1_bridge_pipeline,
)


def _daily_panel() -> pl.DataFrame:
    return pl.DataFrame(
        {
            "KYPERMNO": [1, 1, 1, 1, 2, 4],
            "CALDT": [
                "2024-01-02",
                "2024-01-03",
                "2024-01-04",
                "2024-01-05",
                "2024-01-02",
                "2024-01-02",
            ],
            "KYGVKEY_final": ["1000", "1000", "1000", "1000", "2000", None],
            "LIID": ["A", "A", "A", "A", None, None],
            "CIK_final": [
                "0000001000",
                "0000001000",
                "0000001000",
                "0000001000",
                "0000002000",
                "0000004000",
            ],
            "CUSIP": ["11111111", "11111111", "11111111", "11111111", None, "44444444"],
            "ISIN": [
                "US1111111111",
                "US1111111111",
                "US1111111111",
                "US1111111111",
                "US0000002000",
                None,
            ],
            "TICKER": ["ALFA", "ALFA", "ALFB", "ALFB", "BETA", " DELTA "],
            "LINKTYPE": ["LC", "LC", "LC", "LC", "LU", "LC"],
            "LINKPRIM": ["P", "P", "P", "P", "C", "P"],
            "link_quality_flag": ["canonical_primary_LC"] * 4 + ["canonical_other", "canonical_primary_LC"],
            "HEXCNTRY": ["US", "US", "US", "US", "GB", "US"],
            "n_filings": [0, 2, 0, 1, 0, 0],
            "SHRCD": [10, 10, 12, 12, 11, 10],
            "EXCHCD": [1, 1, 1, 1, 2, 1],
        }
    ).with_columns(pl.col("CALDT").str.strptime(pl.Date, "%Y-%m-%d", strict=True))


def _company_description() -> pl.DataFrame:
    return pl.DataFrame(
        {
            "KYGVKEY": ["1000", "2000"],
            "CONM": ["Alpha Corp", "Beta plc"],
        }
    )


def test_build_refinitiv_step1_bridge_universe_matches_bridge_surface_projection() -> None:
    broad_daily = _daily_panel().with_columns(
        pl.Series("HSIC", [None, None, None, None, None, None], dtype=pl.Int32),
        pl.Series("SRCTYPE_all", [["10-K"], ["10-K"], [], [], [], []], dtype=pl.List(pl.Utf8)),
    )
    bridge_surface = project_ccm_daily_bridge_surface(broad_daily.lazy())

    expected = build_refinitiv_step1_bridge_universe(
        _daily_panel().lazy(),
        company_description_lf=_company_description().lazy(),
    ).collect()
    actual = build_refinitiv_step1_bridge_universe(
        bridge_surface,
        company_description_lf=_company_description().lazy(),
    ).collect()

    assert expected.columns == actual.columns
    assert expected.to_dicts() == actual.to_dicts()


def _bridge_with_manual_review() -> pl.DataFrame:
    bridge = build_refinitiv_step1_bridge_universe(
        _daily_panel()
        .vstack(
            pl.DataFrame(
                {
                    "KYPERMNO": [3, 5],
                    "CALDT": [date(2024, 1, 6), date(2024, 1, 8)],
                    "KYGVKEY_final": ["3000", "5000"],
                    "LIID": [None, None],
                    "CIK_final": ["0000003000", "0000005000"],
                    "CUSIP": [None, None],
                    "ISIN": [None, None],
                    "TICKER": ["GAMMA", "  "],
                    "LINKTYPE": ["LU", "LU"],
                    "LINKPRIM": ["C", "C"],
                    "link_quality_flag": ["ticker_only", "manual_review"],
                    "HEXCNTRY": ["US", "US"],
                    "n_filings": [0, 0],
                    "SHRCD": [10, 10],
                    "EXCHCD": [1, 1],
                }
            )
        )
        .lazy(),
        company_description_lf=_company_description().lazy(),
    ).collect()
    return bridge


def _write_filled_lookup_workbook(path: Path) -> Path:
    workbook = Workbook()
    ric_ws = workbook.active
    ric_ws.title = "ric_lookup"
    ric_ws.append(list(RIC_LOOKUP_COLUMNS))
    records = [
        {
            "bridge_row_id": "1:01:111111111:US1111111111:ALFA-S",
            "KYPERMNO": "1",
            "CUSIP": "111111111",
            "ISIN": "US1111111111",
            "TICKER": "ALFA",
            "first_seen_caldt": "2024-01-01",
            "last_seen_caldt": "2024-01-10",
            "preferred_lookup_id": "US1111111111",
            "preferred_lookup_type": "ISIN",
            "vendor_primary_ric": "AAA.O",
            "vendor_returned_name": "Alpha Inc",
            "vendor_returned_cusip": "111111111",
            "vendor_returned_isin": "US1111111111",
            "vendor_match_status": None,
            "vendor_notes": None,
        },
        {
            "bridge_row_id": "1:01:111111111:US1111111111:ALFA-AFTER",
            "KYPERMNO": "1",
            "CUSIP": "111111111",
            "ISIN": "US1111111111",
            "TICKER": "ALFA",
            "first_seen_caldt": "2023-12-20",
            "last_seen_caldt": "2023-12-25",
            "preferred_lookup_id": "US1111111111",
            "preferred_lookup_type": "ISIN",
            "vendor_primary_ric": "NULL",
            "vendor_returned_name": "The invalid identifier ignored.",
            "vendor_returned_cusip": "The invalid identifier ignored.",
            "vendor_returned_isin": "The invalid identifier ignored.",
            "vendor_match_status": None,
            "vendor_notes": None,
        },
        {
            "bridge_row_id": "1:01:111111111:US1111111111:ALFA-BEFORE",
            "KYPERMNO": "1",
            "CUSIP": "111111111",
            "ISIN": "US1111111111",
            "TICKER": "ALFA",
            "first_seen_caldt": "2024-01-11",
            "last_seen_caldt": "2024-01-15",
            "preferred_lookup_id": "US1111111111",
            "preferred_lookup_type": "ISIN",
            "vendor_primary_ric": "",
            "vendor_returned_name": None,
            "vendor_returned_cusip": None,
            "vendor_returned_isin": None,
            "vendor_match_status": None,
            "vendor_notes": None,
        },
        {
            "bridge_row_id": "2:01:222222222:US2222222222:BETA-S1",
            "KYPERMNO": "2",
            "CUSIP": "222222222",
            "ISIN": "US2222222222",
            "TICKER": "BETA",
            "first_seen_caldt": "2024-02-01",
            "last_seen_caldt": "2024-02-10",
            "preferred_lookup_id": "US2222222222",
            "preferred_lookup_type": "ISIN",
            "vendor_primary_ric": "BBB.O",
            "vendor_returned_name": "Beta Inc",
            "vendor_returned_cusip": "222222222",
            "vendor_returned_isin": "US2222222222",
            "vendor_match_status": None,
            "vendor_notes": None,
        },
        {
            "bridge_row_id": "2:01:223333333:US2233333333:BETN-S2",
            "KYPERMNO": "2",
            "CUSIP": "223333333",
            "ISIN": "US2233333333",
            "TICKER": "BETN",
            "first_seen_caldt": "2024-03-01",
            "last_seen_caldt": "2024-03-10",
            "preferred_lookup_id": "US2233333333",
            "preferred_lookup_type": "ISIN",
            "vendor_primary_ric": "CCC.O",
            "vendor_returned_name": "Beta New",
            "vendor_returned_cusip": "223333333",
            "vendor_returned_isin": "US2233333333",
            "vendor_match_status": None,
            "vendor_notes": None,
        },
        {
            "bridge_row_id": "2:01:222222222:US2222222222:BETA-F1",
            "KYPERMNO": "2",
            "CUSIP": "222222222",
            "ISIN": "US2222222222",
            "TICKER": "BETA",
            "first_seen_caldt": "2024-02-15",
            "last_seen_caldt": "2024-02-20",
            "preferred_lookup_id": "222222222",
            "preferred_lookup_type": "CUSIP",
            "vendor_primary_ric": "",
            "vendor_returned_name": None,
            "vendor_returned_cusip": None,
            "vendor_returned_isin": None,
            "vendor_match_status": None,
            "vendor_notes": None,
        },
        {
            "bridge_row_id": "2:01:222222222:US2222222222:BETA-F2",
            "KYPERMNO": "2",
            "CUSIP": "222222222",
            "ISIN": "US2222222222",
            "TICKER": "BETA",
            "first_seen_caldt": "2024-02-21",
            "last_seen_caldt": "2024-02-25",
            "preferred_lookup_id": "US2222222222",
            "preferred_lookup_type": "ISIN",
            "vendor_primary_ric": None,
            "vendor_returned_name": None,
            "vendor_returned_cusip": None,
            "vendor_returned_isin": None,
            "vendor_match_status": None,
            "vendor_notes": None,
        },
        {
            "bridge_row_id": "3:01:333333333:US3333333333:GAM-F",
            "KYPERMNO": "3",
            "CUSIP": "333333333",
            "ISIN": "US3333333333",
            "TICKER": "GAM",
            "first_seen_caldt": "2024-04-01",
            "last_seen_caldt": "2024-04-05",
            "preferred_lookup_id": "US3333333333",
            "preferred_lookup_type": "ISIN",
            "vendor_primary_ric": None,
            "vendor_returned_name": None,
            "vendor_returned_cusip": None,
            "vendor_returned_isin": None,
            "vendor_match_status": None,
            "vendor_notes": None,
        },
        {
            "bridge_row_id": "4:01:444444444:US4444444444:DLTA-S",
            "KYPERMNO": "4",
            "CUSIP": "444444444",
            "ISIN": "US4444444444",
            "TICKER": "DLTA",
            "first_seen_caldt": "2024-05-01",
            "last_seen_caldt": "2024-05-10",
            "preferred_lookup_id": "US4444444444",
            "preferred_lookup_type": "ISIN",
            "vendor_primary_ric": "DDD.O",
            "vendor_returned_name": "Delta Inc",
            "vendor_returned_cusip": "444444444",
            "vendor_returned_isin": "US4444444444",
            "vendor_match_status": None,
            "vendor_notes": None,
        },
        {
            "bridge_row_id": "4:01:444444444:US4444444444:DLTA-AFTER",
            "KYPERMNO": "4",
            "CUSIP": "444444444",
            "ISIN": "US4444444444",
            "TICKER": "DLTA",
            "first_seen_caldt": "2024-04-20",
            "last_seen_caldt": "2024-04-25",
            "preferred_lookup_id": "US4444444444",
            "preferred_lookup_type": "ISIN",
            "vendor_primary_ric": None,
            "vendor_returned_name": None,
            "vendor_returned_cusip": None,
            "vendor_returned_isin": None,
            "vendor_match_status": None,
            "vendor_notes": None,
        },
        {
            "bridge_row_id": "4:01:444444444:US4444444444:DLTA-BEFORE",
            "KYPERMNO": "4",
            "CUSIP": "444444444",
            "ISIN": "US4444444444",
            "TICKER": "DLTA",
            "first_seen_caldt": "2024-05-11",
            "last_seen_caldt": "2024-05-12",
            "preferred_lookup_id": "444444444",
            "preferred_lookup_type": "CUSIP",
            "vendor_primary_ric": "NULL",
            "vendor_returned_name": "The invalid identifier ignored.",
            "vendor_returned_cusip": "The invalid identifier ignored.",
            "vendor_returned_isin": "The invalid identifier ignored.",
            "vendor_match_status": None,
            "vendor_notes": None,
        },
        {
            "bridge_row_id": "5:01:-:-:EPSI-F",
            "KYPERMNO": "5",
            "CUSIP": None,
            "ISIN": None,
            "TICKER": "EPSI",
            "first_seen_caldt": "2024-06-01",
            "last_seen_caldt": "2024-06-05",
            "preferred_lookup_id": "EPSI",
            "preferred_lookup_type": "TICKER",
            "vendor_primary_ric": None,
            "vendor_returned_name": None,
            "vendor_returned_cusip": None,
            "vendor_returned_isin": None,
            "vendor_match_status": None,
            "vendor_notes": None,
        },
    ]
    for record in records:
        ric_ws.append([record.get(name) for name in RIC_LOOKUP_COLUMNS])
    workbook.create_sheet("README")
    workbook.save(path)
    workbook.close()
    return path


def _bridge_row_id(
    kypermno: str,
    *,
    liid: str | None = "01",
    cusip: str | None = None,
    isin: str | None = None,
    ticker: str | None = None,
) -> str:
    return ":".join(
        (
            kypermno,
            liid or "-",
            cusip or "-",
            isin or "-",
            ticker or "-",
        )
    )


def _extended_resolution_row(
    *,
    kypermno: str,
    first_seen_caldt: date,
    last_seen_caldt: date,
    liid: str | None = "01",
    cusip: str | None = None,
    isin: str | None = None,
    ticker: str | None = None,
    **overrides: object,
) -> dict[str, object]:
    row: dict[str, object] = {name: None for name in RIC_LOOKUP_EXTENDED_COLUMNS}
    row.update(
        {
            "bridge_row_id": _bridge_row_id(
                kypermno,
                liid=liid,
                cusip=cusip,
                isin=isin,
                ticker=ticker,
            ),
            "KYPERMNO": kypermno,
            "CUSIP": cusip,
            "ISIN": isin,
            "TICKER": ticker,
            "first_seen_caldt": first_seen_caldt,
            "last_seen_caldt": last_seen_caldt,
            "ISIN_lookup_input": isin,
            "CUSIP_lookup_input": cusip,
            "TICKER_lookup_input": ticker,
        }
    )
    row.update(overrides)
    return row


def _resolution_input_df(rows: list[dict[str, object]]) -> pl.DataFrame:
    return pl.DataFrame(rows).select(
        [
            pl.col(name).alias(name)
            if name not in {"first_seen_caldt", "last_seen_caldt"}
            else pl.col(name).cast(pl.Date, strict=False).alias(name)
            for name in RIC_LOOKUP_EXTENDED_COLUMNS
        ]
    )


def _write_filled_extended_lookup_workbook(path: Path, rows: list[dict[str, object]]) -> Path:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "lookup_diagnostics"
    ws.append(list(RIC_LOOKUP_EXTENDED_COLUMNS))
    for row in rows:
        ws.append([row.get(name) for name in RIC_LOOKUP_EXTENDED_COLUMNS])
    summary_ws = workbook.create_sheet("summary")
    summary_ws.append(["summary_category", "summary_key", "value"])
    workbook.create_sheet("README")
    workbook.save(path)
    workbook.close()
    return path


def _write_filled_ownership_universe_workbook(
    path: Path,
    handoff_df: pl.DataFrame,
    returned_rows_by_lookup_row_id: dict[str, list[tuple[str, object, object, str]]],
) -> Path:
    workbook = Workbook()
    default_ws = workbook.active
    default_ws.title = "ownership_retrieval"
    block_headers = (
        "input_data",
        "returned_ric",
        "returned_date",
        "returned_value",
        "returned_category",
    )

    eligible_df = handoff_df.filter(pl.col("retrieval_eligible").fill_null(False))
    input_fields = (
        "bridge_row_id",
        "candidate_ric",
        "request_start_date",
        "request_end_date",
        "candidate_slot",
        "lookup_input_source",
        "effective_collection_ric",
        "effective_collection_ric_source",
        "accepted_ric",
        "accepted_ric_source",
    )

    for block_index, row in enumerate(eligible_df.iter_rows(named=True)):
        base_col = 1 + (block_index * 5)
        for offset, header in enumerate(block_headers):
            default_ws.cell(row=1, column=base_col + offset).value = header
        for field_offset, field_name in enumerate(input_fields, start=1):
            default_ws.cell(row=1 + field_offset, column=base_col).value = row.get(field_name)
        rows = returned_rows_by_lookup_row_id.get(str(row["ownership_lookup_row_id"]), [])
        for result_idx, (returned_ric, returned_date, returned_value, returned_category) in enumerate(rows, start=1):
            excel_row = 2 + result_idx
            default_ws.cell(row=excel_row, column=base_col + 1).value = returned_ric
            default_ws.cell(row=excel_row, column=base_col + 2).value = returned_date
            default_ws.cell(row=excel_row, column=base_col + 3).value = returned_value
            default_ws.cell(row=excel_row, column=base_col + 4).value = returned_category

    workbook.save(path)
    workbook.close()
    return path


def test_build_refinitiv_step1_bridge_universe_uses_distinct_identifier_grain() -> None:
    bridge = build_refinitiv_step1_bridge_universe(
        _daily_panel().lazy(),
        company_description_lf=_company_description().lazy(),
    ).collect()

    assert bridge.columns == list(BRIDGE_OUTPUT_COLUMNS)
    assert bridge.height == 4
    assert bridge["bridge_row_id"].n_unique() == bridge.height

    by_key = {row["bridge_row_id"]: row for row in bridge.to_dicts()}
    first = by_key["1:A:11111111:US1111111111:ALFA"]
    second = by_key["1:A:11111111:US1111111111:ALFB"]
    third = by_key["2:-:-:US0000002000:BETA"]
    fourth = by_key["4:-:44444444:-:DELTA"]

    assert first["first_seen_caldt"].isoformat() == "2024-01-02"
    assert first["last_seen_caldt"].isoformat() == "2024-01-03"
    assert first["n_daily_rows"] == 2
    assert first["n_filing_days"] == 1
    assert first["n_filings_total"] == 2
    assert first["company_name"] == "Alpha Corp"
    assert first["TICKER"] == "ALFA"

    assert second["first_seen_caldt"].isoformat() == "2024-01-04"
    assert second["last_seen_caldt"].isoformat() == "2024-01-05"
    assert second["n_daily_rows"] == 2
    assert second["n_filing_days"] == 1
    assert second["n_filings_total"] == 1
    assert second["TICKER"] == "ALFB"

    assert third["company_name"] == "Beta plc"
    assert third["LIID"] is None
    assert third["CUSIP"] is None
    assert third["ISIN"] == "US0000002000"
    assert third["TICKER"] == "BETA"
    assert fourth["KYGVKEY_final"] is None
    assert fourth["CUSIP"] == "44444444"
    assert fourth["TICKER"] == "DELTA"

    for name in BRIDGE_VENDOR_COLUMNS:
        assert bridge[name].null_count() == bridge.height


def test_build_ric_lookup_handoff_frames_prefers_isin_and_separates_manual_review() -> None:
    lookup_df, manual_review_df = _build_ric_lookup_handoff_frames(_bridge_with_manual_review())

    assert lookup_df.columns == list(RIC_LOOKUP_COLUMNS)
    assert manual_review_df.columns == list(RIC_LOOKUP_COLUMNS)
    assert lookup_df.height == 5
    assert manual_review_df.height == 1

    lookup_rows = {row["bridge_row_id"]: row for row in lookup_df.to_dicts()}
    assert lookup_rows["1:A:11111111:US1111111111:ALFA"]["preferred_lookup_id"] == "US1111111111"
    assert lookup_rows["1:A:11111111:US1111111111:ALFA"]["preferred_lookup_type"] == "ISIN"
    assert lookup_rows["1:A:11111111:US1111111111:ALFA"]["vendor_returned_name"] is None
    assert lookup_rows["1:A:11111111:US1111111111:ALFA"]["vendor_returned_cusip"] is None
    assert lookup_rows["1:A:11111111:US1111111111:ALFA"]["vendor_returned_isin"] is None
    assert lookup_rows["2:-:-:US0000002000:BETA"]["preferred_lookup_id"] == "US0000002000"
    assert lookup_rows["2:-:-:US0000002000:BETA"]["preferred_lookup_type"] == "ISIN"
    assert lookup_rows["3:-:-:-:GAMMA"]["preferred_lookup_id"] == "GAMMA"
    assert lookup_rows["3:-:-:-:GAMMA"]["preferred_lookup_type"] == "TICKER"

    manual_row = manual_review_df.to_dicts()[0]
    assert manual_row["bridge_row_id"] == "5:-:-:-:-"
    assert manual_row["preferred_lookup_id"] is None
    assert manual_row["preferred_lookup_type"] is None


def test_build_refinitiv_lookup_extended_diagnostic_artifact_tracks_all_identifier_inputs() -> None:
    lookup_df, manual_review_df = _build_ric_lookup_handoff_frames(_bridge_with_manual_review())

    extended_df, summary_df, summary_payload = build_refinitiv_lookup_extended_diagnostic_artifact(
        lookup_df,
        manual_review_df,
    )

    assert extended_df.columns == list(RIC_LOOKUP_EXTENDED_COLUMNS)
    assert extended_df.height == 6

    rows = {row["bridge_row_id"]: row for row in extended_df.to_dicts()}
    alfa_row = rows["1:A:11111111:US1111111111:ALFA"]
    assert "preferred_lookup_id" not in extended_df.columns
    assert "preferred_lookup_type" not in extended_df.columns
    assert alfa_row["ISIN_lookup_input"] == "US1111111111"
    assert alfa_row["CUSIP_lookup_input"] == "11111111"
    assert alfa_row["TICKER_lookup_input"] == "ALFA"
    assert alfa_row["ISIN_attempted"] is True
    assert alfa_row["CUSIP_attempted"] is True
    assert alfa_row["TICKER_attempted"] is True
    assert alfa_row["ISIN_success"] is False
    assert alfa_row["all_successful_attempts_consistent"] is None
    assert alfa_row["vendor_primary_ric"] is None
    assert alfa_row["vendor_returned_name"] is None

    gamma_row = rows["3:-:-:-:GAMMA"]
    assert gamma_row["ISIN_lookup_input"] is None
    assert gamma_row["CUSIP_lookup_input"] is None
    assert gamma_row["TICKER_lookup_input"] == "GAMMA"
    assert gamma_row["TICKER_attempted"] is True

    manual_row = rows["5:-:-:-:-"]
    assert manual_row["ISIN_lookup_input"] is None
    assert manual_row["CUSIP_lookup_input"] is None
    assert manual_row["TICKER_lookup_input"] is None
    assert manual_row["ISIN_attempted"] is False
    assert manual_row["CUSIP_attempted"] is False
    assert manual_row["TICKER_attempted"] is False

    assert summary_df.filter(pl.col("summary_category") == "attempt_count_by_identifier_type").sort(
        "summary_key"
    ).to_dicts() == [
        {"summary_category": "attempt_count_by_identifier_type", "summary_key": "CUSIP", "value": 3},
        {"summary_category": "attempt_count_by_identifier_type", "summary_key": "ISIN", "value": 3},
        {"summary_category": "attempt_count_by_identifier_type", "summary_key": "TICKER", "value": 5},
    ]
    assert summary_payload["attempt_counts_by_identifier_type"] == {
        "ISIN": 3,
        "CUSIP": 3,
        "TICKER": 5,
    }
    assert summary_payload["success_counts_by_identifier_type"] == {
        "ISIN": 0,
        "CUSIP": 0,
        "TICKER": 0,
    }
    assert summary_payload["rows_where_only_one_identifier_type_succeeds"] == 0


def test_write_refinitiv_ric_lookup_extended_workbook_prefills_formulas(tmp_path: Path) -> None:
    lookup_df, manual_review_df = _build_ric_lookup_handoff_frames(_bridge_with_manual_review())
    extended_df, summary_df, summary_payload = build_refinitiv_lookup_extended_diagnostic_artifact(
        lookup_df,
        manual_review_df,
    )
    out_path = tmp_path / "refinitiv_ric_lookup_handoff_common_stock_extended.xlsx"

    write_refinitiv_ric_lookup_extended_workbook(
        extended_df,
        out_path,
        readme_payload={"extended_lookup_diagnostic_summary": summary_payload},
        text_columns=tuple(
            name
            for name in RIC_LOOKUP_EXTENDED_COLUMNS
            if name not in {"all_successful_attempts_consistent"}
            and not name.endswith("_attempted")
            and not name.endswith("_success")
            and "_same_" not in name
        ),
        summary_df=summary_df,
        summary_text_columns=("summary_category", "summary_key"),
    )

    workbook = load_workbook(out_path, data_only=False)
    try:
        assert "lookup_diagnostics" in workbook.sheetnames
        assert "summary" in workbook.sheetnames
        assert "README" in workbook.sheetnames

        ws = workbook["lookup_diagnostics"]
        headers = [cell.value for cell in ws[1]]
        assert "preferred_lookup_id" not in headers
        assert "preferred_lookup_type" not in headers

        col_idx = {name: idx + 1 for idx, name in enumerate(headers)}
        col_ref = {name: f"{get_column_letter(idx)}2" for name, idx in col_idx.items()}
        assert ws.cell(2, col_idx["ISIN_lookup_input"]).value == "US1111111111"
        assert ws.cell(2, col_idx["ISIN_returned_ric"]).value == (
            f'=IF({col_ref["ISIN_lookup_input"]}="","",_xll.TR({col_ref["ISIN_lookup_input"]},"TR.RIC"))'
        )
        assert ws.cell(2, col_idx["ISIN_returned_name"]).value == (
            f'=IF({col_ref["ISIN_lookup_input"]}="","",_xll.TR({col_ref["ISIN_lookup_input"]},"TR.CommonName"))'
        )
        assert ws.cell(2, col_idx["CUSIP_returned_ric"]).value == (
            f'=IF({col_ref["CUSIP_lookup_input"]}="","",_xll.TR({col_ref["CUSIP_lookup_input"]},"TR.RIC"))'
        )
        assert ws.cell(2, col_idx["TICKER_returned_ric"]).value == (
            f'=IF({col_ref["TICKER_lookup_input"]}="","",_xll.TR({col_ref["TICKER_lookup_input"]},"TR.RIC"))'
        )
        assert ws.cell(2, col_idx["ISIN_attempted"]).value == f'=LEN({col_ref["ISIN_lookup_input"]})>0'
        assert ws.cell(2, col_idx["ISIN_success"]).value == (
            f'=IFERROR({col_ref["ISIN_returned_ric"]}<>"",FALSE)'
        )
        assert ws.cell(2, col_idx["ISIN_vs_CUSIP_same_ric"]).value == (
            f'=IF(AND(IFERROR({col_ref["ISIN_returned_ric"]}<>"",FALSE),'
            f'IFERROR({col_ref["CUSIP_returned_ric"]}<>"",FALSE)),'
            f'IFERROR({col_ref["ISIN_returned_ric"]}={col_ref["CUSIP_returned_ric"]},FALSE),"")'
        )
        pairwise_refs = [
            col_ref["ISIN_vs_CUSIP_same_ric"],
            col_ref["ISIN_vs_CUSIP_same_isin"],
            col_ref["ISIN_vs_CUSIP_same_cusip"],
            col_ref["ISIN_vs_TICKER_same_ric"],
            col_ref["ISIN_vs_TICKER_same_isin"],
            col_ref["ISIN_vs_TICKER_same_cusip"],
            col_ref["CUSIP_vs_TICKER_same_ric"],
            col_ref["CUSIP_vs_TICKER_same_isin"],
            col_ref["CUSIP_vs_TICKER_same_cusip"],
        ]
        pairwise_terms = ",".join(f'IF({cell_ref}="",TRUE,{cell_ref})' for cell_ref in pairwise_refs)
        expected_consistency_formula = (
            f'=IF((N({col_ref["ISIN_success"]})+N({col_ref["CUSIP_success"]})+N({col_ref["TICKER_success"]}))<2,"",'
            f"AND({pairwise_terms}))"
        )
        assert ws.cell(2, col_idx["all_successful_attempts_consistent"]).value == (
            expected_consistency_formula
        )

        readme_ws = workbook["README"]
        readme_values = [
            cell
            for row in readme_ws.iter_rows(values_only=True)
            for cell in row
            if isinstance(cell, str)
        ]
        readme_text = "\n".join(readme_values)
        assert "main manual comparison workbook" in readme_text
        assert "save the evaluated workbook as a new file before any re-import step" in readme_text

        summary_ws = workbook["summary"]
        summary_headers = [cell.value for cell in summary_ws[1]]
        summary_col_idx = {name: idx + 1 for idx, name in enumerate(summary_headers)}
        summary_value = summary_ws.cell(2, summary_col_idx["value"]).value
        assert isinstance(summary_value, str)
        assert summary_value.startswith("=COUNTIF('lookup_diagnostics'!")
    finally:
        workbook.close()


def test_filtered_ric_lookup_profiles_apply_common_stock_and_gvkey_predicates() -> None:
    bridge = build_refinitiv_step1_bridge_universe(
        _daily_panel().lazy(),
        company_description_lf=_company_description().lazy(),
    ).collect()
    profiles = {profile.name: profile for profile in RIC_LOOKUP_FILTER_PROFILES}

    common_lookup_df, _, common_summary = _build_filtered_ric_lookup_profile_artifact(
        bridge,
        _build_lookup_profile_bridge_ids(_daily_panel().lazy(), profiles["common_stock"]),
        profiles["common_stock"],
    )
    strict_lookup_df, _, strict_summary = _build_filtered_ric_lookup_profile_artifact(
        bridge,
        _build_lookup_profile_bridge_ids(_daily_panel().lazy(), profiles["common_stock_with_gvkey"]),
        profiles["common_stock_with_gvkey"],
    )

    assert set(common_lookup_df["bridge_row_id"].to_list()) == {
        "1:A:11111111:US1111111111:ALFA",
        "2:-:-:US0000002000:BETA",
        "4:-:44444444:-:DELTA",
    }
    assert set(strict_lookup_df["bridge_row_id"].to_list()) == {
        "1:A:11111111:US1111111111:ALFA",
        "2:-:-:US0000002000:BETA",
    }
    assert common_summary["predicates_applied"] == [
        "KYPERMNO is not null",
        "common stock (SHRCD in {10, 11})",
    ]
    assert common_summary["rows_before_filter"] == 4
    assert common_summary["rows_after_filter"] == 3
    assert strict_summary["rows_after_filter"] == 2


def test_build_refinitiv_step1_resolution_frame_applies_conventional_and_extension_policy() -> None:
    rows = [
        _extended_resolution_row(
            kypermno="3001",
            first_seen_caldt=date(2024, 1, 1),
            last_seen_caldt=date(2024, 1, 5),
            cusip="111111111",
            isin="US1111111111",
            ticker="ALFA",
            ISIN_returned_ric="AAA.N",
            ISIN_returned_isin="US1111111111",
            ISIN_returned_cusip="111111111",
        ),
        _extended_resolution_row(
            kypermno="3002",
            first_seen_caldt=date(2024, 1, 1),
            last_seen_caldt=date(2024, 1, 5),
            cusip="222222222",
            isin="US2222222222",
            ticker="BETA",
            CUSIP_returned_ric="BBB.N",
            CUSIP_returned_isin="US2222222222",
            CUSIP_returned_cusip="222222222",
        ),
        _extended_resolution_row(
            kypermno="3003",
            first_seen_caldt=date(2024, 1, 1),
            last_seen_caldt=date(2024, 1, 5),
            cusip="333333333",
            isin="US3333333333",
            ticker="GAMM",
            ISIN_returned_ric="CCC.N",
            ISIN_returned_isin="US3333333333",
            ISIN_returned_cusip="333333333",
            CUSIP_returned_ric="CCC.O",
            CUSIP_returned_isin="US3333333333",
            CUSIP_returned_cusip="333333333",
        ),
        _extended_resolution_row(
            kypermno="3004",
            first_seen_caldt=date(2024, 1, 1),
            last_seen_caldt=date(2024, 1, 5),
            cusip="444444444",
            isin="US4444444444",
            ticker="DELT",
            ISIN_returned_ric="DDD.N",
            ISIN_returned_isin="US4444444444",
            ISIN_returned_cusip="444444444",
            CUSIP_returned_ric="EEE.N",
            CUSIP_returned_isin="US5555555555",
            CUSIP_returned_cusip="555555555",
        ),
        _extended_resolution_row(
            kypermno="3005",
            first_seen_caldt=date(2024, 1, 1),
            last_seen_caldt=date(2024, 1, 5),
            cusip="666666666",
            isin="US6666666666",
            ticker="EPSI",
            ISIN_returned_ric="FFF.N",
            ISIN_returned_isin="US6666666666",
            CUSIP_returned_ric="FFF.N",
            CUSIP_returned_cusip="666666666",
        ),
        _extended_resolution_row(
            kypermno="3010",
            first_seen_caldt=date(2024, 1, 1),
            last_seen_caldt=date(2024, 1, 5),
            cusip="101010101",
            isin="US1010101010",
            ticker="PRIA",
            ISIN_returned_ric="GGG.N",
            ISIN_returned_isin="US1010101010",
            ISIN_returned_cusip="101010101",
        ),
        _extended_resolution_row(
            kypermno="3010",
            first_seen_caldt=date(2024, 1, 6),
            last_seen_caldt=date(2024, 1, 10),
            cusip="101010101",
            isin="US1010101010",
            ticker="PRIB",
        ),
        _extended_resolution_row(
            kypermno="3011",
            first_seen_caldt=date(2024, 1, 1),
            last_seen_caldt=date(2024, 1, 5),
            cusip="202020202",
            isin="US2020202020",
            ticker="NXTA",
        ),
        _extended_resolution_row(
            kypermno="3011",
            first_seen_caldt=date(2024, 1, 6),
            last_seen_caldt=date(2024, 1, 10),
            cusip="202020202",
            isin="US2020202020",
            ticker="NXTB",
            CUSIP_returned_ric="HHH.N",
            CUSIP_returned_isin="US2020202020",
            CUSIP_returned_cusip="202020202",
        ),
        _extended_resolution_row(
            kypermno="3012",
            first_seen_caldt=date(2024, 1, 1),
            last_seen_caldt=date(2024, 1, 5),
            cusip="303030303",
            isin="US3030303030",
            ticker="ADJP",
            CUSIP_returned_ric="III.C",
            CUSIP_returned_isin="US3030303030",
            CUSIP_returned_cusip="303030303",
        ),
        _extended_resolution_row(
            kypermno="3012",
            first_seen_caldt=date(2024, 1, 6),
            last_seen_caldt=date(2024, 1, 10),
            cusip=None,
            isin=None,
            ticker="ADJM",
        ),
        _extended_resolution_row(
            kypermno="3012",
            first_seen_caldt=date(2024, 1, 11),
            last_seen_caldt=date(2024, 1, 15),
            cusip="303030303",
            isin="US3030303030",
            ticker="ADJN",
            ISIN_returned_ric="III.I",
            ISIN_returned_isin="US3030303030",
            ISIN_returned_cusip="303030303",
        ),
        _extended_resolution_row(
            kypermno="3013",
            first_seen_caldt=date(2024, 1, 1),
            last_seen_caldt=date(2024, 1, 5),
            cusip="404040404",
            isin="US4040404040",
            ticker="CFLP",
            ISIN_returned_ric="JJJ.N",
            ISIN_returned_isin="US4040404040",
            ISIN_returned_cusip="404040404",
        ),
        _extended_resolution_row(
            kypermno="3013",
            first_seen_caldt=date(2024, 1, 6),
            last_seen_caldt=date(2024, 1, 10),
            cusip=None,
            isin=None,
            ticker="CFLM",
        ),
        _extended_resolution_row(
            kypermno="3013",
            first_seen_caldt=date(2024, 1, 11),
            last_seen_caldt=date(2024, 1, 15),
            cusip="505050505",
            isin="US5050505050",
            ticker="CFLN",
            ISIN_returned_ric="KKK.N",
            ISIN_returned_isin="US5050505050",
            ISIN_returned_cusip="505050505",
        ),
        _extended_resolution_row(
            kypermno="3014",
            first_seen_caldt=date(2024, 1, 1),
            last_seen_caldt=date(2024, 1, 5),
            cusip="606060606",
            isin="US6060606060",
            ticker="BLKP",
            ISIN_returned_ric="LLL.N",
            ISIN_returned_isin="US6060606060",
            ISIN_returned_cusip="606060606",
        ),
        _extended_resolution_row(
            kypermno="3014",
            first_seen_caldt=date(2024, 1, 6),
            last_seen_caldt=date(2024, 1, 10),
            cusip="606060606",
            isin="US6060606060",
            ticker="BLKM",
            ISIN_returned_ric="LLL.N",
            ISIN_returned_isin="US6060606060",
            ISIN_returned_cusip="606060606",
            CUSIP_returned_ric="MMM.N",
            CUSIP_returned_isin="US9090909090",
            CUSIP_returned_cusip="909090909",
        ),
        _extended_resolution_row(
            kypermno="3014",
            first_seen_caldt=date(2024, 1, 11),
            last_seen_caldt=date(2024, 1, 15),
            cusip="606060606",
            isin="US6060606060",
            ticker="BLKN",
            ISIN_returned_ric="LLL.N",
            ISIN_returned_isin="US6060606060",
            ISIN_returned_cusip="606060606",
        ),
        _extended_resolution_row(
            kypermno="3015",
            first_seen_caldt=date(2024, 1, 1),
            last_seen_caldt=date(2024, 1, 5),
            cusip=None,
            isin=None,
            ticker="TKRO",
            TICKER_returned_ric="TICK.ONLY",
            TICKER_returned_isin="USTICKONLY01",
            TICKER_returned_cusip="TICKONLY1",
        ),
        _extended_resolution_row(
            kypermno="3016",
            first_seen_caldt=date(2024, 1, 1),
            last_seen_caldt=date(2024, 1, 5),
            cusip="707070707",
            isin="US7070707070",
            ticker="TKOK",
            ISIN_returned_ric="NNN.N",
            ISIN_returned_isin="US7070707070",
            ISIN_returned_cusip="707070707",
            TICKER_returned_ric="NNN",
            TICKER_returned_isin="US7070707070",
            TICKER_returned_cusip="707070707",
        ),
        _extended_resolution_row(
            kypermno="3017",
            first_seen_caldt=date(2024, 1, 1),
            last_seen_caldt=date(2024, 1, 5),
            cusip="808080808",
            isin="US8080808080",
            ticker="TKCF",
            ISIN_returned_ric="OOO.N",
            ISIN_returned_isin="US8080808080",
            ISIN_returned_cusip="808080808",
            TICKER_returned_ric="PPP.N",
            TICKER_returned_isin="US9999999999",
            TICKER_returned_cusip="999999999",
        ),
    ]

    resolution_df = build_refinitiv_step1_resolution_frame(_resolution_input_df(rows))

    assert resolution_df.columns == list(RIC_LOOKUP_RESOLUTION_OUTPUT_COLUMNS)

    by_id = {row["bridge_row_id"]: row for row in resolution_df.to_dicts()}

    assert by_id[_bridge_row_id("3001", cusip="111111111", isin="US1111111111", ticker="ALFA")][
        "accepted_resolution_status"
    ] == "resolved_from_isin"
    assert by_id[_bridge_row_id("3001", cusip="111111111", isin="US1111111111", ticker="ALFA")][
        "accepted_ric"
    ] == "AAA.N"

    assert by_id[_bridge_row_id("3002", cusip="222222222", isin="US2222222222", ticker="BETA")][
        "accepted_resolution_status"
    ] == "resolved_from_cusip"

    agree_row = by_id[_bridge_row_id("3003", cusip="333333333", isin="US3333333333", ticker="GAMM")]
    assert agree_row["accepted_resolution_status"] == "resolved_conventional_agree"
    assert agree_row["accepted_ric"] == "CCC.N"
    assert agree_row["accepted_ric_source"] == "ISIN"

    conflict_row = by_id[_bridge_row_id("3004", cusip="444444444", isin="US4444444444", ticker="DELT")]
    assert conflict_row["accepted_resolution_status"] == "unresolved_conventional_conflict"
    assert conflict_row["conventional_identity_conflict"] is True
    assert conflict_row["accepted_ric"] is None

    no_overlap_row = by_id[_bridge_row_id("3005", cusip="666666666", isin="US6666666666", ticker="EPSI")]
    assert no_overlap_row["accepted_resolution_status"] == "unresolved_conventional_conflict"
    assert no_overlap_row["accepted_ric"] is None

    extend_prior_row = by_id[_bridge_row_id("3010", cusip="101010101", isin="US1010101010", ticker="PRIB")]
    assert extend_prior_row["extension_status"] == "extended_from_prior_conventional_span"
    assert extend_prior_row["extended_ric"] == "GGG.N"
    assert extend_prior_row["effective_collection_ric_source"] == "EXTENDED_FROM_PRIOR_ISIN"

    extend_next_row = by_id[_bridge_row_id("3011", cusip="202020202", isin="US2020202020", ticker="NXTA")]
    assert extend_next_row["extension_status"] == "extended_from_next_conventional_span"
    assert extend_next_row["extended_ric"] == "HHH.N"
    assert extend_next_row["effective_collection_ric_source"] == "EXTENDED_FROM_NEXT_CUSIP"

    extend_adjacent_row = by_id[_bridge_row_id("3012", cusip=None, isin=None, ticker="ADJM")]
    assert extend_adjacent_row["extension_status"] == "extended_from_adjacent_conventional_span"
    assert extend_adjacent_row["extension_direction"] == "ADJACENT"
    assert extend_adjacent_row["extended_ric"] == "III.I"
    assert extend_adjacent_row["extended_from_bridge_row_id"] == _bridge_row_id(
        "3012",
        cusip="303030303",
        isin="US3030303030",
        ticker="ADJN",
    )
    assert extend_adjacent_row["effective_collection_ric_source"] == "EXTENDED_FROM_ADJACENT_ISIN"

    blocked_adjacent_row = by_id[_bridge_row_id("3013", cusip=None, isin=None, ticker="CFLM")]
    assert blocked_adjacent_row["extension_status"] == "not_extended_due_to_conflict"
    assert blocked_adjacent_row["effective_collection_ric"] is None

    blocked_conflict_row = by_id[_bridge_row_id("3014", cusip="606060606", isin="US6060606060", ticker="BLKM")]
    assert blocked_conflict_row["conventional_identity_conflict"] is True
    assert blocked_conflict_row["extension_status"] == "not_extended_due_to_conflict"
    assert blocked_conflict_row["effective_collection_ric"] is None

    ticker_only_row = by_id[_bridge_row_id("3015", cusip=None, isin=None, ticker="TKRO")]
    assert ticker_only_row["ticker_candidate_available"] is True
    assert ticker_only_row["accepted_ric"] is None
    assert ticker_only_row["extended_ric"] is None
    assert ticker_only_row["effective_collection_ric"] is None
    assert ticker_only_row["extension_status"] == "not_extended_no_adjacent_conventional_source"

    ticker_agree_row = by_id[_bridge_row_id("3016", cusip="707070707", isin="US7070707070", ticker="TKOK")]
    assert ticker_agree_row["ticker_candidate_conflicts_with_conventional"] is False

    ticker_conflict_row = by_id[_bridge_row_id("3017", cusip="808080808", isin="US8080808080", ticker="TKCF")]
    assert ticker_conflict_row["ticker_candidate_conflicts_with_conventional"] is True


def test_run_refinitiv_step1_resolution_pipeline_writes_resolved_artifacts(tmp_path: Path) -> None:
    rows = [
        _extended_resolution_row(
            kypermno="4001",
            first_seen_caldt=date(2024, 2, 1),
            last_seen_caldt=date(2024, 2, 5),
            cusip="111111119",
            isin="US1111111199",
            ticker="RSLV",
            ISIN_returned_ric="RESV.N",
            ISIN_returned_isin="US1111111199",
            ISIN_returned_cusip="111111119",
        ),
        _extended_resolution_row(
            kypermno="4001",
            first_seen_caldt=date(2024, 2, 6),
            last_seen_caldt=date(2024, 2, 10),
            cusip="111111119",
            isin="US1111111199",
            ticker="MISS",
        ),
        _extended_resolution_row(
            kypermno="4002",
            first_seen_caldt=date(2024, 2, 1),
            last_seen_caldt=date(2024, 2, 5),
            cusip=None,
            isin=None,
            ticker="TKRO",
            TICKER_returned_ric="TICK.ONLY",
            TICKER_returned_isin="USTICKER4002",
            TICKER_returned_cusip="TICK4002",
        ),
    ]
    workbook_path = _write_filled_extended_lookup_workbook(
        tmp_path / "refinitiv_ric_lookup_handoff_common_stock_extended_filled_in.xlsx",
        rows,
    )

    out = run_refinitiv_step1_resolution_pipeline(workbook_path, tmp_path)

    assert set(out) == {
        "refinitiv_ric_resolution_common_stock_parquet",
    }
    for path in out.values():
        assert path.exists()
    assert not (tmp_path / "refinitiv_ric_resolution_common_stock.csv").exists()
    assert not (tmp_path / "refinitiv_ric_resolution_common_stock_summary.json").exists()
    assert not (tmp_path / "refinitiv_ric_resolution_common_stock_manifest.json").exists()

    resolved_df = pl.read_parquet(out["refinitiv_ric_resolution_common_stock_parquet"])
    assert resolved_df.columns == list(RIC_LOOKUP_RESOLUTION_OUTPUT_COLUMNS)
    assert resolved_df.height == 3

    by_id = {row["bridge_row_id"]: row for row in resolved_df.to_dicts()}
    assert by_id[_bridge_row_id("4001", cusip="111111119", isin="US1111111199", ticker="RSLV")][
        "effective_collection_ric"
    ] == "RESV.N"
    assert by_id[_bridge_row_id("4001", cusip="111111119", isin="US1111111199", ticker="MISS")][
        "effective_collection_ric_source"
    ] == "EXTENDED_FROM_PRIOR_ISIN"
    assert by_id[_bridge_row_id("4002", cusip=None, isin=None, ticker="TKRO")][
        "effective_collection_ric"
    ] is None


def test_run_refinitiv_step1_ownership_universe_handoff_pipeline_writes_formula_workbook(
    tmp_path: Path,
) -> None:
    resolution_df = build_refinitiv_step1_resolution_frame(
        _resolution_input_df(
            [
                _extended_resolution_row(
                    kypermno="9010",
                    cusip="111111111",
                    isin="US1111111111",
                    ticker="EFF",
                    first_seen_caldt=date(2024, 1, 1),
                    last_seen_caldt=date(2024, 1, 31),
                    ISIN_returned_ric="EFF.N",
                    ISIN_returned_isin="US1111111111",
                    ISIN_returned_cusip="111111111",
                ),
                _extended_resolution_row(
                    kypermno="9011",
                    cusip=None,
                    isin=None,
                    ticker="TKR",
                    first_seen_caldt=date(2024, 2, 1),
                    last_seen_caldt=date(2024, 2, 29),
                    TICKER_returned_ric="TKR.O",
                    TICKER_returned_isin="US4444444444",
                    TICKER_returned_cusip="444444444",
                ),
            ]
        )
    )
    resolution_path = tmp_path / "refinitiv_ric_resolution_common_stock.parquet"
    resolution_df.write_parquet(resolution_path)

    out = run_refinitiv_step1_ownership_universe_handoff_pipeline(
        resolution_artifact_path=resolution_path,
        output_dir=tmp_path / "ownership_universe_common_stock",
    )

    assert set(out) == {
        "refinitiv_ownership_universe_handoff_common_stock_parquet",
        "refinitiv_ownership_universe_handoff_common_stock_xlsx",
    }
    for path in out.values():
        assert path.exists()
    assert not (tmp_path / "ownership_universe_common_stock" / "refinitiv_ownership_universe_handoff_common_stock.csv").exists()
    assert not (
        tmp_path / "ownership_universe_common_stock" / "refinitiv_ownership_universe_handoff_common_stock_summary.json"
    ).exists()
    assert not (
        tmp_path / "ownership_universe_common_stock" / "refinitiv_ownership_universe_handoff_common_stock_manifest.json"
    ).exists()

    handoff_df = pl.read_parquet(out["refinitiv_ownership_universe_handoff_common_stock_parquet"])
    assert handoff_df.columns == list(OWNERSHIP_UNIVERSE_HANDOFF_COLUMNS)
    assert handoff_df.filter(pl.col("retrieval_eligible")).height == 2

    workbook = load_workbook(out["refinitiv_ownership_universe_handoff_common_stock_xlsx"], data_only=False)
    try:
        assert workbook.sheetnames == ["ownership_retrieval", "README"]
        retrieval_ws = workbook["ownership_retrieval"]
        assert retrieval_ws["A1"].value == "input_data"
        assert retrieval_ws["B1"].value == "returned_ric"
        assert retrieval_ws["F1"].value == "input_data"
        assert retrieval_ws["G1"].value == "returned_ric"
        assert retrieval_ws["A2"].value == "9010:01:111111111:US1111111111:EFF"
        assert retrieval_ws["A3"].value == "EFF.N"
        assert retrieval_ws["A6"].value == "UNIVERSE_EFFECTIVE"
        assert retrieval_ws["B2"].value == (
            '=@RDP.Data(A3,'
            '"TR.CategoryOwnershipPct.Date;TR.CategoryOwnershipPct;TR.InstrStatTypeValue",'
            '"StatType=7 SDate="&A4&" EDate="&A5&" CH=Fd RH=IN")'
        )
        assert retrieval_ws["F2"].value == "9011:01:-:-:TKR"
        assert retrieval_ws["F3"].value == "TKR.O"
        assert retrieval_ws["G2"].value == (
            '=@RDP.Data(F3,'
            '"TR.CategoryOwnershipPct.Date;TR.CategoryOwnershipPct;TR.InstrStatTypeValue",'
            '"StatType=7 SDate="&F4&" EDate="&F5&" CH=Fd RH=IN")'
        )
    finally:
        workbook.close()


def test_run_refinitiv_step1_ownership_universe_results_pipeline_builds_row_summary(
    tmp_path: Path,
) -> None:
    resolution_df = build_refinitiv_step1_resolution_frame(
        _resolution_input_df(
            [
                _extended_resolution_row(
                    kypermno="9020",
                    cusip="111111111",
                    isin="US1111111111",
                    ticker="EFF",
                    first_seen_caldt=date(2024, 1, 1),
                    last_seen_caldt=date(2024, 1, 31),
                    ISIN_returned_ric="EFF.N",
                    ISIN_returned_isin="US1111111111",
                    ISIN_returned_cusip="111111111",
                ),
                _extended_resolution_row(
                    kypermno="9021",
                    cusip="222222222",
                    isin="US2222222222",
                    ticker="CNF",
                    first_seen_caldt=date(2024, 2, 1),
                    last_seen_caldt=date(2024, 2, 29),
                    ISIN_returned_ric="CNFA.O",
                    ISIN_returned_isin="US2222222222",
                    ISIN_returned_cusip="222222222",
                    CUSIP_returned_ric="CNFB.N",
                    CUSIP_returned_isin="US3333333333",
                    CUSIP_returned_cusip="333333333",
                ),
                _extended_resolution_row(
                    kypermno="9022",
                    cusip=None,
                    isin=None,
                    ticker="TKR",
                    first_seen_caldt=date(2024, 3, 1),
                    last_seen_caldt=date(2024, 3, 31),
                    TICKER_returned_ric="TKR.O",
                    TICKER_returned_isin="US4444444444",
                    TICKER_returned_cusip="444444444",
                ),
                _extended_resolution_row(
                    kypermno="9023",
                    cusip=None,
                    isin=None,
                    ticker="MISS",
                    first_seen_caldt=date(2024, 4, 1),
                    last_seen_caldt=date(2024, 4, 30),
                ),
            ]
        )
    )
    resolution_path = tmp_path / "refinitiv_ric_resolution_common_stock.parquet"
    resolution_df.write_parquet(resolution_path)
    handoff_paths = run_refinitiv_step1_ownership_universe_handoff_pipeline(
        resolution_artifact_path=resolution_path,
        output_dir=tmp_path / "ownership_universe_common_stock",
    )
    handoff_df = pl.read_parquet(handoff_paths["refinitiv_ownership_universe_handoff_common_stock_parquet"])

    returned_rows_by_lookup_row_id = {
        "9020:01:111111111:US1111111111:EFF|UNIVERSE_EFFECTIVE": [
            ("EFF.N", date(2024, 1, 31), 10, "CAT_A"),
            ("EFF.N", date(2024, 2, 29), 11.5, "CAT_B"),
        ],
        "9021:01:222222222:US2222222222:CNF|UNIVERSE_TARGET_ISIN_CANDIDATE": [
            ("CNFA.O", date(2024, 2, 29), 21, "CAT_A"),
        ],
        "9021:01:222222222:US2222222222:CNF|UNIVERSE_TARGET_CUSIP_CANDIDATE": [
            ("CNFB.N", date(2024, 2, 29), 22.0, "CAT_A"),
        ],
        "9022:01:-:-:TKR|UNIVERSE_TARGET_TICKER_CANDIDATE": [
            ("TKR.O", date(2024, 3, 31), 30, "CAT_C"),
        ],
    }
    filled_workbook_path = (
        tmp_path
        / "ownership_universe_common_stock"
        / "refinitiv_ownership_universe_handoff_common_stock_filled_in.xlsx"
    )
    _write_filled_ownership_universe_workbook(
        filled_workbook_path,
        handoff_df,
        returned_rows_by_lookup_row_id,
    )

    out = run_refinitiv_step1_ownership_universe_results_pipeline(
        filled_workbook_path=filled_workbook_path,
        output_dir=tmp_path / "ownership_universe_common_stock",
    )

    assert set(out) == {
        "refinitiv_ownership_universe_results_parquet",
        "refinitiv_ownership_universe_row_summary_parquet",
    }
    for path in out.values():
        assert path.exists()
    assert not (tmp_path / "ownership_universe_common_stock" / "refinitiv_ownership_universe_results.csv").exists()
    assert not (
        tmp_path / "ownership_universe_common_stock" / "refinitiv_ownership_universe_row_summary.csv"
    ).exists()
    assert not (
        tmp_path / "ownership_universe_common_stock" / "refinitiv_ownership_universe_row_summary.json"
    ).exists()
    assert not (
        tmp_path / "ownership_universe_common_stock" / "refinitiv_ownership_universe_results_manifest.json"
    ).exists()

    results_df = pl.read_parquet(out["refinitiv_ownership_universe_results_parquet"])
    row_summary_df = pl.read_parquet(out["refinitiv_ownership_universe_row_summary_parquet"])

    assert results_df.columns == list(OWNERSHIP_UNIVERSE_RESULTS_COLUMNS)
    assert row_summary_df.columns == list(OWNERSHIP_UNIVERSE_ROW_SUMMARY_COLUMNS)
    assert results_df.height == 5
    assert row_summary_df.height == 5

    summary_by_lookup_id = {row["ownership_lookup_row_id"]: row for row in row_summary_df.to_dicts()}
    assert summary_by_lookup_id["9020:01:111111111:US1111111111:EFF|UNIVERSE_EFFECTIVE"]["ownership_rows_returned"] == 2
    assert summary_by_lookup_id["9020:01:111111111:US1111111111:EFF|UNIVERSE_EFFECTIVE"]["ownership_returned_ric_nunique"] == 1
    assert summary_by_lookup_id["9023:01:-:-:MISS|UNIVERSE_NOT_RETRIEVABLE"]["retrieval_row_present"] is False
    assert summary_by_lookup_id["9023:01:-:-:MISS|UNIVERSE_NOT_RETRIEVABLE"]["ownership_rows_returned"] == 0


def test_run_refinitiv_step1_bridge_pipeline_writes_only_canonical_artifacts(
    tmp_path: Path,
) -> None:
    out = run_refinitiv_step1_bridge_pipeline(
        _daily_panel()
        .vstack(
            pl.DataFrame(
                {
                    "KYPERMNO": [3],
                    "CALDT": [date(2024, 1, 6)],
                    "KYGVKEY_final": ["3000"],
                    "LIID": [None],
                    "CIK_final": ["0000003000"],
                    "CUSIP": [None],
                    "ISIN": [None],
                    "TICKER": [None],
                    "LINKTYPE": [None],
                    "LINKPRIM": [None],
                    "link_quality_flag": [None],
                    "HEXCNTRY": [None],
                    "n_filings": [0],
                    "SHRCD": [10],
                    "EXCHCD": [1],
                }
            )
        )
        .lazy(),
        tmp_path,
    )

    assert set(out) == {
        "refinitiv_bridge_universe_parquet",
        "refinitiv_ric_lookup_handoff_common_stock_extended_xlsx",
        "refinitiv_step1_manifest",
    }
    for path in out.values():
        assert path.exists()
    assert not (tmp_path / "refinitiv_bridge_universe.csv").exists()
    assert not (tmp_path / "refinitiv_bridge_handoff.xlsx").exists()
    assert not (tmp_path / "refinitiv_ric_lookup_handoff.xlsx").exists()
    assert not (tmp_path / "refinitiv_ric_lookup_handoff_common_stock.xlsx").exists()
    assert not (tmp_path / "refinitiv_ric_lookup_handoff_common_stock_with_gvkey.xlsx").exists()

    parquet_df = pl.read_parquet(out["refinitiv_bridge_universe_parquet"])
    assert parquet_df.columns == list(BRIDGE_OUTPUT_COLUMNS)
    assert parquet_df.height == 5

    manifest = json.loads(out["refinitiv_step1_manifest"].read_text(encoding="utf-8"))
    assert manifest["bridge_rows"] == 5
    assert manifest["ric_lookup_rows"] == 3
    assert manifest["ric_manual_review_rows"] == 1
    assert set(manifest["artifacts"]) == {
        "refinitiv_bridge_universe_parquet",
        "refinitiv_ric_lookup_handoff_common_stock_extended_xlsx",
        "refinitiv_step1_manifest",
    }

    with zipfile.ZipFile(out["refinitiv_ric_lookup_handoff_common_stock_extended_xlsx"]) as workbook_zip:
        workbook_xml = workbook_zip.read("xl/workbook.xml").decode("utf-8")
        shared_strings = workbook_zip.read("xl/sharedStrings.xml").decode("utf-8")

    assert "lookup_diagnostics" in workbook_xml
    assert "summary" in workbook_xml
    assert "README" in workbook_xml
    assert "3:-:-:-:-" in shared_strings
