from __future__ import annotations

from datetime import date
from pathlib import Path

import polars as pl
from openpyxl import Workbook, load_workbook

from thesis_pkg.pipeline import (
    run_refinitiv_lm2011_doc_ownership_exact_handoff_pipeline,
    run_refinitiv_lm2011_doc_ownership_fallback_handoff_pipeline,
    run_refinitiv_lm2011_doc_ownership_finalize_pipeline,
)
from thesis_pkg.pipelines.refinitiv import lseg_ownership_api
from thesis_pkg.pipelines.refinitiv.doc_ownership import (
    DOC_OWNERSHIP_BLOCK_HEADERS,
    DOC_OWNERSHIP_INPUT_FIELDS,
    DOC_OWNERSHIP_EXACT_STAGE,
    DOC_OWNERSHIP_FALLBACK_STAGE,
    _clean_institutional_value,
    _most_recent_quarter_end_before,
    _target_effective_date_for_quarter_end,
    build_refinitiv_lm2011_doc_ownership_requests,
)


def _build_doc_ownership_inputs(tmp_path: Path) -> tuple[Path, Path, Path]:
    doc_filing_path = tmp_path / "sec_ccm_matched_clean_filtered.parquet"
    authority_decisions_path = tmp_path / "refinitiv_permno_ownership_authority_decisions.parquet"
    authority_exceptions_path = tmp_path / "refinitiv_permno_ownership_authority_exceptions.parquet"

    pl.DataFrame(
        {
            "doc_id": [
                "doc_exact",
                "doc_fallback",
                "doc_exception",
                "doc_review",
                "doc_no_authority",
                "doc_conflict",
            ],
            "filing_date": [
                date(2024, 4, 15),
                date(2024, 5, 20),
                date(2024, 8, 20),
                date(2024, 4, 10),
                date(2024, 4, 10),
                date(2024, 5, 25),
            ],
            "kypermno": ["100", "100", "200", "300", "400", "100"],
        }
    ).write_parquet(doc_filing_path)

    pl.DataFrame(
        {
            "KYPERMNO": ["100", "200", "300", "400"],
            "authoritative_ric": ["AAA.N", None, None, None],
            "authoritative_source_family": ["CONVENTIONAL", "CONVENTIONAL", None, None],
            "authority_decision_status": [
                "STATIC_CONVENTIONAL",
                "DATE_VARYING_CONVENTIONAL_EXCEPTION",
                "REVIEW_REQUIRED",
                "NO_CONVENTIONAL_AUTHORITY",
            ],
            "requires_review": [False, False, True, False],
        }
    ).write_parquet(authority_decisions_path)

    pl.DataFrame(
        {
            "KYPERMNO": ["200", "200"],
            "authoritative_ric": ["OLD.N", "NEW.N"],
            "authoritative_source_family": ["CONVENTIONAL", "CONVENTIONAL"],
            "authority_window_start_date": [date(2024, 1, 1), date(2024, 6, 1)],
            "authority_window_end_date": [date(2024, 3, 31), date(2024, 12, 31)],
            "authority_exception_status": [
                "DATE_VARYING_CONVENTIONAL_EXCEPTION",
                "DATE_VARYING_CONVENTIONAL_EXCEPTION",
            ],
        }
    ).write_parquet(authority_exceptions_path)
    return doc_filing_path, authority_decisions_path, authority_exceptions_path


def _write_filled_doc_ownership_workbook(
    path: Path,
    request_df: pl.DataFrame,
    *,
    request_stage: str,
    exact_rows_by_doc_id: dict[str, list[tuple[date | None, float | None, str | None]]] | None = None,
    fallback_rows_by_doc_id: dict[str, list[tuple[date | None, float | None, str | None]]] | None = None,
) -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "ownership_retrieval"

    eligible_df = request_df.filter(pl.col("retrieval_eligible").fill_null(False))
    for block_index, row in enumerate(eligible_df.iter_rows(named=True)):
        base_col = 1 + (block_index * len(DOC_OWNERSHIP_BLOCK_HEADERS))
        for offset, header in enumerate(DOC_OWNERSHIP_BLOCK_HEADERS):
            worksheet.cell(row=1, column=base_col + offset).value = header
        for field_offset, field_name in enumerate(DOC_OWNERSHIP_INPUT_FIELDS, start=1):
            worksheet.cell(row=1 + field_offset, column=base_col).value = row.get(field_name)

        doc_id = str(row["doc_id"])
        worksheet.cell(row=2, column=base_col + 1).value = row.get("authoritative_ric")
        if request_stage == DOC_OWNERSHIP_EXACT_STAGE:
            worksheet.cell(row=2, column=base_col + 2).value = "Date"
            worksheet.cell(row=2, column=base_col + 3).value = "Category Percent Of Tr"
            worksheet.cell(row=2, column=base_col + 4).value = "Investor Statistics Category Value"
            for result_index, (returned_date, returned_value, returned_category) in enumerate(
                (exact_rows_by_doc_id or {}).get(doc_id, []),
                start=1,
            ):
                excel_row = 2 + result_index
                worksheet.cell(row=excel_row, column=base_col + 2).value = returned_date
                worksheet.cell(row=excel_row, column=base_col + 3).value = returned_value
                worksheet.cell(row=excel_row, column=base_col + 4).value = returned_category
        elif request_stage == DOC_OWNERSHIP_FALLBACK_STAGE:
            worksheet.cell(row=2, column=base_col + 2).value = "Date"
            worksheet.cell(row=2, column=base_col + 3).value = "Category Percent Of Tr"
            worksheet.cell(row=2, column=base_col + 4).value = "Investor Statistics Category Value"
            for result_index, (returned_date, returned_value, returned_category) in enumerate(
                (fallback_rows_by_doc_id or {}).get(doc_id, []),
                start=1,
            ):
                excel_row = 2 + result_index
                worksheet.cell(row=excel_row, column=base_col + 2).value = returned_date
                worksheet.cell(row=excel_row, column=base_col + 3).value = returned_value
                worksheet.cell(row=excel_row, column=base_col + 4).value = returned_category
        else:
            raise ValueError(f"unsupported request stage: {request_stage}")

    workbook.save(path)
    workbook.close()
    return path


def test_most_recent_quarter_end_before_handles_boundaries() -> None:
    assert _most_recent_quarter_end_before(date(2024, 1, 1)) == date(2023, 12, 31)
    assert _most_recent_quarter_end_before(date(2024, 3, 31)) == date(2023, 12, 31)
    assert _most_recent_quarter_end_before(date(2024, 4, 1)) == date(2024, 3, 31)


def test_target_effective_date_for_quarter_end_handles_boundaries() -> None:
    assert _target_effective_date_for_quarter_end(date(2023, 12, 31)) == date(2024, 1, 1)
    assert _target_effective_date_for_quarter_end(date(2024, 3, 31)) == date(2024, 4, 1)
    assert _target_effective_date_for_quarter_end(date(2024, 6, 30)) == date(2024, 7, 1)
    assert _target_effective_date_for_quarter_end(date(2024, 9, 30)) == date(2024, 10, 1)


def test_build_refinitiv_lm2011_doc_ownership_requests_handles_static_exception_and_ineligible(
    tmp_path: Path,
) -> None:
    doc_filing_path, authority_decisions_path, authority_exceptions_path = _build_doc_ownership_inputs(tmp_path)
    request_df = build_refinitiv_lm2011_doc_ownership_requests(
        pl.read_parquet(doc_filing_path),
        pl.read_parquet(authority_decisions_path),
        pl.read_parquet(authority_exceptions_path),
    )
    rows = {row["doc_id"]: row for row in request_df.to_dicts()}

    assert rows["doc_exact"]["retrieval_eligible"] is True
    assert rows["doc_exact"]["authoritative_ric"] == "AAA.N"
    assert rows["doc_exact"]["target_quarter_end"] == date(2024, 3, 31)
    assert rows["doc_exact"]["target_effective_date"] == date(2024, 4, 1)
    assert rows["doc_exact"]["fallback_window_start"] == date(2024, 4, 1)
    assert rows["doc_exact"]["fallback_window_end"] == date(2024, 4, 15)

    assert rows["doc_exception"]["retrieval_eligible"] is True
    assert rows["doc_exception"]["authoritative_ric"] == "NEW.N"
    assert rows["doc_exception"]["target_quarter_end"] == date(2024, 6, 30)
    assert rows["doc_exception"]["target_effective_date"] == date(2024, 7, 1)

    assert rows["doc_review"]["retrieval_eligible"] is False
    assert rows["doc_review"]["retrieval_exclusion_reason"] == "authority_review_required"

    assert rows["doc_no_authority"]["retrieval_eligible"] is False
    assert rows["doc_no_authority"]["retrieval_exclusion_reason"] == "no_authoritative_ric"


def test_build_refinitiv_lm2011_doc_ownership_requests_applies_request_bounds(tmp_path: Path) -> None:
    doc_filing_path, authority_decisions_path, authority_exceptions_path = _build_doc_ownership_inputs(tmp_path)
    request_df = build_refinitiv_lm2011_doc_ownership_requests(
        pl.read_parquet(doc_filing_path),
        pl.read_parquet(authority_decisions_path),
        pl.read_parquet(authority_exceptions_path),
        request_min_date=date(2024, 4, 1),
        request_max_date=date(2024, 12, 31),
    )
    rows = {row["doc_id"]: row for row in request_df.to_dicts()}

    assert rows["doc_exact"]["retrieval_eligible"] is True
    assert rows["doc_exact"]["target_effective_date"] == date(2024, 4, 1)
    assert rows["doc_exact"]["fallback_window_start"] == date(2024, 4, 1)
    assert rows["doc_exact"]["fallback_window_end"] == date(2024, 4, 15)

    assert rows["doc_exception"]["retrieval_eligible"] is True
    assert rows["doc_exception"]["target_effective_date"] == date(2024, 7, 1)
    assert rows["doc_exception"]["fallback_window_start"] == date(2024, 7, 1)
    assert rows["doc_exception"]["fallback_window_end"] == date(2024, 8, 15)


def test_build_refinitiv_lm2011_doc_ownership_requests_marks_pre_min_docs_ineligible(tmp_path: Path) -> None:
    doc_filing_path, authority_decisions_path, authority_exceptions_path = _build_doc_ownership_inputs(tmp_path)
    request_df = build_refinitiv_lm2011_doc_ownership_requests(
        pl.read_parquet(doc_filing_path),
        pl.read_parquet(authority_decisions_path),
        pl.read_parquet(authority_exceptions_path),
        request_min_date=date(2024, 5, 1),
        request_max_date=date(2024, 12, 31),
    )
    rows = {row["doc_id"]: row for row in request_df.to_dicts()}

    assert rows["doc_exact"]["retrieval_eligible"] is False
    assert rows["doc_exact"]["retrieval_exclusion_reason"] == "target_effective_date_before_request_min_date"
    assert rows["doc_fallback"]["retrieval_eligible"] is False
    assert rows["doc_fallback"]["retrieval_exclusion_reason"] == "target_effective_date_before_request_min_date"


def test_build_refinitiv_lm2011_doc_ownership_requests_marks_post_max_docs_ineligible(tmp_path: Path) -> None:
    doc_filing_path, authority_decisions_path, authority_exceptions_path = _build_doc_ownership_inputs(tmp_path)
    late_doc_df = pl.read_parquet(doc_filing_path).vstack(
        pl.DataFrame({"doc_id": ["doc_post_max"], "filing_date": [date(2025, 2, 15)], "kypermno": ["100"]})
    )
    request_df = build_refinitiv_lm2011_doc_ownership_requests(
        late_doc_df,
        pl.read_parquet(authority_decisions_path),
        pl.read_parquet(authority_exceptions_path),
        request_min_date=date(2024, 1, 1),
        request_max_date=date(2024, 12, 31),
    )
    rows = {row["doc_id"]: row for row in request_df.to_dicts()}

    assert rows["doc_post_max"]["retrieval_eligible"] is False
    assert rows["doc_post_max"]["retrieval_exclusion_reason"] == "target_effective_date_after_request_max_date"


def test_doc_ownership_exact_handoff_pipeline_threads_request_bounds(tmp_path: Path) -> None:
    doc_filing_path, authority_decisions_path, authority_exceptions_path = _build_doc_ownership_inputs(tmp_path)
    output_dir = tmp_path / "refinitiv_doc_ownership_lm2011"

    exact_paths = run_refinitiv_lm2011_doc_ownership_exact_handoff_pipeline(
        doc_filing_artifact_path=doc_filing_path,
        authority_decisions_artifact_path=authority_decisions_path,
        authority_exceptions_artifact_path=authority_exceptions_path,
        output_dir=output_dir,
        request_min_date=date(2024, 5, 1),
        request_max_date=date(2024, 12, 31),
    )
    exact_request_df = pl.read_parquet(exact_paths["refinitiv_lm2011_doc_ownership_exact_requests_parquet"])
    rows = {row["doc_id"]: row for row in exact_request_df.to_dicts()}

    assert rows["doc_exact"]["retrieval_eligible"] is False
    assert rows["doc_exact"]["retrieval_exclusion_reason"] == "target_effective_date_before_request_min_date"
    assert rows["doc_exception"]["retrieval_eligible"] is True


def test_doc_ownership_exact_api_pipeline_threads_request_bounds(tmp_path: Path, monkeypatch) -> None:
    doc_filing_path, authority_decisions_path, authority_exceptions_path = _build_doc_ownership_inputs(tmp_path)
    output_dir = tmp_path / "refinitiv_doc_ownership_lm2011"
    captured: dict[str, object] = {}

    def fake_build_requests(
        doc_filing_df: pl.DataFrame,
        authority_decisions_df: pl.DataFrame,
        authority_exceptions_df: pl.DataFrame,
        *,
        request_min_date=None,
        request_max_date=None,
    ) -> pl.DataFrame:
        captured["request_min_date"] = request_min_date
        captured["request_max_date"] = request_max_date
        return build_refinitiv_lm2011_doc_ownership_requests(
            doc_filing_df,
            authority_decisions_df,
            authority_exceptions_df,
            request_min_date=request_min_date,
            request_max_date=request_max_date,
        )

    class _StageRun:
        def __init__(self, staging_dir: Path) -> None:
            self.staging_dir = staging_dir
            self.run_session_id = "test-session"

    class _AuditResult:
        passed = True

        @staticmethod
        def to_dict() -> dict[str, object]:
            return {}

    monkeypatch.setattr(lseg_ownership_api, "build_refinitiv_lm2011_doc_ownership_requests", fake_build_requests)
    monkeypatch.setattr(lseg_ownership_api, "run_api_batches", lambda **kwargs: _StageRun(output_dir / "staging"))
    monkeypatch.setattr(
        lseg_ownership_api,
        "_assemble_doc_raw",
        lambda staging_dir, **kwargs: pl.DataFrame(schema=lseg_ownership_api._doc_ownership_raw_schema()).select(
            lseg_ownership_api.DOC_OWNERSHIP_RAW_COLUMNS
        ),
    )
    monkeypatch.setattr(lseg_ownership_api, "audit_api_stage", lambda **kwargs: _AuditResult())
    monkeypatch.setattr(lseg_ownership_api, "write_stage_completion_manifest", lambda **kwargs: None)

    lseg_ownership_api.run_refinitiv_lm2011_doc_ownership_exact_api_pipeline(
        doc_filing_artifact_path=doc_filing_path,
        authority_decisions_artifact_path=authority_decisions_path,
        authority_exceptions_artifact_path=authority_exceptions_path,
        output_dir=output_dir,
        request_min_date=date(2024, 5, 1),
        request_max_date=date(2024, 12, 31),
    )

    assert captured["request_min_date"] == date(2024, 5, 1)
    assert captured["request_max_date"] == date(2024, 12, 31)


def test_doc_ownership_workbooks_write_expected_exact_and_fallback_formulas(tmp_path: Path) -> None:
    doc_filing_path, authority_decisions_path, authority_exceptions_path = _build_doc_ownership_inputs(tmp_path)
    output_dir = tmp_path / "refinitiv_doc_ownership_lm2011"

    exact_paths = run_refinitiv_lm2011_doc_ownership_exact_handoff_pipeline(
        doc_filing_artifact_path=doc_filing_path,
        authority_decisions_artifact_path=authority_decisions_path,
        authority_exceptions_artifact_path=authority_exceptions_path,
        output_dir=output_dir,
    )
    workbook = load_workbook(exact_paths["refinitiv_lm2011_doc_ownership_exact_handoff_xlsx"], data_only=False)
    try:
        worksheet = workbook["ownership_retrieval"]
        assert [worksheet.cell(row=1, column=idx).value for idx in range(1, 6)] == list(DOC_OWNERSHIP_BLOCK_HEADERS)
        assert worksheet["C2"].value is not None
        assert "TR.CategoryOwnershipPct.Date;TR.CategoryOwnershipPct;TR.InstrStatTypeValue" in str(worksheet["C2"].value)
        assert 'TEXT(A5,"yyyy-mm-dd")' in str(worksheet["C2"].value)
    finally:
        workbook.close()

    exact_request_df = pl.read_parquet(exact_paths["refinitiv_lm2011_doc_ownership_exact_requests_parquet"])
    exact_filled_path = output_dir / "refinitiv_lm2011_doc_ownership_exact_handoff_filled_in.xlsx"
    _write_filled_doc_ownership_workbook(
        exact_filled_path,
        exact_request_df,
        request_stage=DOC_OWNERSHIP_EXACT_STAGE,
        exact_rows_by_doc_id={
            "doc_exact": [(date(2024, 4, 1), 55.0, "Holdings by Institutions")],
        },
    )
    fallback_paths = run_refinitiv_lm2011_doc_ownership_fallback_handoff_pipeline(
        exact_filled_workbook_path=exact_filled_path,
        output_dir=output_dir,
    )
    fallback_workbook = load_workbook(
        fallback_paths["refinitiv_lm2011_doc_ownership_fallback_handoff_xlsx"],
        data_only=False,
    )
    try:
        fallback_sheet = fallback_workbook["ownership_retrieval"]
        assert "TR.CategoryOwnershipPct.Date;TR.CategoryOwnershipPct;TR.InstrStatTypeValue" in str(
            fallback_sheet["C2"].value
        )
        assert 'TEXT(A6,"yyyy-mm-dd")' in str(fallback_sheet["C2"].value)
        assert 'TEXT(A7,"yyyy-mm-dd")' in str(fallback_sheet["C2"].value)
    finally:
        fallback_workbook.close()


def test_clean_institutional_value_handles_negative_and_cap() -> None:
    assert _clean_institutional_value(-1.0) is None
    assert _clean_institutional_value(125.0) == 100.0


def test_doc_ownership_finalize_pipeline_builds_one_row_per_doc_id(tmp_path: Path) -> None:
    doc_filing_path, authority_decisions_path, authority_exceptions_path = _build_doc_ownership_inputs(tmp_path)
    output_dir = tmp_path / "refinitiv_doc_ownership_lm2011"

    exact_paths = run_refinitiv_lm2011_doc_ownership_exact_handoff_pipeline(
        doc_filing_artifact_path=doc_filing_path,
        authority_decisions_artifact_path=authority_decisions_path,
        authority_exceptions_artifact_path=authority_exceptions_path,
        output_dir=output_dir,
    )
    exact_request_df = pl.read_parquet(exact_paths["refinitiv_lm2011_doc_ownership_exact_requests_parquet"])
    exact_filled_path = output_dir / "refinitiv_lm2011_doc_ownership_exact_handoff_filled_in.xlsx"
    _write_filled_doc_ownership_workbook(
        exact_filled_path,
        exact_request_df,
        request_stage=DOC_OWNERSHIP_EXACT_STAGE,
        exact_rows_by_doc_id={
            "doc_exact": [
                (date(2024, 4, 1), 105.0, "Holdings by Institutions"),
                (date(2024, 4, 1), 20.0, "Holdings by Domestic Investors"),
            ],
            "doc_fallback": [(date(2024, 4, 1), 21.0, "Holdings by Domestic Investors")],
            "doc_exception": [(date(2024, 7, 1), 120.0, "Holdings by Institutions")],
            "doc_conflict": [(date(2024, 4, 1), 18.0, "Holdings by Domestic Investors")],
        },
    )
    fallback_paths = run_refinitiv_lm2011_doc_ownership_fallback_handoff_pipeline(
        exact_filled_workbook_path=exact_filled_path,
        output_dir=output_dir,
    )
    fallback_request_df = pl.read_parquet(fallback_paths["refinitiv_lm2011_doc_ownership_fallback_requests_parquet"])
    assert set(fallback_request_df.get_column("doc_id").to_list()) == {"doc_fallback", "doc_conflict"}

    fallback_filled_path = output_dir / "refinitiv_lm2011_doc_ownership_fallback_handoff_filled_in.xlsx"
    _write_filled_doc_ownership_workbook(
        fallback_filled_path,
        fallback_request_df,
        request_stage=DOC_OWNERSHIP_FALLBACK_STAGE,
        fallback_rows_by_doc_id={
            "doc_fallback": [
                (date(2024, 3, 31), 43.0, "Holdings by Institutions"),
                (date(2024, 4, 30), 44.0, "Holdings by Institutions"),
            ],
            "doc_conflict": [
                (date(2024, 4, 30), 40.0, "Holdings by Institutions"),
                (date(2024, 4, 30), 41.0, "Holdings by Institutions"),
            ],
        },
    )
    finalize_paths = run_refinitiv_lm2011_doc_ownership_finalize_pipeline(
        output_dir=output_dir,
        fallback_filled_workbook_path=fallback_filled_path,
    )

    raw_df = pl.read_parquet(finalize_paths["refinitiv_lm2011_doc_ownership_raw_parquet"])
    final_df = pl.read_parquet(finalize_paths["refinitiv_lm2011_doc_ownership_parquet"])
    assert final_df.height == 6
    assert final_df.select(pl.col("doc_id").n_unique()).item() == 6
    assert set(raw_df.get_column("doc_id").to_list()) == {"doc_exact", "doc_fallback", "doc_exception", "doc_conflict"}

    rows = {row["doc_id"]: row for row in final_df.to_dicts()}
    assert rows["doc_exact"]["retrieval_status"] == "EXACT_TARGET_HIT"
    assert rows["doc_exact"]["target_effective_date"] == date(2024, 4, 1)
    assert rows["doc_exact"]["institutional_ownership_pct"] == 100.0
    assert rows["doc_exact"]["fallback_used"] is False
    assert rows["doc_exact"]["selected_response_date"] == date(2024, 4, 1)

    assert rows["doc_fallback"]["retrieval_status"] == "FALLBACK_WINDOW_HIT"
    assert rows["doc_fallback"]["institutional_ownership_pct"] == 44.0
    assert rows["doc_fallback"]["fallback_used"] is True
    assert rows["doc_fallback"]["selected_response_date"] == date(2024, 4, 30)

    assert rows["doc_exception"]["retrieval_status"] == "EXACT_TARGET_HIT"
    assert rows["doc_exception"]["authoritative_ric"] == "NEW.N"
    assert rows["doc_exception"]["institutional_ownership_pct"] == 100.0
    assert rows["doc_exception"]["selected_response_date"] == date(2024, 7, 1)

    assert rows["doc_review"]["retrieval_status"] == "AUTHORITY_REVIEW_REQUIRED"
    assert rows["doc_review"]["institutional_ownership_pct"] is None

    assert rows["doc_no_authority"]["retrieval_status"] == "NO_AUTHORITATIVE_RIC"
    assert rows["doc_no_authority"]["institutional_ownership_pct"] is None

    assert rows["doc_conflict"]["retrieval_status"] == "FALLBACK_CONFLICT_REVIEW"
    assert rows["doc_conflict"]["institutional_ownership_pct"] is None
