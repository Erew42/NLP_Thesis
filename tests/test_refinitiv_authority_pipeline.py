from __future__ import annotations

from datetime import date
from pathlib import Path

import polars as pl
from openpyxl import Workbook

from thesis_pkg.pipelines.refinitiv_bridge_pipeline import (
    RIC_LOOKUP_EXTENDED_COLUMNS,
    build_refinitiv_step1_ownership_authority_tables,
    build_refinitiv_step1_resolution_frame,
    run_refinitiv_step1_ownership_authority_pipeline,
    run_refinitiv_step1_ownership_universe_handoff_pipeline,
    run_refinitiv_step1_ownership_universe_results_pipeline,
)


def _bridge_row_id(
    kypermno: str,
    *,
    liid: str | None = "01",
    cusip: str | None = None,
    isin: str | None = None,
    ticker: str | None = None,
) -> str:
    return f"{kypermno}:{liid or '-'}:{cusip or '-'}:{isin or '-'}:{ticker or '-'}"


def _extended_resolution_row(
    *,
    kypermno: str,
    first_seen_caldt: date,
    last_seen_caldt: date,
    liid: str | None = "01",
    cusip: str | None = None,
    isin: str | None = None,
    ticker: str | None = None,
    **overrides: object,
) -> dict[str, object]:
    row: dict[str, object] = {name: None for name in RIC_LOOKUP_EXTENDED_COLUMNS}
    row.update(
        {
            "bridge_row_id": _bridge_row_id(
                kypermno,
                liid=liid,
                cusip=cusip,
                isin=isin,
                ticker=ticker,
            ),
            "KYPERMNO": kypermno,
            "CUSIP": cusip,
            "ISIN": isin,
            "TICKER": ticker,
            "first_seen_caldt": first_seen_caldt,
            "last_seen_caldt": last_seen_caldt,
            "ISIN_lookup_input": isin,
            "CUSIP_lookup_input": cusip,
            "TICKER_lookup_input": ticker,
        }
    )
    row.update(overrides)
    return row


def _resolution_input_df(rows: list[dict[str, object]]) -> pl.DataFrame:
    return pl.DataFrame(rows).select(
        [
            pl.col(name).alias(name)
            if name not in {"first_seen_caldt", "last_seen_caldt"}
            else pl.col(name).cast(pl.Date, strict=False).alias(name)
            for name in RIC_LOOKUP_EXTENDED_COLUMNS
        ]
    )


def _write_filled_ownership_universe_workbook(
    path: Path,
    handoff_df: pl.DataFrame,
    returned_rows_by_lookup_row_id: dict[str, list[tuple[str, object, object, str]]],
) -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "ownership_retrieval"
    block_headers = (
        "input_data",
        "returned_ric",
        "returned_date",
        "returned_value",
        "returned_category",
    )
    input_fields = (
        "bridge_row_id",
        "candidate_ric",
        "request_start_date",
        "request_end_date",
        "candidate_slot",
        "lookup_input_source",
        "effective_collection_ric",
        "effective_collection_ric_source",
        "accepted_ric",
        "accepted_ric_source",
    )
    eligible_df = handoff_df.filter(pl.col("retrieval_eligible").fill_null(False))
    for block_index, row in enumerate(eligible_df.iter_rows(named=True)):
        base_col = 1 + (block_index * 5)
        for offset, header in enumerate(block_headers):
            worksheet.cell(row=1, column=base_col + offset).value = header
        for field_offset, field_name in enumerate(input_fields, start=1):
            worksheet.cell(row=1 + field_offset, column=base_col).value = row.get(field_name)
        for result_index, (returned_ric, returned_date, returned_value, returned_category) in enumerate(
            returned_rows_by_lookup_row_id.get(str(row["ownership_lookup_row_id"]), []),
            start=1,
        ):
            excel_row = 2 + result_index
            worksheet.cell(row=excel_row, column=base_col + 1).value = returned_ric
            worksheet.cell(row=excel_row, column=base_col + 2).value = returned_date
            worksheet.cell(row=excel_row, column=base_col + 3).value = returned_value
            worksheet.cell(row=excel_row, column=base_col + 4).value = returned_category
    workbook.save(path)
    workbook.close()
    return path


def _build_authority_inputs(tmp_path: Path) -> tuple[Path, Path, Path]:
    resolution_df = build_refinitiv_step1_resolution_frame(
        _resolution_input_df(
            [
                _extended_resolution_row(
                    kypermno="100",
                    cusip="111111111",
                    isin="US1111111111",
                    ticker="ALIAS1",
                    first_seen_caldt=date(2024, 1, 1),
                    last_seen_caldt=date(2024, 3, 31),
                    ISIN_returned_ric="AAA.N",
                    ISIN_returned_isin="US1111111111",
                    ISIN_returned_cusip="111111111",
                ),
                _extended_resolution_row(
                    kypermno="100",
                    cusip="111111111",
                    isin="US1111111111",
                    ticker="ALIAS2",
                    first_seen_caldt=date(2024, 2, 1),
                    last_seen_caldt=date(2024, 4, 30),
                    ISIN_returned_ric="AAA.OQ",
                    ISIN_returned_isin="US1111111111",
                    ISIN_returned_cusip="111111111",
                ),
                _extended_resolution_row(
                    kypermno="200",
                    cusip="222222222",
                    isin="US2222222222",
                    ticker="OLD",
                    first_seen_caldt=date(2024, 1, 1),
                    last_seen_caldt=date(2024, 2, 29),
                    ISIN_returned_ric="OLD.N",
                    ISIN_returned_isin="US2222222222",
                    ISIN_returned_cusip="222222222",
                ),
                _extended_resolution_row(
                    kypermno="200",
                    cusip="333333333",
                    isin="US3333333333",
                    ticker="NEW",
                    first_seen_caldt=date(2024, 6, 1),
                    last_seen_caldt=date(2024, 7, 31),
                    ISIN_returned_ric="NEW.N",
                    ISIN_returned_isin="US3333333333",
                    ISIN_returned_cusip="333333333",
                ),
                _extended_resolution_row(
                    kypermno="300",
                    cusip=None,
                    isin=None,
                    ticker="TKR",
                    first_seen_caldt=date(2024, 5, 1),
                    last_seen_caldt=date(2024, 5, 31),
                    TICKER_returned_ric="TKR.O",
                    TICKER_returned_isin="US9999999999",
                    TICKER_returned_cusip="999999999",
                ),
                _extended_resolution_row(
                    kypermno="400",
                    cusip="444444444",
                    isin="US4444444444",
                    ticker="CNF1",
                    first_seen_caldt=date(2024, 1, 1),
                    last_seen_caldt=date(2024, 1, 31),
                    ISIN_returned_ric="CNF1.N",
                    ISIN_returned_isin="US4444444444",
                    ISIN_returned_cusip="444444444",
                ),
                _extended_resolution_row(
                    kypermno="400",
                    cusip="555555555",
                    isin="US5555555555",
                    ticker="CNF2",
                    first_seen_caldt=date(2024, 1, 1),
                    last_seen_caldt=date(2024, 1, 31),
                    ISIN_returned_ric="CNF2.N",
                    ISIN_returned_isin="US5555555555",
                    ISIN_returned_cusip="555555555",
                ),
            ]
        )
    )
    resolution_path = tmp_path / "refinitiv_ric_resolution_common_stock.parquet"
    resolution_df.write_parquet(resolution_path)

    handoff_paths = run_refinitiv_step1_ownership_universe_handoff_pipeline(
        resolution_artifact_path=resolution_path,
        output_dir=tmp_path / "ownership_universe_common_stock",
    )
    handoff_df = pl.read_parquet(handoff_paths["refinitiv_ownership_universe_handoff_common_stock_parquet"])
    returned_rows_by_lookup_row_id = {
        f"{_bridge_row_id('100', cusip='111111111', isin='US1111111111', ticker='ALIAS1')}|UNIVERSE_EFFECTIVE": [
            ("AAA.N", date(2024, 2, 1), 10.0, "CAT_A"),
            ("AAA.N", date(2024, 2, 1), 20.0, "CAT_B"),
            ("AAA.N", date(2024, 2, 1), 30.0, "CAT_C"),
            ("AAA.N", date(2024, 2, 29), 11.0, "CAT_A"),
            ("AAA.N", date(2024, 2, 29), 21.0, "CAT_B"),
            ("AAA.N", date(2024, 2, 29), 31.0, "CAT_C"),
            ("AAA.N", date(2024, 3, 31), 12.0, "CAT_A"),
            ("AAA.N", date(2024, 3, 31), 22.0, "CAT_B"),
            ("AAA.N", date(2024, 3, 31), 32.0, "CAT_C"),
            ("AAA.N", date(2024, 4, 30), 13.0, "CAT_A"),
            ("AAA.N", date(2024, 4, 30), 23.0, "CAT_B"),
            ("AAA.N", date(2024, 4, 30), 33.0, "CAT_C"),
        ],
        f"{_bridge_row_id('100', cusip='111111111', isin='US1111111111', ticker='ALIAS2')}|UNIVERSE_EFFECTIVE": [
            ("AAA.OQ", date(2024, 2, 1), 10.0, "CAT_A"),
            ("AAA.OQ", date(2024, 2, 1), 20.0, "CAT_B"),
            ("AAA.OQ", date(2024, 2, 1), 30.0, "CAT_C"),
            ("AAA.OQ", date(2024, 2, 29), 11.0, "CAT_A"),
            ("AAA.OQ", date(2024, 2, 29), 21.0, "CAT_B"),
            ("AAA.OQ", date(2024, 2, 29), 31.0, "CAT_C"),
            ("AAA.OQ", date(2024, 3, 31), 12.0, "CAT_A"),
            ("AAA.OQ", date(2024, 3, 31), 22.0, "CAT_B"),
            ("AAA.OQ", date(2024, 3, 31), 32.0, "CAT_C"),
            ("AAA.OQ", date(2024, 4, 30), 13.0, "CAT_A"),
            ("AAA.OQ", date(2024, 4, 30), 23.0, "CAT_B"),
            ("AAA.OQ", date(2024, 4, 30), 33.0, "CAT_C"),
        ],
        f"{_bridge_row_id('200', cusip='222222222', isin='US2222222222', ticker='OLD')}|UNIVERSE_EFFECTIVE": [
            ("OLD.N", date(2024, 1, 31), 40.0, "CAT_A"),
            ("OLD.N", date(2024, 2, 29), 41.0, "CAT_A"),
        ],
        f"{_bridge_row_id('200', cusip='333333333', isin='US3333333333', ticker='NEW')}|UNIVERSE_EFFECTIVE": [
            ("NEW.N", date(2024, 6, 30), 50.0, "CAT_A"),
            ("NEW.N", date(2024, 7, 31), 51.0, "CAT_A"),
        ],
        f"{_bridge_row_id('300', cusip=None, isin=None, ticker='TKR')}|UNIVERSE_TARGET_TICKER_CANDIDATE": [
            ("TKR.O", date(2024, 5, 31), 60.0, "CAT_A"),
        ],
        f"{_bridge_row_id('400', cusip='444444444', isin='US4444444444', ticker='CNF1')}|UNIVERSE_EFFECTIVE": [
            ("CNF1.N", date(2024, 1, 31), 70.0, "CAT_A"),
        ],
        f"{_bridge_row_id('400', cusip='555555555', isin='US5555555555', ticker='CNF2')}|UNIVERSE_EFFECTIVE": [
            ("CNF2.N", date(2024, 1, 31), 71.0, "CAT_A"),
        ],
    }
    filled_workbook_path = (
        tmp_path
        / "ownership_universe_common_stock"
        / "refinitiv_ownership_universe_handoff_common_stock_filled_in.xlsx"
    )
    _write_filled_ownership_universe_workbook(
        filled_workbook_path,
        handoff_df,
        returned_rows_by_lookup_row_id,
    )
    results_paths = run_refinitiv_step1_ownership_universe_results_pipeline(
        filled_workbook_path=filled_workbook_path,
        output_dir=tmp_path / "ownership_universe_common_stock",
    )
    return (
        resolution_path,
        results_paths["refinitiv_ownership_universe_results_parquet"],
        results_paths["refinitiv_ownership_universe_row_summary_parquet"],
    )


def test_build_refinitiv_step1_ownership_authority_tables_handles_static_exception_and_review(
    tmp_path: Path,
) -> None:
    resolution_path, results_path, row_summary_path = _build_authority_inputs(tmp_path)
    tables, summary = build_refinitiv_step1_ownership_authority_tables(
        pl.read_parquet(resolution_path),
        pl.read_parquet(results_path),
        pl.read_parquet(row_summary_path),
    )

    decisions = {
        row["KYPERMNO"]: row
        for row in tables["authority_decisions"].to_dicts()
    }
    assert decisions["100"]["authority_decision_status"] == "STATIC_CONVENTIONAL"
    assert decisions["100"]["authoritative_ric"] == "AAA.N"
    assert decisions["200"]["authority_decision_status"] == "DATE_VARYING_CONVENTIONAL_EXCEPTION"
    assert decisions["300"]["requires_review"] is True
    assert decisions["300"]["review_flag_ticker_only_without_allowlist"] is True
    assert decisions["400"]["requires_review"] is True
    assert decisions["400"]["review_flag_effective_overlap_conflict"] is True

    alias_rows = tables["alias_diagnostics"].filter(pl.col("KYPERMNO") == "100").to_dicts()
    assert len(alias_rows) == 1
    assert alias_rows[0]["pair_benign_alias_supported"] is True

    exception_rows = tables["authority_exceptions"].filter(pl.col("KYPERMNO") == "200")
    assert exception_rows.height == 2
    assert set(exception_rows.get_column("authoritative_ric").to_list()) == {"OLD.N", "NEW.N"}

    final_panel_permnos = set(tables["final_panel"].get_column("KYPERMNO").to_list())
    assert final_panel_permnos == {"100", "200"}
    assert summary["static_conventional_permno_count"] == 1
    assert summary["date_varying_conventional_exception_permno_count"] == 1
    assert summary["review_required_permno_count"] == 2


def test_run_refinitiv_step1_ownership_authority_pipeline_applies_reviewed_ticker_allowlist(
    tmp_path: Path,
) -> None:
    resolution_path, results_path, row_summary_path = _build_authority_inputs(tmp_path)
    allowlist_path = tmp_path / "ownership_authority_common_stock" / "refinitiv_permno_ownership_ticker_allowlist.parquet"
    allowlist_path.parent.mkdir(parents=True, exist_ok=True)
    pl.DataFrame(
        {
            "KYPERMNO": ["300"],
            "candidate_ric": ["TKR.O"],
            "allowlist_reason": ["manual_review"],
        }
    ).write_parquet(allowlist_path)

    out = run_refinitiv_step1_ownership_authority_pipeline(
        resolution_artifact_path=resolution_path,
        ownership_results_artifact_path=results_path,
        ownership_row_summary_artifact_path=row_summary_path,
        output_dir=tmp_path / "ownership_authority_common_stock",
        reviewed_ticker_allowlist_path=allowlist_path,
    )

    assert set(out) == {
        "refinitiv_permno_ownership_candidate_metrics_parquet",
        "refinitiv_permno_ownership_alias_diagnostics_parquet",
        "refinitiv_permno_ownership_authority_decisions_parquet",
        "refinitiv_permno_ownership_authority_exceptions_parquet",
        "refinitiv_permno_ownership_review_required_parquet",
        "refinitiv_permno_ownership_ticker_candidates_parquet",
        "refinitiv_permno_date_ownership_panel_parquet",
    }
    for path in out.values():
        assert path.exists()

    decisions_df = pl.read_parquet(out["refinitiv_permno_ownership_authority_decisions_parquet"])
    final_panel_df = pl.read_parquet(out["refinitiv_permno_date_ownership_panel_parquet"])
    ticker_row = decisions_df.filter(pl.col("KYPERMNO") == "300").row(0, named=True)
    assert ticker_row["authority_decision_status"] == "REVIEWED_TICKER_ALLOWLIST_ONLY"
    assert ticker_row["reviewed_ticker_allowlist_applied"] is True
    assert "300" in set(final_panel_df.get_column("KYPERMNO").to_list())
